#!/usr/bin/env python3
"""
PayU Product Documentation Coverage Tracker generator.

Source of truth: PayU Developer Documentation repository (/workspace).
Sample Excel used only for reporting philosophy/format inspiration.
"""

from __future__ import annotations

from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

ROOT = Path("/workspace")
OUT_DIR = Path("/workspace/docs-coverage-tracker")
OUT_FILE = OUT_DIR / "PayU_Product_Documentation_Coverage_Tracker.xlsx"

# Coverage dimensions (equal weight among applicable)
DIMENSIONS = [
    "overview",
    "integration_guide",
    "api_reference",
    "sdk",
    "quick_start",
    "webhooks",
    "error_codes",
    "testing",
    "go_live",
    "troubleshooting",
    "faqs",
    "changelog",
]


@dataclass
class Product:
    name: str
    category: str
    product_type: str
    overview: Optional[str] = None
    integration_guide: Optional[str] = None
    api_reference: Optional[str] = None
    sdk: Optional[str] = None
    quick_start: Optional[str] = None
    webhooks: Optional[str] = None  # path or "Yes"/"Shared" marker via exists
    error_codes: Optional[str] = None
    testing: Optional[str] = None
    go_live: Optional[str] = None
    troubleshooting: Optional[str] = None
    faqs: Optional[str] = None
    changelog: Optional[str] = None
    # Applicability: True=required, False=N/A, None=auto from presence
    applicable: dict = field(default_factory=dict)
    # Integration guide recommendation
    recommend_ig: bool = False
    priority: str = "P3"  # P0/P1/P2/P3 / N/A
    recommended_action: str = ""
    notes: str = ""
    # Extra pages for inventory notes
    existing_pages_note: str = ""

    def path_exists(self, path: Optional[str]) -> bool:
        if not path:
            return False
        if path in {"Yes", "Shared", "Partial", "N/A"}:
            return path in {"Yes", "Shared", "Partial"}
        p = ROOT / path
        return p.exists()

    def flag(self, dim: str) -> str:
        """Return Yes / No / Partial / N/A for a dimension."""
        if self.applicable.get(dim) is False:
            return "N/A"
        val = getattr(self, dim)
        if val == "Partial":
            return "Partial"
        if val == "Shared":
            return "Yes"
        if val == "Yes":
            return "Yes"
        if val == "N/A":
            return "N/A"
        if val and self.path_exists(val):
            return "Yes"
        # If marked applicable False already handled; if no path, Missing
        if self.applicable.get(dim) is False:
            return "N/A"
        return "No"

    def link(self, dim: str) -> str:
        val = getattr(self, dim)
        if not val or val in {"Yes", "No", "Partial", "Shared", "N/A"}:
            return ""
        if self.path_exists(val):
            return val
        return ""

    def coverage_score(self) -> float:
        scores = []
        for dim in DIMENSIONS:
            f = self.flag(dim)
            if f == "N/A":
                continue
            if f == "Yes":
                scores.append(1.0)
            elif f == "Partial":
                scores.append(0.5)
            else:
                scores.append(0.0)
        if not scores:
            return 0.0
        return round(100.0 * sum(scores) / len(scores), 1)

    def status(self) -> str:
        score = self.coverage_score()
        # Missing critical overview + integration for developer-facing products
        overview = self.flag("overview")
        ig = self.flag("integration_guide")
        if overview == "No" and ig == "No" and score < 40:
            return "Missing"
        if score >= 85:
            return "Complete"
        if score < 40:
            return "Missing"
        return "Partial"


def exists(path: str) -> bool:
    return (ROOT / path).exists()


def build_products() -> list[Product]:
    """Build product inventory from repository structure (verified paths)."""
    products: list[Product] = []

    # Shared web checkout assets
    WEB_WEBHOOKS = "docs/Collect Payments/introduction-web/webhooks.md"
    WEB_ERRORS = "docs/Collect Payments/introduction-web/error-handling.md"
    WEB_TESTING = "docs/Collect Payments/introduction-web/test-cards-upi-id-and-wallets.md"
    WEB_FAQS = "docs/Collect Payments/introduction-web/faqs-for-web-checkout-integration.md"
    WEB_ERROR_REF = "reference/Collect Payment/error-codes.md"
    COLLECT_API = "reference/Collect Payment"

    # -------------------------------------------------------------------------
    # Getting Started / Platform
    # -------------------------------------------------------------------------
    products.append(
        Product(
            name="Merchant Onboarding / Sign Up",
            category="Getting Started",
            product_type="Platform / Onboarding",
            overview="docs/getting started/introduction/index.md",
            integration_guide="docs/getting started/register-with-payu/register-for-a-merchant-account-on-dashboard.md",
            api_reference=None,
            sdk=None,
            quick_start="docs/getting started/register-with-payu/index.md"
            if exists("docs/getting started/register-with-payu/index.md")
            else "docs/getting started/register-with-payu/register-for-a-merchant-account-on-dashboard.md",
            webhooks=None,
            error_codes=None,
            testing="docs/getting started/payu-dashboard/generate-test-merchant-key-and-salt.md",
            go_live="docs/getting started/register-with-payu/complete-your-kyc.md",
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={
                "api_reference": False,
                "sdk": False,
                "webhooks": False,
                "error_codes": False,
                "changelog": False,
                "troubleshooting": True,
                "faqs": True,
            },
            recommend_ig=False,
            priority="P2",
            recommended_action="Add FAQs and troubleshooting for KYC/activation blockers; keep onboarding journey as step-by-step guide (not a classic Integration Guide).",
            notes="Covers register, KYC, documents checklist under getting started/register-with-payu.",
        )
    )

    products.append(
        Product(
            name="PayU Dashboard",
            category="Getting Started",
            product_type="Platform / Tool",
            overview="docs/getting started/payu-dashboard/index.md",
            integration_guide=None,
            api_reference=None,
            sdk=None,
            quick_start="docs/getting started/payu-dashboard/log-in-to-dashboard.md",
            webhooks="docs/getting started/payu-dashboard/manage-webhooks-using-dashboard/index.md",
            error_codes=None,
            testing="docs/getting started/payu-dashboard/generate-test-merchant-key-and-salt.md",
            go_live=None,
            troubleshooting=None,
            faqs="docs/getting started/payu-dashboard/faqs-for-dashboard.md",
            changelog=None,
            applicable={
                "integration_guide": False,
                "api_reference": False,
                "sdk": False,
                "error_codes": False,
                "go_live": False,
                "changelog": False,
            },
            recommend_ig=False,
            priority="P2",
            recommended_action="Improve module-level how-to consistency (settlements, reports, users); add troubleshooting for common dashboard access issues.",
            notes="Modules include transactions, settlements, beneficiaries, integrations, banking, reports, webhooks, users/permissions.",
        )
    )

    # -------------------------------------------------------------------------
    # Collect Payments — Core Checkout
    # -------------------------------------------------------------------------
    products.append(
        Product(
            name="PayU Hosted Checkout (Prebuilt)",
            category="Collect Payments / Web Checkout",
            product_type="Core Payment Product",
            overview="docs/Collect Payments/introduction-web/prebuilt-checkout-payu-hosted/index.md",
            integration_guide="docs/Collect Payments/introduction-web/prebuilt-checkout-payu-hosted/prebuilt-checkout-page-integration.md",
            api_reference="reference/Collect Payment/_payment_payu_hosted_checkout.md",
            sdk=None,
            quick_start="docs/Collect Payments/introduction-web/prebuilt-checkout-payu-hosted/prebuilt-checkout-page-integration.md",
            webhooks=WEB_WEBHOOKS,
            error_codes=WEB_ERROR_REF,
            testing=WEB_TESTING,
            go_live="Partial",
            troubleshooting="Partial",
            faqs=WEB_FAQS,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P0",
            recommended_action="Strengthen dedicated end-to-end Integration Guide with go-live checklist, troubleshooting, and clearer path from hash → request → callback → verify payment.",
            notes="Also called Prebuilt / Non-Seamless Checkout. Shared webhooks/errors/testing under introduction-web. Recipe exists in recipes/payu-hosted-checkout-curl-request-walkthrough.md. ASK AI duplicate: hosted-checkout-integration-demo.",
        )
    )

    products.append(
        Product(
            name="Merchant Hosted Checkout (Custom / Seamless)",
            category="Collect Payments / Web Checkout",
            product_type="Core Payment Product",
            overview="docs/Collect Payments/introduction-web/custom-checkout-merchant-hosted/index.md",
            integration_guide="docs/Collect Payments/introduction-web/custom-checkout-merchant-hosted/integration-checklist-merchant-hosted-checkout.md",
            api_reference="reference/Collect Payment/_payment_merchant_hosted",
            sdk=None,
            quick_start="docs/Collect Payments/introduction-web/custom-checkout-merchant-hosted/collect-payments-with-cards-seamless.md",
            webhooks=WEB_WEBHOOKS,
            error_codes="reference/Collect Payment/_payment_merchant_hosted/_payment_merchant_hosted_cards/field-7-field-8-error-codes.md",
            testing="docs/Collect Payments/introduction-web/custom-checkout-merchant-hosted/test-integration.md",
            go_live="docs/Collect Payments/introduction-web/custom-checkout-merchant-hosted/integration-checklist-merchant-hosted-checkout.md",
            troubleshooting="Partial",
            faqs=WEB_FAQS,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P0",
            recommended_action="Publish a single merchant-journey Integration Guide that stitches payment-method pages (cards/NB/UPI/wallets/EMI/BNPL), hashing, webhooks, verify payment, and go-live into one narrative.",
            notes="Payment method pages exist individually. Duplicate API surfaces: reference/PayU Merchant Hosted — _payment and General APIs/merchanthostedpostservice.",
        )
    )

    products.append(
        Product(
            name="Server-to-Server (S2S) Checkout",
            category="Collect Payments / Web Checkout",
            product_type="Core Payment Product",
            overview="docs/Collect Payments/introduction-web/server-to-server-integration/index.md",
            integration_guide="docs/Collect Payments/introduction-web/server-to-server-integration/integrate-with-s2s.md",
            api_reference="reference/Collect Payment/_payment_server_to_server",
            sdk=None,
            quick_start="docs/Collect Payments/introduction-web/server-to-server-integration/integrate-with-s2s.md",
            webhooks=WEB_WEBHOOKS,
            error_codes="reference/Collect Payment/_payment_server_to_server/error-codes-for-s2s-link-and-pay.md",
            testing=WEB_TESTING,
            go_live="docs/Collect Payments/introduction-web/server-to-server-integration/integration-checklist-s2s.md",
            troubleshooting="Partial",
            faqs=WEB_FAQS,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P0",
            recommended_action="Create a unified S2S Integration Guide covering classic, decoupled, and direct-authorization flows with decision tree, error handling, and go-live checklist.",
            notes="High complexity (classic/decoupled/direct auth/UPI variants). Legacy and old_ API pages exist in reference. Recycle Bin contains decoupled-flow docs.",
        )
    )

    products.append(
        Product(
            name="Checkout Plus (ICP / Bolt Checkout)",
            category="Collect Payments / Web Checkout",
            product_type="Core Payment Product",
            overview="docs/Collect Payments/introduction-web/checkout-plus-integration/index.md",
            integration_guide="docs/Collect Payments/introduction-web/checkout-plus-integration/integrate-checkout-plus.md",
            api_reference="docs/Collect Payments/introduction-web/checkout-plus-integration/apis-used-for-checkout-plus-integration.md",
            sdk=None,
            quick_start="docs/Collect Payments/introduction-web/checkout-plus-integration/integrate-checkout-plus.md",
            webhooks=WEB_WEBHOOKS,
            error_codes=WEB_ERROR_REF,
            testing=WEB_TESTING,
            go_live=None,
            troubleshooting=None,
            faqs=WEB_FAQS,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P0",
            recommended_action="Consolidate Checkout Plus / ICP / Bolt naming; expand Integration Guide with testing, go-live, troubleshooting; reduce reliance on ASK AI duplicate page.",
            notes="Naming inconsistency: Checkout Plus vs ICP Checkout vs Bolt Checkout. ASK AI page: Integration ASK AI Docs/checkout-plusicp-checkoutbolt-checkout.md.",
        )
    )

    products.append(
        Product(
            name="CommercePro Checkout (Checkout Express)",
            category="Collect Payments / Web Checkout",
            product_type="Core Payment Product",
            overview="docs/Collect Payments/introduction-web/checkout-express/index.md",
            integration_guide="docs/Collect Payments/introduction-web/checkout-express/integrate-commercepro-checkout-using-callback-url.md",
            api_reference="reference/Checkout Express",
            sdk=None,
            quick_start="docs/Collect Payments/introduction-web/checkout-express/integration-checkout-express-response-handler.md",
            webhooks=WEB_WEBHOOKS,
            error_codes=WEB_ERROR_REF,
            testing=WEB_TESTING,
            go_live=None,
            troubleshooting=None,
            faqs=WEB_FAQS,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Unify CommercePro vs Checkout Express naming; add dedicated go-live + troubleshooting; clarify relationship to plugin CommercePro pages.",
            notes="Folder/reference named Checkout Express; guide titles use CommercePro. Also listed under ecommerce-platform-plugins/commercepro-checkout.md.",
        )
    )

    # -------------------------------------------------------------------------
    # No-code
    # -------------------------------------------------------------------------
    products.append(
        Product(
            name="Payment Links",
            category="Collect Payments / No-Code",
            product_type="Core Payment Product",
            overview="docs/Collect Payments/introduction-no-code-payments-integration/payment-links-dashboard/index.md",
            integration_guide="docs/Collect Payments/introduction-no-code-payments-integration/payment-links-dashboard/integration-api-for-payment-links.md",
            api_reference="reference/payment links",
            sdk=None,
            quick_start="docs/Collect Payments/introduction-no-code-payments-integration/payment-links-dashboard/index.md",
            webhooks=WEB_WEBHOOKS,
            error_codes=None,
            testing=WEB_TESTING,
            go_live=None,
            troubleshooting=None,
            faqs="docs/Collect Payments/introduction-no-code-payments-integration/payment-links-dashboard/faqs-payment-links.md",
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P0",
            recommended_action="Ship a dedicated Payment Links Integration Guide covering dashboard + API creation, webhooks, TPV variant, WhatsApp EPL, and go-live.",
            notes="Also appears in partner APIs, WhatsApp Enhanced Payment Links, TPV payment-link guide, WooCommerce payment links.",
        )
    )

    products.append(
        Product(
            name="Payment Buttons",
            category="Collect Payments / No-Code",
            product_type="Payment Feature",
            overview="docs/Collect Payments/introduction-no-code-payments-integration/payment-buttons-dashboard.md",
            integration_guide="docs/Collect Payments/introduction-no-code-payments-integration/payment-buttons-dashboard.md",
            api_reference=None,
            sdk=None,
            quick_start="docs/Collect Payments/introduction-no-code-payments-integration/payment-buttons-dashboard.md",
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={
                "api_reference": False,
                "sdk": False,
                "webhooks": False,
                "error_codes": False,
                "changelog": False,
            },
            recommend_ig=False,
            priority="P3",
            recommended_action="Expand dashboard how-to with screenshots and FAQs; dedicated Integration Guide not required (no-code UI product).",
            notes="Single-page documentation under no-code payments.",
        )
    )

    products.append(
        Product(
            name="Invoices (No-Code)",
            category="Collect Payments / No-Code",
            product_type="Payment Feature",
            overview="docs/Collect Payments/introduction-no-code-payments-integration/invoices-dashboard/index.md",
            integration_guide="docs/Collect Payments/introduction-no-code-payments-integration/invoices-dashboard/create-an-invoice.md",
            api_reference=None,
            sdk=None,
            quick_start="docs/Collect Payments/introduction-no-code-payments-integration/invoices-dashboard/create-an-invoice.md",
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={
                "api_reference": False,
                "sdk": False,
                "webhooks": False,
                "error_codes": False,
                "changelog": False,
            },
            recommend_ig=False,
            priority="P2",
            recommended_action="Clarify Invoice vs Payment Link vs Zion subscription invoices; add FAQs.",
            notes="Partner APIs also mention Manage Invoices or Payment Links — naming overlap risk.",
        )
    )

    # -------------------------------------------------------------------------
    # In-person
    # -------------------------------------------------------------------------
    products.append(
        Product(
            name="UPI QR",
            category="Collect Payments / In-Person",
            product_type="Core Payment Product",
            overview="docs/Collect Payments/in-person-payments/integrate-upi-qr/index.md",
            integration_guide="docs/Collect Payments/in-person-payments/integrate-upi-qr/index.md",
            api_reference="reference/In-person payments/integrate-upi-qr-apis",
            sdk=None,
            quick_start="docs/Collect Payments/in-person-payments/integrate-upi-qr/index.md",
            webhooks="Partial",
            error_codes="reference/In-person payments/integrate-upi-qr-apis/error-codes-for-qr-apis-1.md",
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs="docs/Collect Payments/in-person-payments/integrate-upi-qr/faqs-2.md",
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Expand Integration Guide with callback/webhook handling, sandbox testing, and go-live steps.",
            notes="API reference includes dynamic/static QR and error codes.",
        )
    )

    products.append(
        Product(
            name="Dynamic Storefront QR (DBQR)",
            category="Collect Payments / In-Person",
            product_type="Payment Feature",
            overview="docs/Collect Payments/in-person-payments/integrated-dynamic-storefront/index.md",
            integration_guide="docs/Collect Payments/in-person-payments/integrated-dynamic-storefront/integrated-dynamic-storefront-customer-journey.md",
            api_reference=None,
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "api_reference": True, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Publish a dedicated Integration Guide; resolve DBQR / Offline DBQR / Dynamic Bharat QR naming; surface ASK AI offline-dbqr content into canonical docs.",
            notes="ASK AI duplicate: Integration ASK AI Docs/offline-dbqr.md. Limited API mapping in guides.",
        )
    )

    products.append(
        Product(
            name="POS Terminal Integration",
            category="Collect Payments / In-Person",
            product_type="Core Payment Product",
            overview="docs/Collect Payments/in-person-payments/pos-terminal-integration/index.md",
            integration_guide="docs/Collect Payments/in-person-payments/pos-terminal-integration/apis-for-pos-terminal-integration.md",
            api_reference="reference/In-person payments/pos-terminal-integration-apis",
            sdk=None,
            quick_start="docs/Collect Payments/in-person-payments/pos-terminal-integration/apis-for-pos-terminal-integration.md",
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Add end-to-end Integration Guide with device setup, test transactions, error codes, and go-live.",
            notes="Overview + API map exist; operational docs thin.",
        )
    )

    products.append(
        Product(
            name="Android POS SDK",
            category="Collect Payments / In-Person",
            product_type="Integration Channel (SDK)",
            overview="docs/Collect Payments/in-person-payments/android-pos-sdk/index.md",
            integration_guide="docs/Collect Payments/in-person-payments/android-pos-sdk/index.md",
            api_reference="reference/In-person payments/android-pos-sdk-apis",
            sdk="docs/Collect Payments/in-person-payments/android-pos-sdk/index.md",
            quick_start="docs/Collect Payments/in-person-payments/android-pos-sdk/index.md",
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"webhooks": False, "changelog": True},
            recommend_ig=True,
            priority="P1",
            recommended_action="Expand SDK Integration Guide with sample app walkthrough, error handling, testing, and changelog.",
            notes="Installation/activation pages under android-pos-sdk folder.",
        )
    )

    # -------------------------------------------------------------------------
    # Ecommerce plugins
    # -------------------------------------------------------------------------
    plugins = [
        ("Shopify", "shopify", "index.md", "faqs-for-shopify.md", None, "P1", True,
         "High-adoption plugin; ensure CommercePro/offers/reconciliation paths are clear; keep Integration Guide as install→configure→go-live."),
        ("WooCommerce", "woocommerce", "index.md", None, "troubleshooting-woocommerce-integration.md", "P1", True,
         "High-adoption; reconcile duplicate ASK AI woocommerce guide into canonical docs."),
        ("Magento", "magento", "index.md", None, "troubleshooting-magento-integration.md", "P2", True,
         "Maintain install + CommercePro + troubleshooting; Integration Guide valuable for merchants."),
        ("Wix", "wix", "index.md", "faqs-for-wix-integration.md", None, "P2", False,
         "Overview + FAQ sufficient; expand only if support volume warrants."),
        ("BigCommerce", "bigcommerce", "index.md", None, "troubleshooting-bigcommerce-integration.md", "P2", False,
         "Keep troubleshooting current; dedicated IG low value beyond install guide."),
        ("Shopmatic", "shopmatic", "index.md", None, "troubleshooting-shopmatic-integration.md", "P3", False,
         "Maintain troubleshooting; no dedicated IG beyond install."),
        ("Fynd Store", "fynd-integration", "index.md", None, None, "P3", False,
         "Thin docs; expand only with merchant demand."),
        ("OpenCart", "opencart", "index.md", None, "troubleshooting-opencart-integration.md", "P3", False,
         "Install + troubleshooting adequate for plugin channel."),
        ("PrestaShop", "prestashop", "index.md", None, "troubleshooting-prestashop-integration.md", "P3", False,
         "Install + troubleshooting adequate."),
        ("Zoho", "zoho-integration", "index.md", None, None, "P2", True,
         "Multiple Zoho surfaces (Marketplace/One/Billing/Inventory); recommend one Integration Guide with product matrix."),
        ("Odoo", "odoo", "index.md", None, None, "P3", False,
         "Installation page present; expand if adoption grows."),
        ("Bagisto", "bagisto.md", None, None, None, "P3", False,
         "Single-page guide; low priority."),
        ("Interakt for WhatsApp Business", "interakt-for-whatsapp-business", "index.md", None, None, "P2", False,
         "Cross-links with WhatsApp payments; clarify channel boundary."),
    ]

    for name, folder, overview_file, faqs_file, trouble_file, pri, rec_ig, action in plugins:
        base = f"docs/Collect Payments/ecommerce-platform-plugins/{folder}"
        if folder.endswith(".md"):
            overview = f"docs/Collect Payments/ecommerce-platform-plugins/{folder}"
            ig = overview
            faqs = None
            trouble = None
        else:
            overview = f"{base}/{overview_file}" if overview_file else f"{base}/index.md"
            # find a likely integration page
            ig_candidates = [
                f"{base}/index.md",
            ]
            ig = next((c for c in ig_candidates if exists(c)), overview)
            faqs = f"{base}/{faqs_file}" if faqs_file else None
            trouble = f"{base}/{trouble_file}" if trouble_file else None
        products.append(
            Product(
                name=f"{name} Plugin",
                category="Collect Payments / eCommerce Plugins",
                product_type="Integration Channel (Plugin)",
                overview=overview if exists(overview) else None,
                integration_guide=ig if exists(ig) else None,
                api_reference=None,
                sdk=None,
                quick_start=ig if exists(ig) else None,
                webhooks=None,
                error_codes=None,
                testing=None,
                go_live="docs/Collect Payments/ecommerce-platform-plugins/integration-checklist-plugins.md",
                troubleshooting=trouble if trouble and exists(trouble) else None,
                faqs=faqs if faqs and exists(faqs) else None,
                changelog=None,
                applicable={
                    "api_reference": False,
                    "sdk": False,
                    "webhooks": False,
                    "error_codes": False,
                    "changelog": False,
                },
                recommend_ig=rec_ig,
                priority=pri,
                recommended_action=action,
                notes="Shared plugin checklist: ecommerce-platform-plugins/integration-checklist-plugins.md.",
            )
        )

    # -------------------------------------------------------------------------
    # Server SDKs (grouped)
    # -------------------------------------------------------------------------
    for lang, file in [
        ("Go SDK", "go-sdk.md"),
        ("Java SDK", "java-sdk.md"),
        ("PHP SDK", "php-sdk.md"),
        ("Python SDK", "python-sdk.md"),
        ("Node.js SDK", "node-js-sdk.md"),
    ]:
        path = f"docs/Collect Payments/explore-server-integrations/{file}"
        products.append(
            Product(
                name=lang,
                category="Collect Payments / Server-Side SDKs",
                product_type="Integration Channel (SDK)",
                overview="docs/Collect Payments/explore-server-integrations/index.md"
                if exists("docs/Collect Payments/explore-server-integrations/index.md")
                else path,
                integration_guide=path,
                api_reference=COLLECT_API,
                sdk=path,
                quick_start=path,
                webhooks=WEB_WEBHOOKS,
                error_codes=WEB_ERROR_REF,
                testing=WEB_TESTING,
                go_live=None,
                troubleshooting=None,
                faqs=None,
                changelog=None,
                applicable={"changelog": True, "troubleshooting": True, "faqs": True, "go_live": True},
                recommend_ig=True if lang in {"PHP SDK", "Node.js SDK", "Java SDK"} else False,
                priority="P1" if lang in {"PHP SDK", "Node.js SDK", "Java SDK"} else "P2",
                recommended_action="Upgrade from single-page SDK note to full Integration Guide (install, auth, payment sample, verify, errors, changelog) for high-usage languages.",
                notes="Currently thin single-page SDK docs under explore-server-integrations.",
            )
        )

    # -------------------------------------------------------------------------
    # Mobile SDKs — major products (platform-level)
    # -------------------------------------------------------------------------
    mobile = [
        (
            "Android CheckoutPro SDK",
            "docs/Collect Payments/mobile-sdks/explore-android-sdks/android-checkoutpro-sdk/index.md",
            "docs/Collect Payments/mobile-sdks/explore-android-sdks/android-checkoutpro-sdk/index.md",
            "docs/Collect Payments/mobile-sdks/explore-android-sdks/android-checkoutpro-sdk/android-checkoutpro-troubleshoot-errors.md",
            "docs/Collect Payments/mobile-sdks/explore-android-sdks/faqs-android-sdk/index.md",
            None,
            "P0",
            True,
            "Primary Android integration path; ensure Integration Guide covers offers/TPV/sample app and keep troubleshooting current.",
        ),
        (
            "Android Core SDK",
            "docs/Collect Payments/mobile-sdks/explore-android-sdks/android-core-sdk/index.md",
            "docs/Collect Payments/mobile-sdks/explore-android-sdks/android-core-sdk/index.md",
            None,
            "docs/Collect Payments/mobile-sdks/explore-android-sdks/faqs-android-sdk/index.md",
            None,
            "P1",
            True,
            "Document decision: CheckoutPro vs Core; Integration Guide should include TPV and web services.",
        ),
        (
            "iOS CheckoutPro SDK",
            "docs/Collect Payments/mobile-sdks/explore-ios-sdks/ios-checkoutpro-sdk/index.md",
            "docs/Collect Payments/mobile-sdks/explore-ios-sdks/ios-checkoutpro-sdk/index.md",
            None,
            "docs/Collect Payments/mobile-sdks/explore-ios-sdks/ios-sdk-faqs.md",
            None,
            "P0",
            True,
            "Primary iOS path; add explicit go-live checklist and changelog parity with Android.",
        ),
        (
            "iOS Core SDK",
            "docs/Collect Payments/mobile-sdks/explore-ios-sdks/ios-core-sdk/index.md",
            "docs/Collect Payments/mobile-sdks/explore-ios-sdks/ios-core-sdk/index.md",
            None,
            "docs/Collect Payments/mobile-sdks/explore-ios-sdks/ios-sdk-faqs.md",
            None,
            "P1",
            True,
            "Cover recurring/TPV; align release notes.",
        ),
        (
            "iOS Custom Browser SDK",
            "docs/Collect Payments/mobile-sdks/explore-ios-sdks/ios-custombrowser-sdk/index.md",
            "docs/Collect Payments/mobile-sdks/explore-ios-sdks/ios-custombrowser-sdk/index.md",
            None,
            "docs/Collect Payments/mobile-sdks/explore-ios-sdks/ios-sdk-faqs.md",
            "docs/Collect Payments/mobile-sdks/explore-ios-sdks/ios-custombrowser-sdk/ios-custombrowser-golive-checklist.md",
            "P1",
            False,
            "Has go-live checklist and test integration — model for other SDKs; dedicated IG optional if pages stay cohesive.",
        ),
        (
            "React Native CheckoutPro SDK",
            "docs/Collect Payments/mobile-sdks/explore-reactnative-sdks/react-native-checkoutpro-sdk/index.md",
            "docs/Collect Payments/mobile-sdks/explore-reactnative-sdks/react-native-checkoutpro-sdk/index.md",
            None,
            "docs/Collect Payments/mobile-sdks/explore-reactnative-sdks/faqs-react-native-sdk.md",
            None,
            "P1",
            True,
            "High DevEx impact for cross-platform apps; consolidate duplicate FAQ pages.",
        ),
        (
            "Flutter CheckoutPro SDK",
            "docs/Collect Payments/mobile-sdks/flutter-sdk-introduction/flutter-checkoutpro-sdk/index.md",
            "docs/Collect Payments/mobile-sdks/flutter-sdk-introduction/flutter-checkoutpro-sdk/index.md",
            None,
            None,
            None,
            "P1",
            True,
            "Add FAQs/troubleshooting; version history exists — promote as changelog.",
        ),
        (
            "Cordova CheckoutPro SDK",
            "docs/Collect Payments/mobile-sdks/cordova-mobile-sdks/cordova-sdk-introduction/index.md",
            "docs/Collect Payments/mobile-sdks/cordova-mobile-sdks/cordova-sdk-introduction/index.md",
            None,
            None,
            None,
            "P2",
            False,
            "Maintain; lower adoption vs RN/Flutter.",
        ),
        (
            "UPI Bolt SDK (Android/iOS/RN/Flutter/Ionic)",
            "docs/Collect Payments/mobile-sdks/explore-android-sdks/payu-bolt-sdk/index.md",
            "docs/Collect Payments/mobile-sdks/explore-android-sdks/payu-bolt-sdk/index.md",
            None,
            None,
            None,
            "P1",
            True,
            "Resolve Ionic/Capacitor/Cordova naming inconsistency; publish one cross-platform UPI Bolt Integration Guide with platform tabs.",
        ),
    ]

    for name, overview, ig, trouble, faqs, golive, pri, rec_ig, action in mobile:
        products.append(
            Product(
                name=name,
                category="Collect Payments / Mobile SDKs",
                product_type="Integration Channel (SDK)",
                overview=overview if exists(overview) else None,
                integration_guide=ig if exists(ig) else None,
                api_reference=None,
                sdk=overview if exists(overview) else None,
                quick_start=ig if exists(ig) else None,
                webhooks=None,
                error_codes=None,
                testing="docs/Collect Payments/mobile-sdks/explore-ios-sdks/ios-custombrowser-sdk/ios-custombrowser-test-integration.md"
                if "iOS Custom" in name
                else WEB_TESTING,
                go_live=golive if golive and exists(golive) else None,
                troubleshooting=trouble if trouble and exists(trouble) else None,
                faqs=faqs if faqs and exists(faqs) else None,
                changelog=(
                    "docs/Collect Payments/mobile-sdks/flutter-sdk-introduction/flutter-checkoutpro-sdk/flutter-checkout-pro-version-history.md"
                    if "Flutter CheckoutPro" in name
                    and exists(
                        "docs/Collect Payments/mobile-sdks/flutter-sdk-introduction/flutter-checkoutpro-sdk/flutter-checkout-pro-version-history.md"
                    )
                    else (
                        "docs/Collect Payments/mobile-sdks/explore-android-sdks/android-checkoutpro-sdk/change-logs.md"
                        if "Android CheckoutPro" in name
                        and exists(
                            "docs/Collect Payments/mobile-sdks/explore-android-sdks/android-checkoutpro-sdk/change-logs.md"
                        )
                        else (
                            "docs/Collect Payments/mobile-sdks/explore-reactnative-sdks/react-native-checkoutpro-sdk/reactnative-checkoutpro-change-logs.md"
                            if "React Native CheckoutPro" in name
                            and exists(
                                "docs/Collect Payments/mobile-sdks/explore-reactnative-sdks/react-native-checkoutpro-sdk/reactnative-checkoutpro-change-logs.md"
                            )
                            else None
                        )
                    )
                ),
                applicable={
                    "api_reference": False,
                    "webhooks": False,
                    "error_codes": True,
                },
                recommend_ig=rec_ig,
                priority=pri,
                recommended_action=action,
                notes="Sibling SDKs (UPI, Native OTP Assist, PhonePe, Google Pay, 3DS/FlashPay, Ola Money, Custom Browser) also exist under mobile-sdks.",
            )
        )

    # -------------------------------------------------------------------------
    # Offerings
    # -------------------------------------------------------------------------
    offerings = [
        Product(
            name="EMI / Cardless EMI",
            category="Offerings / Affordability",
            product_type="Payment Feature",
            overview="docs/Offerings/introduction-to-affordability/emi-api-integration/index.md",
            integration_guide="docs/Offerings/introduction-to-affordability/emi-api-integration/integration-checklist-emi.md",
            api_reference="reference/General/emi-apis",
            sdk=None,
            quick_start="docs/Offerings/introduction-to-affordability/emi-api-integration/index.md",
            webhooks=WEB_WEBHOOKS,
            error_codes=None,
            testing=None,
            go_live="docs/Offerings/introduction-to-affordability/emi-api-integration/integration-checklist-emi.md",
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P0",
            recommended_action="Create dedicated EMI Integration Guide covering hosted/MH/S2S + NTB flow; link Affordability suite.",
            notes="EMI NTB has separate reference under Affordability/emi-ntb-flow-apis.",
        ),
        Product(
            name="Offer Engine / Offers",
            category="Offerings / Affordability",
            product_type="Platform / Payment Feature",
            overview="docs/Offerings/introduction-to-affordability/offers-integration-1/index.md",
            integration_guide="docs/Offerings/introduction-to-affordability/offers-integration-1/offers-integration/integration-checklist-offers.md",
            api_reference="reference/Affordability/offer-apis",
            sdk=None,
            quick_start="docs/Offerings/introduction-to-affordability/offers-integration-1/index.md",
            webhooks=None,
            error_codes="docs/Offerings/introduction-to-affordability/offers-integration-1/offers-integration/error-codes-for-offers-integration.md",
            testing=None,
            go_live="docs/Offerings/introduction-to-affordability/offers-integration-1/offers-integration/integration-checklist-offers.md",
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P0",
            recommended_action="Ship Offer Engine Integration Guide (dashboard create → apply at checkout → validate → settle/refund). ASK AI duplicate should redirect to canonical.",
            notes="Dashboard docs under offers-dashboard. ASK AI: affordability-offer-engine-14.md.",
        ),
        Product(
            name="Affordability Widget",
            category="Offerings / Affordability",
            product_type="Platform / Tool",
            overview="docs/Offerings/introduction-to-affordability/affordability-suite/index.md",
            integration_guide="docs/Offerings/introduction-to-affordability/affordability-suite/index.md",
            api_reference=None,
            sdk="docs/Offerings/introduction-to-affordability/affordability-suite/index.md",
            quick_start="docs/Offerings/introduction-to-affordability/affordability-suite/index.md",
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"api_reference": False, "webhooks": False, "error_codes": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Dedicated JS/React widget Integration Guide with plugin variants (Shopify/Woo).",
            notes="JS and React integrations under affordability-suite.",
        ),
        Product(
            name="BNPL / Pay Later",
            category="Offerings / Affordability",
            product_type="Payment Feature",
            overview="docs/Offerings/introduction-to-affordability/payu-bnpl-integration-introduction/index.md",
            integration_guide="docs/Offerings/introduction-to-affordability/payu-bnpl-integration-introduction/index.md",
            api_reference="reference/Affordability/bnpl-integration-apis",
            sdk=None,
            quick_start="docs/Offerings/introduction-to-affordability/payu-bnpl-integration-introduction/index.md",
            webhooks=None,
            error_codes="docs/Offerings/introduction-to-affordability/payu-bnpl-integration-introduction/error-codes-1.md",
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Complete Integration Guide for Link & Pay + quick checkout with testing and go-live.",
            notes="Includes BNPL Link & Pay flows.",
        ),
        Product(
            name="LazyPay Pay-in-3",
            category="Offerings / Affordability",
            product_type="Partner Payment Feature",
            overview="docs/Offerings/introduction-to-affordability/lazypay-pay-in-3/index.md",
            integration_guide="docs/Offerings/introduction-to-affordability/lazypay-pay-in-3/index.md",
            api_reference="reference/Affordability/lazypay-pay-in-3",
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=False,
            priority="P2",
            recommended_action="Expand Merchant Hosted guide; FAQs and testing checklist recommended.",
            notes="Partner product under Affordability.",
        ),
        Product(
            name="MobiKwik Link & Pay",
            category="Offerings / Affordability",
            product_type="Partner Payment Feature",
            overview="docs/Offerings/introduction-to-affordability/mobikwik-link-pay-integration/index.md",
            integration_guide="docs/Offerings/introduction-to-affordability/mobikwik-link-pay-integration/index.md",
            api_reference="reference/Third-Party Wallets/mobikwik-link-wallet-apis",
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing="docs/Offerings/introduction-to-affordability/mobikwik-link-pay-integration/testing-checklist-mobikwik-link-pay.md",
            go_live="docs/Offerings/introduction-to-affordability/mobikwik-link-pay-integration/testing-checklist-mobikwik-link-pay.md",
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=False,
            priority="P2",
            recommended_action="Clarify boundary vs MobiKwik Wallet APIs; keep testing checklist current.",
            notes="Distinct from MobiKwik wallet reference APIs.",
        ),
        Product(
            name="Loyalty Edge",
            category="Offerings / Affordability",
            product_type="Platform / Payment Feature",
            overview="docs/Offerings/introduction-to-affordability/loyalty-edge-introduction/index.md",
            integration_guide=None,
            api_reference=None,
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "api_reference": False, "webhooks": False, "error_codes": False, "changelog": False},
            recommend_ig=False,
            priority="P2",
            recommended_action="Add clear enablement + dashboard journey; Integration Guide only if API/SDK surface expands.",
            notes="Workflow/enablement/dashboard pages under loyalty-edge-introduction.",
        ),
        Product(
            name="Subscriptions / Recurring Payments",
            category="Offerings / Subscriptions",
            product_type="Core Payment Product",
            overview="docs/Offerings/introduction-recurring-payments-integration/index.md",
            integration_guide="docs/Offerings/introduction-recurring-payments-integration/using-api-integration-recurring-payments/index.md",
            api_reference="reference/Subscriptions",
            sdk=None,
            quick_start="docs/Offerings/introduction-recurring-payments-integration/subscriptions-integration/index.md"
            if exists("docs/Offerings/introduction-recurring-payments-integration/subscriptions-integration/index.md")
            else "docs/Offerings/introduction-recurring-payments-integration/using-api-integration-recurring-payments/index.md",
            webhooks="reference/Subscriptions/set-up-webhook-to-receive-cancellation-or-modification-update-from-the-issuer-bank.md",
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs="docs/Offerings/introduction-recurring-payments-integration/faqs-recurring-payments.md",
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P0",
            recommended_action="MUST have a single Integration Guide that unifies SI/Recurring/Subscriptions naming, consent→PDN→recurring, UPI AutoPay, eNACH, and cards; retire ASK AI duplicate into canonical.",
            notes="Duplicate aggregate API: reference/Pre-debit and Recurring Payments Transaction API. Naming: Subscriptions vs Recurring vs SI.",
        ),
        Product(
            name="Zion Subscription Automation",
            category="Offerings / Subscriptions",
            product_type="Platform / Tool",
            overview="docs/Offerings/introduction-recurring-payments-integration/using-zion-subscription-automation-platform/index.md",
            integration_guide="docs/Offerings/introduction-recurring-payments-integration/using-zion-subscription-automation-platform/index.md",
            api_reference="reference/ZION",
            sdk=None,
            quick_start=None,
            webhooks="docs/Offerings/introduction-recurring-payments-integration/using-zion-subscription-automation-platform/webhooks-for-subscription.md",
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs="docs/Offerings/introduction-recurring-payments-integration/using-zion-subscription-automation-platform/faqs-zion-integration.md",
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Dedicated Zion Integration Guide (plans, invoices, webhooks lifecycle) with consistent Zion/ZION casing.",
            notes="API category capitalized as ZION.",
        ),
        Product(
            name="Dynamic Currency Conversion / International Payments",
            category="Offerings / International",
            product_type="Payment Feature",
            overview="docs/Offerings/introduction-dynamic-currency-conversion/index.md",
            integration_guide="docs/Offerings/introduction-dynamic-currency-conversion/payu-hosted-checkout-integration-dynamic-currency-conversion.md",
            api_reference="reference/international payments",
            sdk=None,
            quick_start="docs/Offerings/introduction-dynamic-currency-conversion/dynamic-currency-conversion-workflow.md",
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs="docs/Offerings/introduction-dynamic-currency-conversion/faqs-dynamic-currency-conversion.md",
            changelog=None,
            applicable={"sdk": False, "changelog": False, "webhooks": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Unify DCC vs International Payments naming; clarify MCP (multi-currency pricing) vs MCP (Model Context Protocol) collision in docs IA.",
            notes="MCP Lookup API under international payments. ASK AI single-mid-for-mcc-merchants-flow is related.",
        ),
        Product(
            name="Cross-Border Payments Import (PACB)",
            category="Offerings / International",
            product_type="Core Payment Product",
            overview="docs/Offerings/introduction-cross-border-payments-import/index.md",
            integration_guide="docs/Offerings/introduction-cross-border-payments-import/integrate-cross-border-payments-for-payubiz/index.md"
            if exists(
                "docs/Offerings/introduction-cross-border-payments-import/integrate-cross-border-payments-for-payubiz/index.md"
            )
            else "docs/Offerings/introduction-cross-border-payments-import/index.md",
            api_reference="reference/Cross-border Payments",
            sdk=None,
            quick_start="docs/Offerings/introduction-cross-border-payments-import/workflow-for-cross-border-payments-import.md",
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs="docs/Offerings/introduction-cross-border-payments-import/faqs-for-cross-border-payments.md",
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P0",
            recommended_action="Dedicated Integration Guide covering card/UPI/NB, LRS, virtual accounts, subscriptions-CB, and onboarding; standardize CB/PACB/Import naming.",
            notes="Sub-products: LRS, CB subscriptions, virtual accounts, merchant onboarding APIs.",
        ),
        Product(
            name="EFTNET / NEFT-RTGS Collect",
            category="Offerings",
            product_type="Payment Feature",
            overview="docs/Offerings/introduction-to-eftnet/index.md",
            integration_guide="docs/Offerings/introduction-to-eftnet/payu-hosted-checkout-eftnet.md",
            api_reference="reference/Collect Payment",
            sdk=None,
            quick_start="docs/Offerings/introduction-to-eftnet/collect-payments-with-eftnet-neftrtgs-seamless.md",
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False, "webhooks": False},
            recommend_ig=False,
            priority="P2",
            recommended_action="Clarify Collect EFTNET vs Payouts Dashboard EFTNET funding feature.",
            notes="Also appears under payouts/payouts-dashboard/eftnet.md.",
        ),
        Product(
            name="Pre-Authorize / Auth & Capture",
            category="Offerings",
            product_type="Payment Feature",
            overview="docs/Offerings/auth-and-capture-pre-authorize-card-payments/index.md",
            integration_guide="docs/Offerings/auth-and-capture-pre-authorize-card-payments/pre-authorize-card-transactions/index.md"
            if exists(
                "docs/Offerings/auth-and-capture-pre-authorize-card-payments/pre-authorize-card-transactions/index.md"
            )
            else "docs/Offerings/auth-and-capture-pre-authorize-card-payments/index.md",
            api_reference="reference/Pre-Authorize Payment",
            sdk=None,
            quick_start="docs/Offerings/auth-and-capture-pre-authorize-card-payments/index.md",
            webhooks=None,
            error_codes="reference/Pre-Authorize Payment/error-codes-pre-authorize-payment.md",
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Integration Guide for card pre-auth + capture/cancel with S2S variant.",
            notes="Includes UPI OTM under same parent folder.",
        ),
        Product(
            name="UPI One-Time Mandate (OTM / Reserve Pay)",
            category="Offerings",
            product_type="Payment Feature",
            overview="docs/Offerings/auth-and-capture-pre-authorize-card-payments/upi-one-time-mandate-integration/index.md",
            integration_guide="docs/Offerings/auth-and-capture-pre-authorize-card-payments/upi-one-time-mandate-integration/index.md",
            api_reference="reference/Pre-Authorize Payment/pre-authorize-payments-for-upi",
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes="reference/Pre-Authorize Payment/error-codes-pre-authorize-payment.md",
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Promote OTM to first-class Integration Guide; merge ASK AI upi-otm into canonical docs.",
            notes="Reserve Pay page under upi-one-time-mandate-integration.",
        ),
        Product(
            name="Split Settlements (Aggregator / Marketplace)",
            category="Offerings",
            product_type="Core Payment Product",
            overview="docs/Offerings/split-settlments/index.md",
            integration_guide="docs/Offerings/split-settlments/introduction-split-settlements/index.md"
            if exists("docs/Offerings/split-settlments/introduction-split-settlements/index.md")
            else "docs/Offerings/split-settlments/split-settlements-payment-integration.md",
            api_reference="reference/split settlements",
            sdk=None,
            quick_start="docs/Offerings/split-settlments/payu-hosted-integration-split-settlements",
            webhooks=None,
            error_codes="docs/Offerings/split-settlments/error-codes-for-refunds-status.md",
            testing="reference/split settlements/split_after_transaction_api/test-the-integration-split-after-transaction-1.md",
            go_live=None,
            troubleshooting=None,
            faqs="docs/Offerings/split-settlments/faqs-for-split-settlements.md",
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P0",
            recommended_action="Fix folder typo split-settlments; publish marketplace Integration Guide (parent/child onboarding → split during/after txn → refunds/settlements).",
            notes="ASK AI duplicate: split-settlements-aggregator.md. Strong API coverage with test pages.",
        ),
        Product(
            name="Tokenization / Save Cards (Vault)",
            category="Offerings",
            product_type="Payment Feature / Platform",
            overview="docs/Offerings/introduction-save-cards/index.md",
            integration_guide="docs/Offerings/introduction-save-cards/index.md",
            api_reference="reference/Tokenization",
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P0",
            recommended_action="Dedicated Integration Guide for Models 1–3 + push tokenization; critical for PCI-reducing integrations and recurring.",
            notes="Naming: Tokenization / Save Cards / Vault / Store Card across files.",
        ),
        Product(
            name="Third-Party Verification (TPV)",
            category="Offerings",
            product_type="Payment Feature",
            overview="docs/Offerings/introduction-to-payu-tpv/index.md",
            integration_guide="docs/Offerings/introduction-to-payu-tpv/collect-payments-with-tpv-payu-hosted-checkout/index.md"
            if exists(
                "docs/Offerings/introduction-to-payu-tpv/collect-payments-with-tpv-payu-hosted-checkout/index.md"
            )
            else "docs/Offerings/introduction-to-payu-tpv/index.md",
            api_reference="Partial",
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs="docs/Offerings/introduction-to-payu-tpv/faqs-tpv-integration.md",
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P0",
            recommended_action="Create TPV Integration Guide covering Hosted/MH/S2S/Payment Link/NEFT; consolidate dispersed API JSON (tpv-*.json) into reference IA; merge ASK AI guides.",
            notes="ASK AI: upi-net-banking-tpv-integration.md, payment-link-tpv-complete-guide.md. API specs dispersed.",
        ),
        Product(
            name="Recommendation Engine",
            category="Offerings",
            product_type="Platform / Payment Feature",
            overview="docs/Offerings/recommendation-engine/index.md",
            integration_guide="docs/Offerings/recommendation-engine/fetch-recommendation-engine-api.md",
            api_reference="docs/Offerings/recommendation-engine/fetch-recommendation-engine-api.md",
            sdk=None,
            quick_start="docs/Offerings/recommendation-engine/re-customer-journey.md",
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "webhooks": False, "changelog": False},
            recommend_ig=False,
            priority="P2",
            recommended_action="Document CheckoutPro/CommercePro embedding paths; light Integration Guide optional.",
            notes="Also embedded in CheckoutPro SDKs.",
        ),
        Product(
            name="Native OTP Flow",
            category="Offerings",
            product_type="Payment Feature",
            overview="docs/Offerings/native-otp-flow-integration/index.md",
            integration_guide="docs/Offerings/native-otp-flow-integration/collect-payments-with-card-native-otp-flow.md",
            api_reference="reference/Collect Payment/native-otp-flow-apis",
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Clarify Native OTP Flow (web/API) vs Native OTP Assist mobile SDKs; ship Integration Guide for card/debit/cardless EMI.",
            notes="Separate from Android/iOS Native OTP Assist SDKs.",
        ),
        Product(
            name="Apple Pay",
            category="Offerings",
            product_type="Partner Payment Method",
            overview="docs/Offerings/apple-pay-integration/index.md",
            integration_guide="docs/Offerings/apple-pay-integration/prerequisites-and-set-up-for-apple-pay-integration.md",
            api_reference="reference/Apple Pay",
            sdk=None,
            quick_start="docs/Offerings/apple-pay-integration/apple-pay-integration-payu-hosted-checkout.md",
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Keep prerequisites + Hosted/MH/session/UI as one Integration Guide with Apple certification/testing steps.",
            notes="Session management and UI seamless pages exist.",
        ),
        Product(
            name="Account Funding Transaction (AFT)",
            category="Offerings",
            product_type="Payment Feature",
            overview="docs/Offerings/account-funding-transaction-integration/index.md",
            integration_guide="docs/Offerings/account-funding-transaction-integration/collection-payments-with-account-funding-transaction.md",
            api_reference="reference/Collect Payment/_payment_merchant_hosted/_payment_api_merchant_hosted_aft.md",
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False, "webhooks": False},
            recommend_ig=False,
            priority="P2",
            recommended_action="Add compliance notes and testing guidance; dedicated IG only if adoption rises.",
            notes="Thin but present guide + API.",
        ),
        Product(
            name="Mutual Fund Payments (WealthTech)",
            category="Offerings",
            product_type="Partner / Vertical Product",
            overview="docs/Offerings/mutual-funds-payments/index.md",
            integration_guide="docs/Offerings/mutual-funds-payments/payu-hosted-integration-mutual-funds-payment.md",
            api_reference="reference/Wealth Tech",
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Align Mutual Fund Payments guide naming with Wealth Tech API category; Integration Guide covering eNACH + UPI AutoPay.",
            notes="Hosted, MH, eNACH, UPI AutoPay pages present.",
        ),
        Product(
            name="Merchant Wallet",
            category="Offerings",
            product_type="Core Payment Product",
            overview="docs/Offerings/introduction-to-merchant-wallet/index.md",
            integration_guide="docs/Offerings/introduction-to-merchant-wallet/closed-loop-wallet-management/index.md",
            api_reference="reference/Merchant Wallet",
            sdk=None,
            quick_start=None,
            webhooks="Partial",
            error_codes="Partial",
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Integration Guides for Closed-Loop vs Open/Semi-Closed issuance journeys; document callbacks/status codes clearly.",
            notes="Closed-loop and open-loop API folders under Merchant Wallet.",
        ),
        Product(
            name="Virtual Cards / GPR Cards",
            category="Offerings",
            product_type="Partner Product",
            overview="docs/Offerings/virtual-cards-introduction/index.md",
            integration_guide="docs/Offerings/virtual-cards-introduction/web-integration-virtual-cards/index.md"
            if exists("docs/Offerings/virtual-cards-introduction/web-integration-virtual-cards/index.md")
            else "docs/Offerings/virtual-cards-introduction/apis-used-in-virtual-cards-integration.md",
            api_reference="docs/Offerings/virtual-cards-introduction/apis-used-in-virtual-cards-integration.md",
            sdk="docs/Offerings/virtual-cards-introduction/virtual-card-integration-in-android.md",
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"changelog": False},
            recommend_ig=True,
            priority="P2",
            recommended_action="One Integration Guide with web + mobile SDK matrix; promote API map to reference section.",
            notes="Android/iOS/Flutter/RN SDK guides exist under virtual-cards-introduction.",
        ),
        Product(
            name="Refunds",
            category="Offerings",
            product_type="Payment Feature",
            overview="docs/Offerings/introduction-refunds/index.md",
            integration_guide="docs/Offerings/introduction-refunds/index.md",
            api_reference="reference/General/refund-apis",
            sdk=None,
            quick_start=None,
            webhooks="docs/Offerings/introduction-refunds/webhooks-for-refunds.md",
            error_codes="reference/General/refund-apis/error-codes-for-refund-initiation.md",
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs="docs/Offerings/introduction-refunds/faqs-for-refunds.md",
            changelog=None,
            applicable={"sdk": False, "changelog": False, "integration_guide": False},
            recommend_ig=False,
            priority="P1",
            recommended_action="Keep API + webhooks + product-specific refund pages strong; cross-link from every collect product. Dedicated IG not needed beyond current structure.",
            notes="Product-specific refunds documented under refunds-in-payu-products. ASK AI refund.md duplicate.",
        ),
        Product(
            name="Chargebacks",
            category="Offerings",
            product_type="Payment Feature",
            overview="docs/Offerings/chargeback/index.md",
            integration_guide=None,
            api_reference="reference/Chargeback",
            sdk=None,
            quick_start=None,
            webhooks="docs/Offerings/chargeback/webhooks-for-chargeback",
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "integration_guide": False, "changelog": False, "quick_start": False},
            recommend_ig=False,
            priority="P2",
            recommended_action="Add FAQs and operational runbooks; Integration Guide not required (ops/API product).",
            notes="Process, types, statuses, reasons, dashboard, webhooks present.",
        ),
        Product(
            name="Rewards / RewardX Partner Integration",
            category="Offerings",
            product_type="Partner Payment Feature",
            overview="docs/Offerings/rewards-partner-integration/index.md",
            integration_guide="docs/Offerings/rewards-partner-integration/rewards-pay-redemption-integration.md",
            api_reference="reference/REWARD PARTNERS",
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P2",
            recommended_action="Consolidate RewardX, Pay with Rewards/TWID, and Flipkart SuperCoins IA to reduce duplication.",
            notes="Overlaps with reference/Pay with rewards and Flipkart supercoins.",
        ),
        Product(
            name="Banking Connect (IBMB / NBBL)",
            category="Offerings",
            product_type="Partner Payment Feature",
            overview="docs/Offerings/banking-connect-ibmb-or-nbbl/index.md",
            integration_guide="docs/Offerings/banking-connect-ibmb-or-nbbl/payu-hosted-customer-journey-banking-connect.md",
            api_reference=None,
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "api_reference": True, "changelog": False, "webhooks": False},
            recommend_ig=False,
            priority="P3",
            recommended_action="Docs appear hidden/limited; expand when product is generally available.",
            notes="NBBL overview present. Naming: Banking Connect / IBMB / NBBL.",
        ),
        Product(
            name="Pluxee Card (Sodexo)",
            category="Collect Payments / Payment Methods",
            product_type="Partner Payment Method",
            overview="docs/Collect Payments/introduction-web/custom-checkout-merchant-hosted/integrate-with-merchant-hosted-checkout-for-pluxee-card.md",
            integration_guide="docs/Collect Payments/introduction-web/custom-checkout-merchant-hosted/integrate-with-merchant-hosted-checkout-for-pluxee-card.md",
            api_reference="reference/Sodexo",
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False, "webhooks": False},
            recommend_ig=False,
            priority="P3",
            recommended_action="Rename Sodexo reference to Pluxee; keep as method page under Merchant Hosted.",
            notes="Brand Pluxee vs legacy Sodexo API category.",
        ),
    ]
    products.extend(offerings)

    # -------------------------------------------------------------------------
    # Payouts
    # -------------------------------------------------------------------------
    products.append(
        Product(
            name="PayU Payouts",
            category="Payouts",
            product_type="Core Payment Product",
            overview="docs/payouts/introduction-to-payouts.md",
            integration_guide="docs/payouts/payouts-integration/single-transfer-integration-for-payouts.md",
            api_reference="reference/payouts",
            sdk=None,
            quick_start="docs/payouts/process-flow-for-payouts.md",
            webhooks="docs/payouts/payouts-integration/payouts-webhooks.md",
            error_codes="reference/payouts/error-codes-for-payouts.md",
            testing="docs/payouts/test-credentials-for-payouts.md",
            go_live=None,
            troubleshooting=None,
            faqs="docs/payouts/faqs-for-payouts.md",
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P0",
            recommended_action="Elevate Single Transfer + Smart Send + Beneficiary + Webhooks into a master Payouts Integration Guide with go-live; remove releasepending- filename residue from Pay to Phone.",
            notes="Sub-features: Smart Send, Pay to Phone, Bulk, Beneficiary, Verification, Partner Onboarding, MCP tools. Custom page duplicate: custom_pages/payouts-introduction.md.",
        )
    )
    products.append(
        Product(
            name="Smart Send",
            category="Payouts",
            product_type="Payment Feature",
            overview="docs/payouts/payouts-integration/smart-send-introduction/index.md",
            integration_guide="docs/payouts/payouts-integration/smart-send-introduction/index.md",
            api_reference="reference/payouts/smart-send-apis",
            sdk=None,
            quick_start=None,
            webhooks="docs/payouts/payouts-integration/payouts-webhooks.md",
            error_codes="reference/payouts/smart-send-apis/smart-send-error-codes.md",
            testing="docs/payouts/test-credentials-for-payouts.md",
            go_live=None,
            troubleshooting=None,
            faqs="docs/payouts/faqs-for-payouts.md",
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=False,
            priority="P1",
            recommended_action="Keep as chapter of Payouts IG rather than separate IG.",
            notes="Error codes present under smart-send-apis.",
        )
    )
    products.append(
        Product(
            name="Pay to Phone",
            category="Payouts",
            product_type="Payment Feature",
            overview="docs/payouts/releasepending-pay-to-phone-integration/index.md"
            if exists("docs/payouts/releasepending-pay-to-phone-integration/index.md")
            else "docs/payouts/releasepending-pay-to-phone-integration/releasepending-pay-to-phone-initiation.md",
            integration_guide="docs/payouts/releasepending-pay-to-phone-integration/releasepending-pay-to-phone-initiation.md",
            api_reference="reference/payouts/releasepending-pay-to-phone-configuration-apis",
            sdk=None,
            quick_start=None,
            webhooks="docs/payouts/payouts-integration/payouts-webhooks.md",
            error_codes="reference/payouts/error-codes-for-payouts.md",
            testing="docs/payouts/test-credentials-for-payouts.md",
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Rename releasepending- paths; complete Integration Guide for initiation/tracking/configuration.",
            notes="Editorial residue in filenames (releasepending-).",
        )
    )

    # -------------------------------------------------------------------------
    # Partners
    # -------------------------------------------------------------------------
    products.append(
        Product(
            name="PayU Partner Program & Portal",
            category="Partners",
            product_type="Partner Product / Platform",
            overview="docs/partners/payu-partner-program-overview.md",
            integration_guide=None,
            api_reference=None,
            sdk=None,
            quick_start="docs/partners/partner-portal/index.md",
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs="docs/partners/faqs-partner-integration.md",
            changelog=None,
            applicable={
                "integration_guide": False,
                "api_reference": False,
                "sdk": False,
                "webhooks": False,
                "error_codes": False,
                "changelog": False,
            },
            recommend_ig=False,
            priority="P2",
            recommended_action="Keep portal how-tos current; link to API onboarding guides.",
            notes="Incentives, referral links, portal user management documented.",
        )
    )
    products.append(
        Product(
            name="Partner Merchant Onboarding (API / OAuth)",
            category="Partners",
            product_type="Partner Product",
            overview="docs/partners/internal-reviewpartner-integration-overview/index.md"
            if exists("docs/partners/internal-reviewpartner-integration-overview/index.md")
            else "docs/partners/refer-merchants-using-api.md",
            integration_guide="docs/partners/internal-reviewpartner-integration-overview/quick-start-partner-integration.md",
            api_reference="reference/ParTner integration",
            sdk=None,
            quick_start="docs/partners/internal-reviewpartner-integration-overview/quick-start-partner-integration.md",
            webhooks="reference/ParTner integration/using-webhooks-for-merchant-status",
            error_codes="reference/ParTner integration/using-webhooks-for-merchant-status/kyc-errors-and-solutions.md",
            testing="docs/partners/internal-reviewpartner-integration-overview/testing-and-go-live-partner-integration.md",
            go_live="docs/partners/internal-reviewpartner-integration-overview/testing-and-go-live-partner-integration.md",
            troubleshooting="docs/partners/internal-reviewpartner-integration-overview/errors-and-troubleshooting-partner-integration.md",
            faqs="docs/partners/faqs-partner-integration.md",
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P0",
            recommended_action="Finalize internal-review partner overview as canonical Integration Guide; merge duplicate Partner Integration - Merchant Onboarding APIs tree; fix ParTner folder casing.",
            notes="Parallel API trees: ParTner integration vs Partner Integration - Merchant Onboarding APIs (16-step). Co-branded OAuth has its own folder.",
        )
    )
    products.append(
        Product(
            name="Partner Payments Integration",
            category="Partners",
            product_type="Partner Product",
            overview="docs/partners/partner-payments-integration.md",
            integration_guide="docs/partners/partner-payments-integration.md",
            api_reference="reference/ParTner integration/partner-payment-integration-apis",
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs="docs/partners/faqs-partner-integration.md",
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Expand Integration Guide for Hosted/UPI S2S/UPI TPV/refunds partner flows.",
            notes="Payment APIs under partner-payment-integration-apis.",
        )
    )

    # -------------------------------------------------------------------------
    # WhatsApp, BBPS, MCP/CLI, Monitoring, Air India
    # -------------------------------------------------------------------------
    products.append(
        Product(
            name="WhatsApp Payments",
            category="WhatsApp Integration",
            product_type="Partner / Integration Channel",
            overview="docs/Whatsapp integration/whatsapp-integration-introduction.md",
            integration_guide="docs/Whatsapp integration/whatsapp-native-payments/index.md",
            api_reference=None,
            sdk=None,
            quick_start="docs/Whatsapp integration/whatsapp-native-payments/integrate-whatsapp-payments-p2m-or-up-intent.md"
            if exists(
                "docs/Whatsapp integration/whatsapp-native-payments/integrate-whatsapp-payments-p2m-or-up-intent.md"
            )
            else None,
            webhooks=WEB_WEBHOOKS,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "api_reference": True, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Dedicated Integration Guide for Native P2M + Enhanced Payment Links; clarify Interakt plugin boundary.",
            notes="Includes Enhanced Payment Links on WhatsApp page.",
        )
    )
    products.append(
        Product(
            name="BBPS Connect Agent",
            category="BBPS",
            product_type="Core / Partner Product",
            overview="docs/BBPS/connect-agent-api-integration/index.md",
            integration_guide="docs/BBPS/connect-agent-api-integration/bbps-integration-flow.md",
            api_reference="reference/BBPS",
            sdk=None,
            quick_start="docs/BBPS/connect-agent-api-integration/bbps-integration-flow.md",
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Expand Integration Guide with auth, bill fetch/pay, complaints, sandbox, error catalog.",
            notes="Rich API reference under reference/BBPS.",
        )
    )
    products.append(
        Product(
            name="BBPS Prepaid Recharge",
            category="BBPS",
            product_type="Payment Feature",
            overview="docs/BBPS/recharge-api-integration/index.md",
            integration_guide="docs/BBPS/recharge-api-integration/prepaid-recharge-workflow.md",
            api_reference="reference/BBPS/bbps-prepaid-apis",
            sdk=None,
            quick_start="docs/BBPS/recharge-api-integration/prepaid-recharge-workflow.md",
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"sdk": False, "changelog": False},
            recommend_ig=False,
            priority="P2",
            recommended_action="Keep as chapter of BBPS IG; add testing/errors.",
            notes="Workflow + prepaid APIs present.",
        )
    )
    products.append(
        Product(
            name="PayU Remote MCP Server",
            category="MCP & CLI",
            product_type="Platform / Tool",
            overview="docs/MCP & CLI/payu-remote-mcp-server-integration.md",
            integration_guide="docs/MCP & CLI/payu-remote-mcp-server-integration.md",
            api_reference=None,
            sdk=None,
            quick_start="docs/MCP & CLI/payu-remote-mcp-server-integration.md",
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"api_reference": False, "sdk": False, "webhooks": False, "error_codes": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Treat current page as Integration Guide seed; add tool catalog, auth (OAuth 2.1), examples, troubleshooting. Disambiguate MCP acronym from Multi-Currency Pricing.",
            notes="Single-page today. Distinct from payouts/mcp-tools.md.",
        )
    )
    products.append(
        Product(
            name="PayU CLI",
            category="MCP & CLI",
            product_type="Platform / Tool",
            overview="docs/MCP & CLI/payu-cli.md",
            integration_guide="docs/MCP & CLI/payu-cli.md",
            api_reference=None,
            sdk=None,
            quick_start="docs/MCP & CLI/payu-cli.md",
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={
                "api_reference": False,
                "sdk": False,
                "webhooks": False,
                "error_codes": False,
                "changelog": False,
            },
            recommend_ig=False,
            priority="P2",
            recommended_action="Expand command reference and examples; dedicated IG optional.",
            notes="Install/config/commands on single page.",
        )
    )
    products.append(
        Product(
            name="PayU Devguide Builder MCP",
            category="MCP & CLI",
            product_type="Platform / Tool (Internal Docs)",
            overview="docs/MCP & CLI/payu-devguide-builder-mcp-configuration.md",
            integration_guide="docs/MCP & CLI/payu-devguide-builder-mcp-configuration.md",
            api_reference=None,
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={
                "api_reference": False,
                "sdk": False,
                "webhooks": False,
                "error_codes": False,
                "testing": False,
                "go_live": False,
                "faqs": False,
                "changelog": False,
            },
            recommend_ig=False,
            priority="P3",
            recommended_action="Internal docs tooling — maintain configuration page only.",
            notes="Documentation authoring tooling, not merchant-facing product.",
        )
    )
    products.append(
        Product(
            name="Agentic Commerce Suite",
            category="MCP & CLI",
            product_type="Platform / Partner Product",
            overview="docs/MCP & CLI/agentic-commerce/index.md",
            integration_guide="docs/MCP & CLI/agentic-commerce/build-your-own-chatgpt-merchant-app.md",
            api_reference=None,
            sdk=None,
            quick_start="docs/MCP & CLI/agentic-commerce/build-your-own-chatgpt-merchant-app.md",
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={"api_reference": True, "sdk": False, "changelog": False},
            recommend_ig=True,
            priority="P1",
            recommended_action="Emerging DevEx surface — invest in Integration Guide covering ChatGPT merchant app and agent-compatible checkout.",
            notes="Overview includes AI-app commerce, WhatsApp commerce, merchant-owned assistants.",
        )
    )
    products.append(
        Product(
            name="PayU Overwatch (Monitoring & Alerts)",
            category="Monitoring & Alerts",
            product_type="Platform / Tool",
            overview="docs/Monitoring & Alerts/payu-monitoring-alerts-overwatch/index.md",
            integration_guide=None,
            api_reference=None,
            sdk=None,
            quick_start=None,
            webhooks="docs/Monitoring & Alerts/payu-monitoring-alerts-overwatch/webhook-alerts.md",
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={
                "integration_guide": False,
                "api_reference": False,
                "sdk": False,
                "error_codes": False,
                "changelog": False,
                "quick_start": False,
            },
            recommend_ig=False,
            priority="P2",
            recommended_action="Add setup guide and alert catalog; IG not required.",
            notes="Webhook alerts documented.",
        )
    )
    products.append(
        Product(
            name="Air India Checkout API Suite",
            category="Merchant-Specific",
            product_type="Partner / Merchant-Specific",
            overview="docs/AIR India/air-india-integration-apis.md",
            integration_guide=None,
            api_reference="docs/AIR India",
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={
                "sdk": False,
                "integration_guide": False,
                "webhooks": False,
                "changelog": False,
                "faqs": False,
            },
            recommend_ig=False,
            priority="P3",
            recommended_action="Keep merchant-specific; do not generalize into public product catalog without product decision.",
            notes="Hidden merchant-specific APIs (order, offers, EMI, capture, refund).",
        )
    )
    products.append(
        Product(
            name="Settlements / Priority Settlements",
            category="Getting Started / Dashboard",
            product_type="Payment Feature / Platform",
            overview="docs/getting started/payu-dashboard/settlements-dashboard/index.md"
            if exists("docs/getting started/payu-dashboard/settlements-dashboard/index.md")
            else "docs/getting started/payu-dashboard/settlements-dashboard/priority-settlements.md",
            integration_guide=None,
            api_reference="reference/Settlements",
            sdk=None,
            quick_start=None,
            webhooks=None,
            error_codes=None,
            testing=None,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={
                "sdk": False,
                "integration_guide": False,
                "webhooks": False,
                "changelog": False,
            },
            recommend_ig=False,
            priority="P2",
            recommended_action="Improve API + dashboard cross-links; Priority Settlements page should link Settlements APIs.",
            notes="Settlement APIs: transaction details, date range, upcoming, release.",
        )
    )
    products.append(
        Product(
            name="General Utility APIs (Verify / BIN / Health / Bank Verification)",
            category="API Basics / Utilities",
            product_type="Utility",
            overview="docs/API basics/rest-api-format.md",
            integration_guide=None,
            api_reference="reference/General",
            sdk=None,
            quick_start="docs/API basics/api-authentication-and-security.md",
            webhooks=WEB_WEBHOOKS,
            error_codes=WEB_ERROR_REF,
            testing=WEB_TESTING,
            go_live=None,
            troubleshooting=None,
            faqs=None,
            changelog=None,
            applicable={
                "sdk": False,
                "integration_guide": False,
                "changelog": False,
                "go_live": False,
            },
            recommend_ig=False,
            priority="P1",
            recommended_action="Do not create product IG; instead improve discoverability from checkout IGs (verify payment, BIN, health check).",
            notes="Includes check-transaction, transaction-detail, bin, get_checkout_details, health-check, bank-verification APIs. Hash tool under API basics.",
        )
    )

    # Validate paths lightly — mark missing explicit paths as None already handled by flag()
    return products


# ---------------------------------------------------------------------------
# Excel styling helpers
# ---------------------------------------------------------------------------

FILL_HEADER = PatternFill("solid", fgColor="0B3D5C")
FILL_HEADER2 = PatternFill("solid", fgColor="145A86")
FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_RED = PatternFill("solid", fgColor="FFC7CE")
FILL_BLUE = PatternFill("solid", fgColor="DDEBF7")
FILL_ORANGE = PatternFill("solid", fgColor="FCE4D6")
FILL_GREY = PatternFill("solid", fgColor="F2F2F2")
FILL_P0 = PatternFill("solid", fgColor="FF6B6B")
FILL_P1 = PatternFill("solid", fgColor="FFD93D")
FILL_P2 = PatternFill("solid", fgColor="6BCB77")
FILL_P3 = PatternFill("solid", fgColor="B0B0B0")
FILL_TITLE = PatternFill("solid", fgColor="072A40")
FILL_KPI = PatternFill("solid", fgColor="E8F4FC")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")

FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_TITLE = Font(name="Calibri", bold=True, color="FFFFFF", size=18)
FONT_SECTION = Font(name="Calibri", bold=True, color="0B3D5C", size=13)
FONT_KPI = Font(name="Calibri", bold=True, color="0B3D5C", size=20)
FONT_BODY = Font(name="Calibri", size=10)
FONT_BOLD = Font(name="Calibri", bold=True, size=10)
FONT_LINK = Font(name="Calibri", size=10, color="0563C1", underline="single")
FONT_GREEN = Font(name="Calibri", size=10, color="006100")
FONT_YELLOW = Font(name="Calibri", size=10, color="9C5700")
FONT_RED = Font(name="Calibri", size=10, color="9C0006")

THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def autosize(ws, min_width=10, max_width=48, extra=2):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col:
            if cell.value is None:
                continue
            length = max(length, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, length + extra))


def style_header_row(ws, row, start_col=1, end_col=None):
    end_col = end_col or ws.max_column
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN


def apply_yn_fill(cell):
    v = str(cell.value or "").strip()
    if v == "Yes":
        cell.fill = FILL_GREEN
        cell.font = FONT_GREEN
    elif v == "Partial":
        cell.fill = FILL_YELLOW
        cell.font = FONT_YELLOW
    elif v == "No":
        cell.fill = FILL_RED
        cell.font = FONT_RED
    elif v == "N/A":
        cell.fill = FILL_GREY
    cell.alignment = CENTER
    cell.border = THIN


def apply_status_fill(cell):
    v = str(cell.value or "").strip()
    if v == "Complete":
        cell.fill = FILL_GREEN
        cell.font = FONT_GREEN
    elif v == "Partial":
        cell.fill = FILL_YELLOW
        cell.font = FONT_YELLOW
    elif v == "Missing":
        cell.fill = FILL_RED
        cell.font = FONT_RED
    cell.alignment = CENTER
    cell.border = THIN


def apply_priority_fill(cell):
    v = str(cell.value or "").strip()
    if v == "P0":
        cell.fill = FILL_P0
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    elif v == "P1":
        cell.fill = FILL_P1
        cell.font = Font(name="Calibri", bold=True, size=10)
    elif v == "P2":
        cell.fill = FILL_P2
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    elif v == "P3":
        cell.fill = FILL_P3
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    cell.alignment = CENTER
    cell.border = THIN


def set_hyperlink(cell, path: str):
    if not path:
        cell.value = "—"
        cell.alignment = CENTER
        cell.border = THIN
        return
    # Relative repo path as display; hyperlink as file path for Excel
    cell.value = path
    cell.hyperlink = path
    cell.font = FONT_LINK
    cell.alignment = WRAP
    cell.border = THIN


def product_row_data(p: Product) -> dict:
    return {
        "Product Name": p.name,
        "Product Category": p.category,
        "Product Type": p.product_type,
        "Overview Page Exists": p.flag("overview"),
        "Overview Page Link": p.link("overview"),
        "Integration Guide Exists": p.flag("integration_guide"),
        "Integration Guide Link": p.link("integration_guide"),
        "API Reference Exists": p.flag("api_reference"),
        "API Reference Link": p.link("api_reference"),
        "SDK Exists": p.flag("sdk"),
        "SDK Link": p.link("sdk"),
        "Quick Start Exists": p.flag("quick_start"),
        "Quick Start Link": p.link("quick_start"),
        "Webhooks Documented": p.flag("webhooks"),
        "Webhooks Link": p.link("webhooks"),
        "Error Codes Documented": p.flag("error_codes"),
        "Error Codes Link": p.link("error_codes"),
        "Testing Guide Exists": p.flag("testing"),
        "Testing Guide Link": p.link("testing"),
        "Go Live Guide Exists": p.flag("go_live"),
        "Go Live Guide Link": p.link("go_live"),
        "Troubleshooting Exists": p.flag("troubleshooting"),
        "Troubleshooting Link": p.link("troubleshooting"),
        "FAQs Exists": p.flag("faqs"),
        "FAQs Link": p.link("faqs"),
        "Changelog Exists": p.flag("changelog"),
        "Changelog Link": p.link("changelog"),
        "Documentation Status": p.status(),
        "Documentation Coverage (%)": p.coverage_score(),
        "Recommend Dedicated Integration Guide": "Yes" if p.recommend_ig else "No",
        "Recommended Priority": p.priority,
        "Recommended Action": p.recommended_action,
        "Notes": p.notes,
    }


# --- Fix-first ranking used only on the Coverage Scoring sheet ---
_TIER_WEIGHT = {"P0": 1000, "P1": 700, "P2": 400, "P3": 100}
_CORE_BOOST = {
    "PayU Hosted Checkout (Prebuilt)": 50,
    "Merchant Hosted Checkout (Custom / Seamless)": 50,
    "Server-to-Server (S2S) Checkout": 48,
    "Payment Links": 45,
    "Checkout Plus (ICP / Bolt Checkout)": 42,
    "Android CheckoutPro SDK": 40,
    "iOS CheckoutPro SDK": 40,
    "Subscriptions / Recurring Payments": 45,
    "PayU Payouts": 45,
    "Partner Merchant Onboarding (API / OAuth)": 42,
    "Tokenization / Save Cards (Vault)": 40,
    "Third-Party Verification (TPV)": 40,
    "Split Settlements (Aggregator / Marketplace)": 38,
    "Cross-Border Payments Import (PACB)": 38,
    "EMI / Cardless EMI": 35,
    "Offer Engine / Offers": 35,
}
_COMPLEX_KEYS = ("S2S", "TPV", "Partner", "Subscriptions", "Tokenization", "Split", "Cross-Border")


def fix_priority_score(p: Product) -> float:
    """Higher = fix sooner. Used to order the Coverage Scoring sheet."""
    tier = _TIER_WEIGHT.get(p.priority, 0)
    gap = (100.0 - p.coverage_score()) * 2.0
    core = _CORE_BOOST.get(p.name, 0)
    ig = 30 if p.recommend_ig else 0
    complexity = 20 if any(k in p.name for k in _COMPLEX_KEYS) else 0
    return tier + gap + core + ig + complexity


def why_fix_first(p: Product) -> str:
    parts = []
    parts.append(f"{p.priority} tier")
    cov = p.coverage_score()
    if cov < 55:
        parts.append(f"large coverage gap ({cov}%)")
    elif cov < 75:
        parts.append(f"moderate coverage gap ({cov}%)")
    else:
        parts.append(f"coverage already stronger ({cov}%) — maintain/polish")
    if p.name in _CORE_BOOST:
        parts.append("core payment / revenue journey")
    if p.recommend_ig:
        parts.append("dedicated Integration Guide recommended")
    if any(k in p.name for k in _COMPLEX_KEYS):
        parts.append("high developer complexity / support dependency")
    return "; ".join(parts)


def what_to_fix(p: Product) -> str:
    critical = [
        ("Overview", "overview"),
        ("Integration Guide", "integration_guide"),
        ("API Reference", "api_reference"),
        ("Testing", "testing"),
        ("Go Live", "go_live"),
        ("Webhooks", "webhooks"),
        ("Error Codes", "error_codes"),
        ("Troubleshooting", "troubleshooting"),
        ("FAQs", "faqs"),
        ("Changelog", "changelog"),
    ]
    missing = [label for label, dim in critical if p.flag(dim) == "No"]
    partial = [label for label, dim in critical if p.flag(dim) == "Partial"]
    bits = []
    if missing:
        bits.append("Missing: " + ", ".join(missing[:5]))
    if partial:
        bits.append("Partial: " + ", ".join(partial[:3]))
    return "; ".join(bits) if bits else "No critical gaps — keep current quality"


def build_workbook(products: list[Product]) -> Workbook:
    wb = Workbook()
    rows = [product_row_data(p) for p in products]

    total = len(rows)
    with_overview = sum(1 for r in rows if r["Overview Page Exists"] == "Yes")
    missing_overview = sum(1 for r in rows if r["Overview Page Exists"] == "No")
    with_ig = sum(1 for r in rows if r["Integration Guide Exists"] == "Yes")
    missing_ig = sum(1 for r in rows if r["Integration Guide Exists"] == "No")
    with_api = sum(1 for r in rows if r["API Reference Exists"] == "Yes")
    with_sdk = sum(1 for r in rows if r["SDK Exists"] == "Yes")
    complete = sum(1 for r in rows if r["Documentation Status"] == "Complete")
    partial = sum(1 for r in rows if r["Documentation Status"] == "Partial")
    missing = sum(1 for r in rows if r["Documentation Status"] == "Missing")
    requiring = partial + missing
    overall_coverage = round(sum(r["Documentation Coverage (%)"] for r in rows) / total, 1) if total else 0

    p0 = [p for p in products if p.priority == "P0" and p.recommend_ig]
    p1 = [p for p in products if p.priority == "P1" and p.recommend_ig]
    p2 = [p for p in products if p.priority == "P2" and p.recommend_ig]

    # =====================================================================
    # Sheet 1 — Executive Dashboard
    # =====================================================================
    ws = wb.active
    ws.title = "Executive Dashboard"

    ws.merge_cells("A1:L1")
    ws["A1"] = "PayU Product Documentation Coverage Tracker — Executive Dashboard"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:L2")
    ws["A2"] = (
        f"Source of truth: PayU Developer Documentation repository (docs/, reference/, recipes/, custom_pages/). "
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}. "
        f"DevEx goal: Enable developers to successfully integrate PayU products without requiring support intervention."
    )
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="333333")
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 36

    # KPI headers
    kpis = [
        ("Total Products", total),
        ("Overall Coverage %", f"{overall_coverage}%"),
        ("Complete", complete),
        ("Partial", partial),
        ("Missing / Require Docs", requiring),
        ("With Overview", with_overview),
        ("Missing Overview", missing_overview),
        ("With Integration Guide", with_ig),
        ("Missing Integration Guide", missing_ig),
        ("With API Reference", with_api),
        ("With SDK Docs", with_sdk),
        ("P0 IG Recommendations", len(p0)),
    ]
    for i, (label, value) in enumerate(kpis):
        col = i + 1
        ws.cell(row=4, column=col, value=label).font = FONT_BOLD
        ws.cell(row=4, column=col).fill = FILL_HEADER2
        ws.cell(row=4, column=col).font = FONT_HEADER
        ws.cell(row=4, column=col).alignment = CENTER
        ws.cell(row=4, column=col).border = THIN
        c = ws.cell(row=5, column=col, value=value)
        c.font = FONT_KPI
        c.fill = FILL_KPI
        c.alignment = CENTER
        c.border = THIN
    ws.row_dimensions[5].height = 28

    # Status summary table
    ws["A7"] = "Documentation Maturity Summary"
    ws["A7"].font = FONT_SECTION
    ws.merge_cells("A7:D7")

    headers = ["Metric", "Count", "% of Portfolio", "Leadership Signal"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=8, column=i, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = THIN
        cell.alignment = CENTER

    maturity = [
        ("Complete Documentation", complete, f"{round(100*complete/total,1)}%", "Protect & maintain; use as templates"),
        ("Partial Documentation", partial, f"{round(100*partial/total,1)}%", "Close gaps — highest ROI for DevEx"),
        ("Missing / Thin Documentation", missing, f"{round(100*missing/total,1)}%", "Prioritize or confirm N/A with product"),
        ("Products Missing Overview", missing_overview, f"{round(100*missing_overview/total,1)}%", "Blockers for discoverability"),
        ("Products Missing Integration Guide", missing_ig, f"{round(100*missing_ig/total,1)}%", "Primary support-ticket driver"),
        ("Recommend Dedicated Integration Guide", sum(1 for p in products if p.recommend_ig), "—", "Do NOT IG everything — focus list below"),
    ]
    for r_i, row in enumerate(maturity, 9):
        for c_i, val in enumerate(row, 1):
            cell = ws.cell(row=r_i, column=c_i, value=val)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = FONT_BODY
            if c_i == 1 and "Complete" in str(row[0]):
                cell.fill = FILL_GREEN
            elif c_i == 1 and "Partial" in str(row[0]):
                cell.fill = FILL_YELLOW
            elif c_i == 1 and "Missing" in str(row[0]):
                cell.fill = FILL_RED

    # Chart data
    ws["F7"] = "Status Distribution"
    ws["F7"].font = FONT_SECTION
    ws["F8"] = "Status"
    ws["G8"] = "Count"
    style_header_row(ws, 8, 6, 7)
    ws["F9"] = "Complete"
    ws["G9"] = complete
    ws["F10"] = "Partial"
    ws["G10"] = partial
    ws["F11"] = "Missing"
    ws["G11"] = missing
    for r in range(9, 12):
        for c in range(6, 8):
            ws.cell(row=r, column=c).border = THIN
            ws.cell(row=r, column=c).alignment = CENTER
    apply_status_fill(ws["F9"])
    apply_status_fill(ws["F10"])
    apply_status_fill(ws["F11"])

    pie = PieChart()
    pie.title = "Documentation Status Mix"
    labels = Reference(ws, min_col=6, min_row=9, max_row=11)
    data = Reference(ws, min_col=7, min_row=8, max_row=11)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    pie.width = 12
    pie.height = 8
    ws.add_chart(pie, "I7")

    # P0 / P1 / P2 recommendations
    start = 16
    ws.cell(row=start, column=1, value="P0 Recommendations — Must Have Dedicated Integration Guides Immediately").font = FONT_SECTION
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=8)
    ws.cell(row=start, column=1).fill = FILL_P0
    ws.cell(row=start, column=1).font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")

    rec_headers = [
        "Product",
        "Category",
        "Coverage %",
        "Why P0",
        "Expected DevEx Impact",
        "Expected Support Reduction",
        "Expected Integration TAT Impact",
        "Recommended Action",
    ]
    for i, h in enumerate(rec_headers, 1):
        cell = ws.cell(row=start + 1, column=i, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = THIN
        cell.alignment = CENTER

    p0_rationales = {
        "PayU Hosted Checkout (Prebuilt)": (
            "Core payment journey; highest merchant adoption entry path",
            "First-success for majority of new merchants",
            "High — reduces hash/callback/verify tickets",
            "Cuts first payment time by clarifying end-to-end path",
        ),
        "Merchant Hosted Checkout (Custom / Seamless)": (
            "Core journey for customized UX; high complexity across methods",
            "Enables self-serve multi-method integration",
            "High — cards/UPI/NB method confusion",
            "Reduces multi-page scavenger hunt for method docs",
        ),
        "Server-to-Server (S2S) Checkout": (
            "Highest developer complexity; classic/decoupled/direct auth",
            "Critical for advanced merchants; currently fragmented",
            "Very high — S2S errors dominate complex tickets",
            "Decision tree + flow guide shortens S2S TAT significantly",
        ),
        "Checkout Plus (ICP / Bolt Checkout)": (
            "Strategic checkout product with naming confusion (Plus/ICP/Bolt)",
            "Removes ambiguity for modern checkout adopters",
            "Medium-High — naming/support confusion",
            "Faster adoption of Checkout Plus vs wrong product choice",
        ),
        "Payment Links": (
            "High-adoption no-code + API; used across WhatsApp/TPV/partners",
            "Unblocks non-engineering and hybrid integrations",
            "High — link creation/webhook/status tickets",
            "Dashboard+API unified guide accelerates go-live",
        ),
        "Android CheckoutPro SDK": (
            "Primary Android collect path; mobile is core revenue channel",
            "Self-serve Android first payment without SE handholding",
            "High — SDK setup/hash/callback tickets",
            "Sample-app-led IG cuts Android TAT",
        ),
        "iOS CheckoutPro SDK": (
            "Primary iOS collect path; parity with Android required",
            "Parity DevEx for iOS merchants",
            "High — iOS release/privacy/hash issues",
            "Go-live checklist reduces App Store integration delays",
        ),
        "EMI / Cardless EMI": (
            "Affordability core; conversion/revenue impact",
            "Merchants enable EMI without support dependency",
            "Medium-High — EMI codes/eligibility confusion",
            "Hosted/MH/S2S matrix reduces EMI enablement TAT",
        ),
        "Offer Engine / Offers": (
            "Revenue lever; dashboard+API complexity",
            "Self-serve offer create→apply→validate",
            "High — offer validation/refund edge cases",
            "End-to-end offer IG reduces campaign launch delays",
        ),
        "Subscriptions / Recurring Payments": (
            "Core recurring revenue product; SI/Recurring/Subscriptions naming debt",
            "Unifies consent→PDN→recurring mental model",
            "Very high — mandate/PDN/UPI AutoPay tickets",
            "Single narrative cuts recurring integration TAT",
        ),
        "Cross-Border Payments Import (PACB)": (
            "Strategic international revenue; multi-flow (LRS/VA/subs)",
            "Reduces specialist dependency for CB merchants",
            "High — CB onboarding/payment mode tickets",
            "Structured CB IG accelerates import merchant go-live",
        ),
        "Split Settlements (Aggregator / Marketplace)": (
            "Marketplace/aggregator core; onboarding+split complexity",
            "Parent/child settlement self-serve",
            "High — child onboarding/split refund tickets",
            "Marketplace IG reduces aggregator launch TAT",
        ),
        "Tokenization / Save Cards (Vault)": (
            "PCI-reducing foundational capability; recurring dependency",
            "Clear Model 1/2/3 choice prevents wrong vault path",
            "High — tokenization model/support loops",
            "Correct-first-time vault choice shortens TAT",
        ),
        "Third-Party Verification (TPV)": (
            "Regulated/mutual-fund/finance flows; APIs dispersed",
            "Hosted/MH/S2S/Payment Link TPV in one place",
            "Very high — TPV parameter/support dependency",
            "Consolidated TPV IG + API IA cuts finance vertical TAT",
        ),
        "PayU Payouts": (
            "Core disbursement product; multi-feature surface",
            "Single Transfer→webhook→reconcile self-serve",
            "High — beneficiary/transfer status tickets",
            "Master Payouts IG reduces disbursement go-live TAT",
        ),
        "Partner Merchant Onboarding (API / OAuth)": (
            "Partner-sourced merchant growth; duplicate API trees",
            "Partners onboard merchants without PayU ops bottleneck",
            "Very high — KYC/OAuth partner escalations",
            "Canonical partner IG + merged API trees cut partner TAT",
        ),
    }

    r = start + 2
    for p in p0:
        why, devex, support, tat = p0_rationales.get(
            p.name,
            (
                "Core journey / high business impact with documentation gaps",
                "Improves self-serve success",
                "Reduces repetitive support",
                "Shortens integration TAT",
            ),
        )
        vals = [
            p.name,
            p.category,
            p.coverage_score(),
            why,
            devex,
            support,
            tat,
            p.recommended_action,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = FONT_BODY
        r += 1
    p0_end = r

    r += 1
    ws.cell(row=r, column=1, value="P1 Recommendations — Next Wave Documentation Improvements").font = Font(
        name="Calibri", bold=True, size=12
    )
    ws.cell(row=r, column=1).fill = FILL_P1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    for i, h in enumerate(rec_headers, 1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = THIN
    r += 1
    for p in p1:
        vals = [
            p.name,
            p.category,
            p.coverage_score(),
            "High DevEx/business impact; gaps remain after P0 core journeys",
            "Improves advanced/vertical self-serve success",
            "Medium-High reduction in specialty tickets",
            "Meaningful TAT reduction for mid-complexity products",
            p.recommended_action,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = FONT_BODY
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="P2 Recommendations — Improve Later").font = Font(
        name="Calibri", bold=True, size=12, color="FFFFFF"
    )
    ws.cell(row=r, column=1).fill = FILL_P2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    for i, h in enumerate(rec_headers, 1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = THIN
    r += 1
    for p in p2:
        vals = [
            p.name,
            p.category,
            p.coverage_score(),
            "Useful but lower adoption/urgency vs core journeys",
            "Incremental DevEx polish",
            "Moderate/localized support reduction",
            "Smaller TAT impact; schedule after P0/P1",
            p.recommended_action,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = FONT_BODY
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="Key Documentation Gaps").font = FONT_SECTION
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    gaps = [
        "1. Go-Live documentation is rare outside Merchant Hosted/S2S checklists, iOS Custom Browser, partner testing, and plugin checklist — most products lack explicit production readiness guides.",
        "2. Changelog / release notes are largely absent (except limited SDK version history) — developers cannot track breaking changes.",
        "3. Troubleshooting pages are sparse vs FAQs; support dependency remains for error diagnosis.",
        "4. Naming inconsistency increases cognitive load: Collect Payments vs Collect Payment; Checkout Express vs CommercePro; Checkout Plus/ICP/Bolt; Subscriptions/Recurring/SI; Zion/ZION; Pluxee/Sodexo; Mutual Funds/Wealth Tech; MCP dual meaning.",
        "5. Duplicate / parallel surfaces: Integration ASK AI Docs overlaps canonical guides; Partner API trees duplicated; Merchant Hosted API surfaces duplicated; Rewards vs Pay with Rewards vs Flipkart SuperCoins overlap.",
        "6. Server-side SDK docs are thin single pages — insufficient for DevEx goal of zero-support first integration.",
        "7. TPV and Cross-Border API information architecture is dispersed (JSON collections + guide folders) rather than a clean reference IA.",
        "8. Folder typo and editorial residue: Offerings/split-settlments; payouts releasepending- prefixes; ParTner integration casing; internal-review* unpublished partner overview.",
    ]
    for g in gaps:
        ws.cell(row=r, column=1, value=g).alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Overall Observations").font = FONT_SECTION
    r += 1
    observations = [
        "The repository is broad and product-rich: Collect Payments, Offerings, Payouts, Partners, BBPS, WhatsApp, MCP/CLI, and utilities are all represented with substantial page volume (~900+ guide files, ~850+ reference files).",
        "Coverage quality is uneven: core hosted/MH/S2S and several Offerings have strong overview+integration scaffolds, but 'complete' maturity (testing + go-live + troubleshooting + FAQs + errors + webhooks) is uncommon.",
        "Integration Guides should NOT be created for every product. Plugins with install+troubleshooting, dashboard modules, chargebacks ops, and utility APIs are better served by focused how-tos and cross-links.",
        "Highest DevEx ROI is consolidating the core payment journey (Hosted/MH/S2S/Checkout Plus/Payment Links) plus Subscriptions, TPV, Tokenization, Split Settlements, Payouts, and Partner Onboarding into decision-tree-led Integration Guides.",
        "Aligning to the DevEx goal requires treating Integration Guides as merchant journeys (choose → integrate → test → go-live → troubleshoot), not as indexes of API pages.",
        "RECYCLE BIN and Integration ASK AI Docs indicate active restructuring — prioritize promoting reviewed ASK AI content into canonical IA and deleting/redirecting duplicates.",
    ]
    for o in observations:
        ws.cell(row=r, column=1, value=o).alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 36
        r += 1

    autosize(ws, min_width=14, max_width=42)
    ws.column_dimensions["D"].width = 46
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 36
    ws.column_dimensions["G"].width = 36
    ws.column_dimensions["H"].width = 48
    ws.freeze_panes = "A4"

    # =====================================================================
    # Sheet 2 — Product Documentation Inventory
    # =====================================================================
    inv = wb.create_sheet("Product Documentation Inventory")
    inv_headers = list(rows[0].keys())
    for c, h in enumerate(inv_headers, 1):
        cell = inv.cell(row=1, column=c, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN
    inv.row_dimensions[1].height = 40
    inv.freeze_panes = "A2"
    inv.auto_filter.ref = f"A1:{get_column_letter(len(inv_headers))}{len(rows)+1}"

    link_cols = {
        "Overview Page Link",
        "Integration Guide Link",
        "API Reference Link",
        "SDK Link",
        "Quick Start Link",
        "Webhooks Link",
        "Error Codes Link",
        "Testing Guide Link",
        "Go Live Guide Link",
        "Troubleshooting Link",
        "FAQs Link",
        "Changelog Link",
    }
    yn_cols = {
        "Overview Page Exists",
        "Integration Guide Exists",
        "API Reference Exists",
        "SDK Exists",
        "Quick Start Exists",
        "Webhooks Documented",
        "Error Codes Documented",
        "Testing Guide Exists",
        "Go Live Guide Exists",
        "Troubleshooting Exists",
        "FAQs Exists",
        "Changelog Exists",
        "Recommend Dedicated Integration Guide",
    }

    for r_i, row in enumerate(rows, 2):
        for c_i, h in enumerate(inv_headers, 1):
            val = row[h]
            cell = inv.cell(row=r_i, column=c_i)
            cell.border = THIN
            cell.font = FONT_BODY
            if h in link_cols:
                set_hyperlink(cell, val)
            else:
                cell.value = val if val != "" else ("—" if h.endswith("Link") else val)
                cell.alignment = WRAP if h in {"Recommended Action", "Notes", "Product Category", "Product Type"} else CENTER
            if h in yn_cols:
                apply_yn_fill(cell)
            if h == "Documentation Status":
                apply_status_fill(cell)
            if h == "Recommended Priority":
                apply_priority_fill(cell)
            if h == "Documentation Coverage (%)":
                cell.number_format = "0.0"
                cell.alignment = CENTER
                if isinstance(val, (int, float)):
                    if val >= 85:
                        cell.fill = FILL_GREEN
                    elif val >= 40:
                        cell.fill = FILL_YELLOW
                    else:
                        cell.fill = FILL_RED

    autosize(inv, min_width=12, max_width=40)
    for h in ["Recommended Action", "Notes"]:
        idx = inv_headers.index(h) + 1
        inv.column_dimensions[get_column_letter(idx)].width = 48

    # Conditional formatting already applied cell-wise; add score bar chart data sheet later

    # =====================================================================
    # Sheet 3 — Gap Analysis
    # =====================================================================
    gap = wb.create_sheet("Gap Analysis")
    gap["A1"] = "Documentation Gap Analysis"
    gap["A1"].font = FONT_TITLE
    gap["A1"].fill = FILL_TITLE
    gap.merge_cells("A1:F1")
    gap.row_dimensions[1].height = 28

    def write_gap_section(ws, start_row, title, items, headers_local):
        ws.cell(row=start_row, column=1, value=title).font = FONT_SECTION
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers_local))
        for i, h in enumerate(headers_local, 1):
            cell = ws.cell(row=start_row + 1, column=i, value=h)
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.border = THIN
            cell.alignment = CENTER
        rr = start_row + 2
        for item in items:
            for c, v in enumerate(item, 1):
                cell = ws.cell(row=rr, column=c, value=v)
                cell.border = THIN
                cell.alignment = WRAP
                cell.font = FONT_BODY
            rr += 1
        return rr + 1

    no_overview = [
        (
            p.name,
            p.category,
            p.recommended_action or "Create overview: definition, use cases, how it works, next-step CTA.",
        )
        for p in products
        if p.flag("overview") == "No"
    ]
    no_ig = [
        (
            p.name,
            p.category,
            "Yes" if p.recommend_ig else "No",
            p.priority,
            p.recommended_action,
        )
        for p in products
        if p.flag("integration_guide") == "No"
    ]
    no_api = [
        (
            p.name,
            p.category,
            p.product_type,
            "Add API reference map" if "API" in p.product_type or p.applicable.get("api_reference", True) is not False else "N/A — non-API product",
        )
        for p in products
        if p.flag("api_reference") == "No"
    ]
    no_sdk = [
        (p.name, p.category, p.notes or "SDK docs missing or N/A")
        for p in products
        if p.flag("sdk") == "No" and p.product_type.startswith("Integration Channel (SDK)")
    ]

    r = 3
    r = write_gap_section(
        gap,
        r,
        "Products without Overview Pages",
        no_overview or [("None — all inventoried products have an overview/intro path", "—", "—")],
        ["Product", "Category", "Recommendation"],
    )
    r = write_gap_section(
        gap,
        r,
        "Products without Integration Guides (Exists=No)",
        no_ig or [("None", "—", "—", "—", "—")],
        ["Product", "Category", "Recommend Dedicated IG?", "Priority", "Recommendation"],
    )
    r = write_gap_section(
        gap,
        r,
        "Products missing API References (where applicable)",
        no_api,
        ["Product", "Category", "Type", "Recommendation"],
    )
    r = write_gap_section(
        gap,
        r,
        "SDK products missing SDK documentation signals",
        no_sdk or [("None for major SDK products inventoried", "—", "—")],
        ["Product", "Category", "Recommendation"],
    )

    duplicates = [
        (
            "Checkout Plus / ICP / Bolt",
            "docs/Collect Payments/introduction-web/checkout-plus-integration vs Integration ASK AI Docs/checkout-plusicp-checkoutbolt-checkout.md",
            "Canonicalize on Checkout Plus; redirect ASK AI page; retire alternate names in H1s",
        ),
        (
            "CommercePro vs Checkout Express",
            "docs/.../checkout-express (title CommercePro) vs reference/Checkout Express vs plugins/commercepro-checkout.md",
            "Pick one public name; alias the other in intro; align folder/reference names",
        ),
        (
            "Subscriptions / Recurring / SI",
            "Offerings/introduction-recurring-payments-integration + ASK AI subscription page + Pre-debit aggregate API",
            "One IA hub with tabs for Hosted/MH/API/Zion; deprecate duplicate aggregate",
        ),
        (
            "Partner Onboarding API trees",
            "reference/ParTner integration vs reference/Partner Integration - Merchant Onboarding APIs",
            "Merge to one partner onboarding reference; fix ParTner casing",
        ),
        (
            "Merchant Hosted API surfaces",
            "reference/Collect Payment/_payment_merchant_hosted + PayU Merchant Hosted — _payment + General APIs/merchanthostedpostservice",
            "Single MH API landing with deprecation notices on duplicates",
        ),
        (
            "Rewards surfaces",
            "Offerings/rewards-partner-integration + reference/REWARD PARTNERS + Pay with rewards + Flipkart supercoins",
            "Hub page explaining partner programs; link child APIs",
        ),
        (
            "WooCommerce guides",
            "ecommerce-platform-plugins/woocommerce vs Integration ASK AI Docs/woocommerce-payu-plugin-integration-guide.md",
            "Merge unique ASK AI content into canonical plugin guide",
        ),
        (
            "TPV guides",
            "Offerings/introduction-to-payu-tpv vs ASK AI upi-net-banking-tpv + payment-link-tpv guides + root tpv-*.json",
            "Canonical TPV hub + structured API reference",
        ),
        (
            "Payouts introduction",
            "docs/payouts/introduction-to-payouts.md vs custom_pages/payouts-introduction.md",
            "One introduction; remove or redirect custom page duplicate",
        ),
        (
            "Collect Payments naming",
            "docs/Collect Payments vs reference/Collect Payment",
            "Standardize pluralization in nav labels",
        ),
    ]
    r = write_gap_section(
        gap,
        r,
        "Products / Areas with Duplicate Documentation",
        duplicates,
        ["Theme", "Evidence (repo paths)", "Recommendation"],
    )

    inconsistent = [
        (
            "MCP acronym collision",
            "MCP & CLI (Model Context Protocol) vs international payments MCP Lookup (Multi-Currency Pricing)",
            "Always expand acronym on first use; rename MCP Lookup to Multi-Currency Pricing Lookup in titles",
        ),
        (
            "split-settlments typo",
            "docs/Offerings/split-settlments",
            "Rename folder/slug to split-settlements with redirect",
        ),
        (
            "releasepending- payouts residue",
            "docs/payouts/releasepending-pay-to-phone-integration/* and matching reference APIs",
            "Rename to pay-to-phone-*; remove editorial prefix",
        ),
        (
            "Pluxee vs Sodexo",
            "Guide uses Pluxee; reference/Sodexo",
            "Rebrand reference category to Pluxee (Sodexo legacy)",
        ),
        (
            "Mutual Funds vs Wealth Tech",
            "docs/Offerings/mutual-funds-payments vs reference/Wealth Tech",
            "Align category labels; cross-link explicitly",
        ),
        (
            "UPI Bolt channel naming",
            "Ionic vs Capacitor vs Cordova UPI Bolt pages",
            "One UPI Bolt hub with platform tabs",
        ),
        (
            "DBQR naming",
            "Dynamic Storefront QR / DBQR / Offline DBQR / Dynamic Bharat QR",
            "Glossary + canonical product name on overview",
        ),
        (
            "internal-review* partner overview",
            "docs/partners/internal-reviewpartner-integration-overview",
            "Publish as canonical or unhide; remove internal-review prefix",
        ),
    ]
    r = write_gap_section(
        gap,
        r,
        "Inconsistent Documentation / Naming",
        inconsistent,
        ["Issue", "Evidence", "Recommendation"],
    )

    restructure = [
        (
            "Core Web Checkout",
            "introduction-web contains product folders + shared webhooks/errors/testing/FAQs",
            "Keep shared utilities; ensure each checkout product has journey IG that deep-links shared pages",
        ),
        (
            "Integration ASK AI Docs",
            "Parallel hidden guides duplicating canonical topics",
            "Promotion workflow: review → merge → redirect → delete",
        ),
        (
            "Affordability suite",
            "EMI, Offers, BNPL, LazyPay, MobiKwik, Loyalty, Widget under introduction-to-affordability",
            "Hub overview already exists — add decision tree to choose affordability product",
        ),
        (
            "Mobile SDKs",
            "Many sibling SDKs per platform",
            "Platform landing with 'Choose your SDK' matrix (CheckoutPro vs Core vs method SDKs)",
        ),
        (
            "Reference root JSON sprawl",
            "Hundreds of versioned/duplicate Postman/OpenAPI JSON at reference root",
            "Archive superseded collections; keep one current per API family",
        ),
        (
            "RECYCLE BIN",
            "Contains proposed improvements, internal subscription copies, FAQs to be added",
            "Mine for content gaps (e.g., faqs-to-be-added) then keep bin out of public nav",
        ),
    ]
    r = write_gap_section(
        gap,
        r,
        "Products / Areas Requiring Restructuring",
        restructure,
        ["Area", "Current State", "Recommendation"],
    )

    autosize(gap, min_width=16, max_width=50)
    gap.freeze_panes = "A3"

    # =====================================================================
    # Sheet 4 — Integration Guide Prioritization
    # =====================================================================
    igp = wb.create_sheet("IG Prioritization")
    igp["A1"] = "Integration Guide Recommendations & Prioritization"
    igp["A1"].font = FONT_TITLE
    igp["A1"].fill = FILL_TITLE
    igp.merge_cells("A1:K1")
    igp["A2"] = (
        "Rule: Recommend a dedicated Integration Guide only where it provides meaningful DevEx value "
        "(core journey, high complexity, high support dependency, or revenue-critical). "
        "Plugins/utilities/ops products may use install guides or API maps instead. "
        "Priority basis: P0 = must have IG now (core/high-support); P1 = next (high DevEx/vertical impact); "
        "P2 = later; P3 = low / no dedicated IG needed. See 'Why this priority' for each row."
    )
    igp["A2"].alignment = WRAP
    igp.merge_cells("A2:K2")
    igp.row_dimensions[2].height = 48

    ig_headers = [
        "Product Name",
        "Category",
        "Recommend Dedicated IG",
        "Priority",
        "Why this priority",
        "Merchant Adoption Signal",
        "Core Payment Journey",
        "Developer Complexity",
        "Support Dependency Risk",
        "Existing Gap Severity",
        "Rationale / Action",
    ]
    for c, h in enumerate(ig_headers, 1):
        cell = igp.cell(row=4, column=c, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = CENTER
        cell.border = THIN
    igp.row_dimensions[4].height = 36
    igp.freeze_panes = "A5"
    igp.auto_filter.ref = f"A4:K{4+len(products)}"

    def signals(p: Product):
        core = "Yes" if "Core" in p.product_type or p.category.startswith("Collect Payments / Web") or p.name in {
            "Payment Links",
            "PayU Payouts",
            "Subscriptions / Recurring Payments",
            "Split Settlements (Aggregator / Marketplace)",
            "Cross-Border Payments Import (PACB)",
        } else "No"
        complexity = (
            "High"
            if any(x in p.name for x in ["S2S", "Subscriptions", "TPV", "Partner", "Split", "Cross-Border", "Tokenization", "Wallet", "Zion"])
            else ("Medium" if p.recommend_ig else "Low")
        )
        adoption = (
            "High"
            if p.priority == "P0"
            else ("Medium" if p.priority == "P1" else "Lower / Niche")
        )
        support = (
            "High"
            if p.priority in {"P0", "P1"} and p.recommend_ig
            else ("Medium" if p.recommend_ig else "Low")
        )
        gap = (
            "High"
            if p.coverage_score() < 55
            else ("Medium" if p.coverage_score() < 80 else "Low")
        )
        return adoption, core, complexity, support, gap

    def why_ig_priority(p: Product, adoption: str, core: str, complexity: str, support: str, gap: str) -> str:
        """Plain-language reason for the assigned IG priority tier."""
        drivers = []
        if adoption == "High":
            drivers.append("high merchant adoption")
        elif adoption == "Medium":
            drivers.append("medium adoption")
        else:
            drivers.append("lower/niche adoption")

        if core == "Yes":
            drivers.append("on the core payment journey")
        if complexity == "High":
            drivers.append("high developer complexity")
        elif complexity == "Medium":
            drivers.append("moderate complexity")
        if support == "High":
            drivers.append("high support dependency without a clear IG")
        elif support == "Medium":
            drivers.append("moderate support risk")
        if gap == "High":
            drivers.append(f"large doc gap ({p.coverage_score()}% coverage)")
        elif gap == "Medium":
            drivers.append(f"moderate doc gap ({p.coverage_score()}% coverage)")
        else:
            drivers.append(f"docs relatively stronger ({p.coverage_score()}%)")

        if not p.recommend_ig:
            if p.priority in {"P2", "P3"}:
                return (
                    f"{p.priority}: dedicated IG not recommended — "
                    + "; ".join(drivers)
                    + ". Prefer install how-to / API map / cross-links instead."
                )
            return (
                f"{p.priority}: dedicated IG not the primary ask — "
                + "; ".join(drivers)
                + "."
            )

        if p.priority == "P0":
            lead = "P0 — must have a dedicated Integration Guide immediately"
        elif p.priority == "P1":
            lead = "P1 — important next; schedule after P0 core journeys"
        elif p.priority == "P2":
            lead = "P2 — improve later; lower urgency vs core journeys"
        else:
            lead = "P3 — low priority / maintain only"

        return f"{lead} because of " + "; ".join(drivers) + "."

    # Sort: recommend Yes first, then priority
    pri_order = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}
    sorted_products = sorted(
        products,
        key=lambda p: (0 if p.recommend_ig else 1, pri_order.get(p.priority, 9), p.name),
    )
    for r_i, p in enumerate(sorted_products, 5):
        adoption, core, complexity, support, gap_sev = signals(p)
        vals = [
            p.name,
            p.category,
            "Yes" if p.recommend_ig else "No",
            p.priority,
            why_ig_priority(p, adoption, core, complexity, support, gap_sev),
            adoption,
            core,
            complexity,
            support,
            gap_sev,
            p.recommended_action,
        ]
        for c, v in enumerate(vals, 1):
            cell = igp.cell(row=r_i, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = FONT_BODY
        apply_yn_fill(igp.cell(row=r_i, column=3))
        apply_priority_fill(igp.cell(row=r_i, column=4))
        if p.recommend_ig and p.priority in {"P0", "P1"}:
            igp.row_dimensions[r_i].height = 48

    autosize(igp, min_width=12, max_width=42)
    igp.column_dimensions["E"].width = 56
    igp.column_dimensions["K"].width = 55

    # =====================================================================
    # Sheet 5 — Coverage Scoring (with fix-first prioritization)
    # =====================================================================
    meth = wb.create_sheet("Coverage Scoring")
    ranked_fix = sorted(
        products,
        key=lambda p: (-fix_priority_score(p), p.coverage_score(), p.name),
    )

    meth["A1"] = "Coverage Scoring — Fix-First Priority"
    meth["A1"].font = FONT_TITLE
    meth["A1"].fill = FILL_TITLE
    meth.merge_cells("A1:J1")
    meth.row_dimensions[1].height = 28

    meth.merge_cells("A2:J2")
    meth["A2"] = (
        "This sheet ranks products by what to fix first. "
        "Higher Fix Priority Rank / Score = address sooner. "
        "Other workbook sheets are unchanged; only this sheet adds the fix-first order and rationale."
    )
    meth["A2"].font = Font(name="Calibri", italic=True, size=10)
    meth["A2"].alignment = WRAP
    meth.row_dimensions[2].height = 32

    # --- How coverage % is calculated ---
    meth["A4"] = "A. How Coverage % is calculated"
    meth["A4"].font = FONT_SECTION
    methodology_lines = [
        "Equal weight across applicable dimensions (N/A excluded): Overview, Integration Guide, API Reference, SDK, Quick Start, Webhooks, Error Codes, Testing, Go Live, Troubleshooting, FAQs, Changelog.",
        "Yes = 1.0 · Partial = 0.5 · No = 0.0 · Coverage % = 100 × sum(scores) / count(applicable). Status: Complete ≥ 85%; Partial 40–84.9%; Missing < 40%.",
    ]
    for i, line in enumerate(methodology_lines):
        meth.cell(row=5 + i, column=1, value=line).alignment = WRAP
        meth.merge_cells(start_row=5 + i, start_column=1, end_row=5 + i, end_column=10)
        meth.row_dimensions[5 + i].height = 30

    # --- Why ranking / basis ---
    meth["A8"] = "B. Why products are prioritized to fix first (ranking basis)"
    meth["A8"].font = FONT_SECTION
    basis_lines = [
        "Fix Priority Score = Tier weight + Gap severity + Core-journey boost + Integration Guide need + Complexity signal.",
        "1) Tier weight (primary): P0=1000, P1=700, P2=400, P3=100 — from merchant adoption, core payment journey, revenue impact, developer complexity, support dependency, DevEx impact.",
        "2) Gap severity: (100 − Coverage%) × 2 — within the same tier, weaker docs rise so incomplete critical products are fixed before already-strong ones.",
        "3) Core-journey boost (0–50): Hosted/MH/S2S, Payment Links, CheckoutPro, Subscriptions, Payouts, Partner Onboarding, Tokenization, TPV, Split Settlements, Cross-Border, EMI, Offers.",
        "4) Integration Guide need: +30 if a dedicated Integration Guide is recommended.",
        "5) Complexity signal: +20 for S2S / TPV / Partner / Subscriptions / Tokenization / Split / Cross-Border (high support dependency without clear docs).",
    ]
    for i, line in enumerate(basis_lines):
        meth.cell(row=9 + i, column=1, value=line).alignment = WRAP
        meth.merge_cells(start_row=9 + i, start_column=1, end_row=9 + i, end_column=10)
        meth.row_dimensions[9 + i].height = 28

    # --- Fix-first queue ---
    meth["A16"] = "C. Fix-first priority queue (highest rank = fix first)"
    meth["A16"].font = FONT_SECTION

    fix_headers = [
        "Fix Priority Rank",
        "Product Name",
        "Priority Tier",
        "Coverage %",
        "Status",
        "Fix Priority Score",
        "Why fix first (basis)",
        "What to fix",
        "Recommended Action",
        "Recommend Dedicated IG",
    ]
    for c, h in enumerate(fix_headers, 1):
        cell = meth.cell(row=17, column=c, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = CENTER
        cell.border = THIN
    meth.row_dimensions[17].height = 36
    meth.freeze_panes = "A18"

    for i, p in enumerate(ranked_fix, 1):
        r = 17 + i
        vals = [
            i,
            p.name,
            p.priority,
            p.coverage_score(),
            p.status(),
            round(fix_priority_score(p), 1),
            why_fix_first(p),
            what_to_fix(p),
            p.recommended_action,
            "Yes" if p.recommend_ig else "No",
        ]
        for c, v in enumerate(vals, 1):
            cell = meth.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.font = FONT_BODY
            cell.alignment = WRAP if c in {2, 7, 8, 9} else CENTER
        apply_priority_fill(meth.cell(row=r, column=3))
        apply_status_fill(meth.cell(row=r, column=5))
        score_cell = meth.cell(row=r, column=4)
        score_cell.number_format = "0.0"
        if p.coverage_score() >= 85:
            score_cell.fill = FILL_GREEN
        elif p.coverage_score() >= 40:
            score_cell.fill = FILL_YELLOW
        else:
            score_cell.fill = FILL_RED
        apply_yn_fill(meth.cell(row=r, column=10))
        if i <= 15:
            meth.row_dimensions[r].height = 40

    fix_end = 17 + len(ranked_fix)
    meth.auto_filter.ref = f"A17:J{fix_end}"

    # --- Dimension breakdown (same data, ordered by fix priority) ---
    br_start = fix_end + 2
    meth.cell(row=br_start, column=1, value="D. Per-product dimension breakdown (same fix-first order)").font = FONT_SECTION
    break_headers = (
        ["Fix Priority Rank", "Product Name", "Coverage %", "Status"]
        + [d.replace("_", " ").title() for d in DIMENSIONS]
        + ["Applicable Dimensions"]
    )
    for c, h in enumerate(break_headers, 1):
        cell = meth.cell(row=br_start + 1, column=c, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = CENTER
        cell.border = THIN

    for i, p in enumerate(ranked_fix, 1):
        r_i = br_start + 1 + i
        flags = [p.flag(d) for d in DIMENSIONS]
        applicable_n = sum(1 for f in flags if f != "N/A")
        vals = [i, p.name, p.coverage_score(), p.status()] + flags + [applicable_n]
        for c, v in enumerate(vals, 1):
            cell = meth.cell(row=r_i, column=c, value=v)
            cell.border = THIN
            cell.alignment = CENTER if c != 2 else WRAP
            cell.font = FONT_BODY
        apply_status_fill(meth.cell(row=r_i, column=4))
        for c in range(5, 5 + len(DIMENSIONS)):
            apply_yn_fill(meth.cell(row=r_i, column=c))
        score_cell = meth.cell(row=r_i, column=3)
        score_cell.number_format = "0.0"
        if p.coverage_score() >= 85:
            score_cell.fill = FILL_GREEN
        elif p.coverage_score() >= 40:
            score_cell.fill = FILL_YELLOW
        else:
            score_cell.fill = FILL_RED

    autosize(meth, min_width=10, max_width=36)
    meth.column_dimensions["A"].width = 14
    meth.column_dimensions["B"].width = 40
    meth.column_dimensions["G"].width = 48
    meth.column_dimensions["H"].width = 42
    meth.column_dimensions["I"].width = 48

    # =====================================================================
    # Sheet 6 — Repository Analysis Summary
    # =====================================================================
    summary = wb.create_sheet("Repository Analysis")
    summary["A1"] = "Repository Analysis Summary"
    summary["A1"].font = FONT_TITLE
    summary["A1"].fill = FILL_TITLE
    summary.merge_cells("A1:D1")

    summary["A3"] = "Scope Analyzed"
    summary["A3"].font = FONT_SECTION
    scope_rows = [
        ("docs/", "Merchant-facing guides & tutorials (excluding RECYCLE BIN from inventory)"),
        ("reference/", "API reference pages + OpenAPI/Postman collections"),
        ("recipes/", "Code walkthrough recipes supporting checkout/verify flows"),
        ("custom_pages/", "Standalone pages (PG chooser, payouts intro)"),
        ("payment-error-codes.json", "Global payment error code dataset at repo root"),
        ("_order.yaml trees", "Navigation/order metadata used to identify product boundaries"),
    ]
    summary["A4"] = "Area"
    summary["B4"] = "Role in analysis"
    style_header_row(summary, 4, 1, 2)
    for i, (a, b) in enumerate(scope_rows, 5):
        summary.cell(row=i, column=1, value=a).border = THIN
        summary.cell(row=i, column=2, value=b).border = THIN
        summary.cell(row=i, column=2).alignment = WRAP

    summary["A12"] = "Top-Level Documentation IA (from docs/_order.yaml)"
    summary["A12"].font = FONT_SECTION
    ia = [
        "getting started",
        "Collect Payments",
        "Offerings",
        "partners",
        "Whatsapp integration",
        "payouts",
        "MCP & CLI",
        "BBPS",
        "API basics",
        "Monitoring & Alerts",
        "RECYCLE BIN (excluded from product SoT)",
        "Integration ASK AI Docs (shadow/duplicate content)",
        "AIR India (merchant-specific)",
    ]
    summary["A13"] = "Nav Section"
    summary["B13"] = "Inventory Treatment"
    style_header_row(summary, 13, 1, 2)
    treatments = [
        "Onboarding + Dashboard products",
        "Core collect: checkout, no-code, in-person, plugins, SDKs",
        "Value-added payment products & features",
        "Partner program, portal, onboarding & payments APIs",
        "WhatsApp native payments & EPL",
        "Disbursements product family",
        "Agentic/devtooling surfaces",
        "Bill pay / recharge agent APIs",
        "Cross-cutting API auth, hashing, redirect handling",
        "Overwatch monitoring",
        "Excluded from coverage credit",
        "Tracked as duplication risk, not primary coverage",
        "Merchant-specific; low public IG priority",
    ]
    for i, (a, b) in enumerate(zip(ia, treatments), 14):
        summary.cell(row=i, column=1, value=a).border = THIN
        summary.cell(row=i, column=2, value=b).border = THIN

    summary["A29"] = "Reference API Categories Observed"
    summary["A29"].font = FONT_SECTION
    ref_cats = sorted(
        [
            p.name
            for p in Path("/workspace/reference").iterdir()
            if p.is_dir() and p.name not in {"Archive", "ReadMeConfig", "introduction"}
        ]
    )
    summary["A30"] = "API Category Folder"
    style_header_row(summary, 30, 1, 1)
    for i, name in enumerate(ref_cats, 31):
        summary.cell(row=i, column=1, value=name).border = THIN

    # Category rollup chart data
    from collections import Counter

    cat_counts = Counter(p.category.split(" / ")[0] for p in products)
    summary["C29"] = "Products by Top Category"
    summary["C29"].font = FONT_SECTION
    summary["C30"] = "Category"
    summary["D30"] = "Products"
    style_header_row(summary, 30, 3, 4)
    for i, (cat, cnt) in enumerate(sorted(cat_counts.items()), 31):
        summary.cell(row=i, column=3, value=cat).border = THIN
        summary.cell(row=i, column=4, value=cnt).border = THIN

    chart = BarChart()
    chart.type = "col"
    chart.title = "Products by Category"
    n = len(cat_counts)
    data = Reference(summary, min_col=4, min_row=30, max_row=30 + n)
    cats = Reference(summary, min_col=3, min_row=31, max_row=30 + n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 15
    chart.height = 10
    summary.add_chart(chart, "F29")

    summary["A50"] = "How to Use This Tracker"
    summary["A50"].font = FONT_SECTION
    how_to = [
        "1. Leadership reviews Executive Dashboard KPIs and P0 list weekly/biweekly.",
        "2. Docs team filters Product Documentation Inventory by Status=Partial/Missing and Priority=P0/P1.",
        "3. Use IG Prioritization to assign writing projects — only where Recommend Dedicated IG=Yes.",
        "4. Gap Analysis drives IA cleanup (duplicates, naming, restructures) in parallel with new IG content.",
        "5. Re-run generate_coverage_tracker.py after major docs releases to refresh scores from the repository.",
        "6. Do not treat this file's historical sample Excel as data input — repository remains the source of truth.",
    ]
    for i, line in enumerate(how_to, 51):
        summary.cell(row=i, column=1, value=line).alignment = WRAP
        summary.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)

    autosize(summary, min_width=18, max_width=60)
    summary.freeze_panes = "A3"

    # =====================================================================
    # Sheet 7 — Executive Recommendations (detailed)
    # =====================================================================
    exe = wb.create_sheet("Executive Recommendations")
    exe["A1"] = "Executive Recommendations — P0 / P1 / P2"
    exe["A1"].font = FONT_TITLE
    exe["A1"].fill = FILL_TITLE
    exe.merge_cells("A1:G1")
    exe["A2"] = (
        "Aligned to PayU DevEx goal: Enable developers to successfully integrate PayU products without requiring support intervention. "
        "Priorities balance merchant adoption, core payment journey criticality, business/revenue impact, developer complexity, support dependency, and documentation gaps."
    )
    exe["A2"].alignment = WRAP
    exe.merge_cells("A2:G2")
    exe.row_dimensions[2].height = 40

    def write_exec_block(ws, start, title, fill, plist, default_rationale):
        ws.cell(row=start, column=1, value=title).fill = fill
        ws.cell(row=start, column=1).font = Font(name="Calibri", bold=True, size=12, color="FFFFFF" if fill != FILL_P1 else "000000")
        ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=7)
        headers_e = [
            "Product",
            "Coverage %",
            "Why this priority",
            "DevEx Impact",
            "Support Dependency Reduction",
            "Integration TAT Impact",
            "Recommended Action",
        ]
        for c, h in enumerate(headers_e, 1):
            cell = ws.cell(row=start + 1, column=c, value=h)
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.border = THIN
            cell.alignment = CENTER
        r = start + 2
        for p in plist:
            why, devex, support, tat = default_rationale(p)
            for c, v in enumerate(
                [p.name, p.coverage_score(), why, devex, support, tat, p.recommended_action],
                1,
            ):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = THIN
                cell.alignment = WRAP
                cell.font = FONT_BODY
            ws.row_dimensions[r].height = 48
            r += 1
        return r + 2

    def p0_r(p):
        return p0_rationales.get(
            p.name,
            (
                "Critical product with material gaps or fragmentation",
                "High — unlocks self-serve success",
                "High — removes repeat support loops",
                "High — shortens time-to-first-successful-integration",
            ),
        )

    def p1_r(p):
        return (
            "Important after core journeys; complexity or vertical revenue impact",
            "Medium-High — improves specialty self-serve",
            "Medium-High — reduces specialty escalations",
            "Medium — improves TAT for mid-complexity integrations",
        )

    def p2_r(p):
        return (
            "Valuable polish / niche adoption; schedule after P0/P1",
            "Medium — incremental clarity",
            "Medium/Low — localized ticket reduction",
            "Lower — smaller portfolio TAT impact",
        )

    r = 4
    r = write_exec_block(exe, r, "P0 — MUST have dedicated Integration Guide immediately", FILL_P0, p0, p0_r)
    r = write_exec_block(exe, r, "P1 — Important documentation improvements next", FILL_P1, p1, p1_r)
    r = write_exec_block(exe, r, "P2 — Improve later", FILL_P2, p2, p2_r)

    exe.cell(row=r, column=1, value="Portfolio Recommendations (Cross-Cutting)").font = FONT_SECTION
    r += 1
    cross = [
        "Establish an Integration Guide template: Overview → Choose integration → Prerequisites → Step-by-step → Test/Sandbox → Go-Live → Webhooks → Errors → Troubleshooting → FAQs → API map.",
        "Create a public 'Choose your integration' decision tree that routes Hosted vs MH vs S2S vs Checkout Plus vs Plugins vs SDKs (partial drafts exist under getting started/choose-your-payment-gateway).",
        "Institute a quarterly coverage refresh by re-running this tracker against main/v1.",
        "Assign owners per P0 product for Integration Guide delivery; track to Complete (≥85%) status.",
        "Run an IA cleanup sprint for naming + duplicates in parallel with P0 writing — otherwise new IGs will inherit confusion.",
        "Promote go-live checklists from MH/S2S/iOS CB as the standard artifact every P0/P1 IG must include.",
    ]
    for line in cross:
        exe.cell(row=r, column=1, value=line).alignment = WRAP
        exe.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        exe.row_dimensions[r].height = 30
        r += 1

    autosize(exe, min_width=14, max_width=40)
    exe.column_dimensions["C"].width = 42
    exe.column_dimensions["G"].width = 48
    exe.freeze_panes = "A4"

    # =====================================================================
    # Sheet 8 — Coverage by Category (helper)
    # =====================================================================
    cat = wb.create_sheet("Coverage by Category")
    cat["A1"] = "Coverage Rollup by Product Category"
    cat["A1"].font = FONT_TITLE
    cat["A1"].fill = FILL_TITLE
    cat.merge_cells("A1:G1")
    cat_headers = [
        "Category",
        "Products",
        "Avg Coverage %",
        "Complete",
        "Partial",
        "Missing",
        "P0 IG Recs",
    ]
    for c, h in enumerate(cat_headers, 1):
        cell = cat.cell(row=3, column=c, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = THIN
    from collections import defaultdict

    buckets = defaultdict(list)
    for p in products:
        buckets[p.category].append(p)
    for r_i, (category, plist) in enumerate(sorted(buckets.items()), 4):
        vals = [
            category,
            len(plist),
            round(sum(p.coverage_score() for p in plist) / len(plist), 1),
            sum(1 for p in plist if p.status() == "Complete"),
            sum(1 for p in plist if p.status() == "Partial"),
            sum(1 for p in plist if p.status() == "Missing"),
            sum(1 for p in plist if p.priority == "P0" and p.recommend_ig),
        ]
        for c, v in enumerate(vals, 1):
            cell = cat.cell(row=r_i, column=c, value=v)
            cell.border = THIN
            cell.alignment = CENTER if c > 1 else WRAP
            cell.font = FONT_BODY
        sc = cat.cell(row=r_i, column=3)
        if vals[2] >= 85:
            sc.fill = FILL_GREEN
        elif vals[2] >= 40:
            sc.fill = FILL_YELLOW
        else:
            sc.fill = FILL_RED
    cat.freeze_panes = "A4"
    cat.auto_filter.ref = f"A3:G{3+len(buckets)}"
    autosize(cat)

    return wb


def main():
    products = build_products()
    # Sanity: verify flagged Yes paths exist
    problems = []
    for p in products:
        for dim in DIMENSIONS:
            val = getattr(p, dim)
            if val and val not in {"Yes", "No", "Partial", "Shared", "N/A"}:
                if not (ROOT / val).exists():
                    problems.append(f"{p.name} :: {dim} :: {val}")
    if problems:
        print("WARNING: missing paths detected:")
        for pr in problems:
            print(" -", pr)
    else:
        print("All explicit paths resolve.")

    wb = build_workbook(products)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_FILE)
    print(f"Wrote {OUT_FILE}")
    print(f"Products: {len(products)}")
    complete = sum(1 for p in products if p.status() == "Complete")
    partial = sum(1 for p in products if p.status() == "Partial")
    missing = sum(1 for p in products if p.status() == "Missing")
    print(f"Status Complete/Partial/Missing: {complete}/{partial}/{missing}")
    print(
        f"Overall avg coverage: {round(sum(p.coverage_score() for p in products)/len(products),1)}%"
    )
    print(
        f"P0/P1/P2 IG recs: {sum(1 for p in products if p.priority=='P0' and p.recommend_ig)}/"
        f"{sum(1 for p in products if p.priority=='P1' and p.recommend_ig)}/"
        f"{sum(1 for p in products if p.priority=='P2' and p.recommend_ig)}"
    )


if __name__ == "__main__":
    main()