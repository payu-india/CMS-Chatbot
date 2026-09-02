#!/usr/bin/env python3
"""
PayU Product Docs — Leadership Coverage Report (combined).

Combines:
  1) Progress Report checklist coverage (source of truth for Present/Absent scores)
  2) Tracker-style prioritization (what to chase, why, actions)

Output: a lean VP-ready workbook — where we stand, what to improve,
priorities, why, and action items. No unnecessary sheets.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from generate_progress_report import NA, ProductBlock, build_products

OUT = Path("/workspace/docs-coverage-tracker/PayU_Product_Docs_Leadership_Coverage_Report.xlsx")

# Short labels for checklist items (Progress Report dimensions)
SHORT = {
    "Introduction Page - Definition, Examples, Features and Advantages": "Intro",
    "Use Cases (Top 3-5 Verticals)": "Use Cases",
    "How it Works (Desc. + Process Diagram)": "How it Works",
    "Product Video": "Video",
    "Product Videos": "Video",
    "Feature Video": "Video",
    "Videos": "Video",
    "All Dashboard Action (&Video)": "Dashboard",
    "Other Features and Capabilties": "Features",
    "API Page": "API",
    "Webhooks": "Webhooks",
    "FAQs": "FAQs",
    "Glossary": "Glossary",
    "Change Log": "Changelog",
    "S2S/Hosted/Custom [if applicable]": "S2S/Hosted",
    "S2S/Hosted/Custom": "S2S/Hosted",
    "Sign up Process is explained": "Sign-up flow",
    "All steps are documented": "All steps",
    "All required documents are documented": "Documents",
    "Descriptions and How Tos": "How-tos",
    "Flows": "Flows",
    "Integrations": "Integrations",
    "Features": "Features",
}

# Explicit leadership priorities (aligned to tracker P0/P1 intent)
PRIORITY = {
    # P0 — chase now
    "PayU Hosted Checkout (Prebuilt)": "P0",
    "Merchant Hosted Checkout (Custom / Seamless)": "P0",
    "Server-to-Server (S2S) Checkout": "P0",
    "Checkout Plus (ICP / Bolt Checkout)": "P0",
    "Payment Links": "P0",
    "Android CheckoutPro SDK": "P0",
    "iOS CheckoutPro SDK": "P0",
    "EMI / Cardless EMI": "P0",
    "Offer Engine / Offers": "P0",
    "Subscriptions / Recurring Payments": "P0",
    "Cross-Border Payments Import (PACB)": "P0",
    "Split Settlements (Aggregator / Marketplace)": "P0",
    "Tokenization / Save Cards (Vault)": "P0",
    "Third-Party Verification (TPV)": "P0",
    "PayU Payouts": "P0",
    "Partner Merchant Onboarding (API / OAuth)": "P0",
    # P1 — next
    "CommercePro Checkout (Checkout Express)": "P1",
    "UPI QR": "P1",
    "Dynamic Storefront QR (DBQR)": "P1",
    "POS Terminal Integration": "P1",
    "Android POS SDK": "P1",
    "Shopify Plugin": "P1",
    "WooCommerce Plugin": "P1",
    "PHP SDK": "P1",
    "Node.js SDK": "P1",
    "Java SDK": "P1",
    "Android Core SDK": "P1",
    "iOS Core SDK": "P1",
    "React Native CheckoutPro SDK": "P1",
    "Flutter CheckoutPro SDK": "P1",
    "UPI Bolt SDK (Cross-Platform)": "P1",
    "Affordability Widget": "P1",
    "BNPL / Pay Later": "P1",
    "Zion Subscription Automation": "P1",
    "Dynamic Currency Conversion / International Payments": "P1",
    "Pre-Authorize / Auth & Capture": "P1",
    "UPI One-Time Mandate (OTM / Reserve Pay)": "P1",
    "Native OTP Flow": "P1",
    "Apple Pay": "P1",
    "Mutual Fund Payments (WealthTech)": "P1",
    "Merchant Wallet": "P1",
    "Smart Send": "P1",
    "Pay to Phone": "P1",
    "Partner Payments Integration": "P1",
    "WhatsApp Payments": "P1",
    "BBPS Connect Agent": "P1",
    "PayU Remote MCP Server": "P1",
    "Agentic Commerce Suite": "P1",
    "Refunds": "P1",
}

ACTIONS = {
    "PayU Hosted Checkout (Prebuilt)": "Ship end-to-end Hosted IG: hash → pay → callback → verify → go-live; add video + use cases.",
    "Merchant Hosted Checkout (Custom / Seamless)": "One journey IG stitching payment methods, hashing, webhooks, verify, go-live.",
    "Server-to-Server (S2S) Checkout": "Unified S2S IG with classic/decoupled/direct-auth decision tree + errors.",
    "Checkout Plus (ICP / Bolt Checkout)": "Canonicalize Plus/ICP/Bolt naming; expand testing, go-live, troubleshooting.",
    "Payment Links": "Dashboard + API + webhooks + WhatsApp/TPV variants in one IG; go-live checklist.",
    "Android CheckoutPro SDK": "Sample-app-led IG; keep troubleshooting/changelog current.",
    "iOS CheckoutPro SDK": "Parity with Android; explicit go-live + release notes.",
    "EMI / Cardless EMI": "EMI IG covering Hosted/MH/S2S + NTB; link Affordability hub.",
    "Offer Engine / Offers": "Offer create → apply → validate → refund journey; FAQs + changelog.",
    "Subscriptions / Recurring Payments": "Unify SI/Recurring/Subscriptions; consent→PDN→recurring + UPI AutoPay/eNACH.",
    "Cross-Border Payments Import (PACB)": "CB IG for card/UPI/NB, LRS, VA; standardize CB/PACB naming.",
    "Split Settlements (Aggregator / Marketplace)": "Marketplace IG: child onboarding → split → refunds; fix folder typo.",
    "Tokenization / Save Cards (Vault)": "Models 1–3 + push tokenization IG (PCI-reducing foundation).",
    "Third-Party Verification (TPV)": "Consolidate Hosted/MH/S2S/Payment Link TPV + dispersed APIs into one hub.",
    "PayU Payouts": "Master Payouts IG (transfer, beneficiary, webhooks, go-live); clean Pay-to-Phone paths.",
    "Partner Merchant Onboarding (API / OAuth)": "Publish canonical partner IG; merge duplicate API trees.",
}

FILL_HEADER = PatternFill("solid", fgColor="0B3D5C")
FILL_TITLE = PatternFill("solid", fgColor="072A40")
FILL_KPI = PatternFill("solid", fgColor="E8F4FC")
FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_RED = PatternFill("solid", fgColor="FFC7CE")
FILL_GREY = PatternFill("solid", fgColor="D9D9D9")
FILL_P0 = PatternFill("solid", fgColor="C00000")
FILL_P1 = PatternFill("solid", fgColor="FFC000")
FILL_P2 = PatternFill("solid", fgColor="548235")
FILL_BLUE = PatternFill("solid", fgColor="DDEBF7")
FILL_ALT = PatternFill("solid", fgColor="F7F9FB")

FONT_TITLE = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_SECTION = Font(name="Calibri", bold=True, color="0B3D5C", size=12)
FONT_KPI = Font(name="Calibri", bold=True, color="0B3D5C", size=18)
FONT_BODY = Font(name="Calibri", size=10)
FONT_BOLD = Font(name="Calibri", bold=True, size=10)
FONT_WHITE_BOLD = Font(name="Calibri", bold=True, size=10, color="FFFFFF")

THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def as_num(v):
    if v in (1, "1"):
        return 1
    if v in (0, "0"):
        return 0
    return None


def coverage_stats(p: ProductBlock):
    applicable = 0
    score = 0
    present, absent, partial_na = [], [], []
    for label, val, _rec in p.items:
        short = SHORT.get(label, label)
        n = as_num(val)
        if n is None:
            partial_na.append(short)
            continue
        applicable += 1
        if n == 1:
            score += 1
            present.append(short)
        else:
            absent.append(short)
    pct = round(100.0 * score / applicable, 1) if applicable else 0.0
    if pct >= 85:
        maturity = "Complete"
    elif pct < 40:
        maturity = "Critical gaps"
    else:
        maturity = "Partial"
    return applicable, score, pct, maturity, present, absent


def priority_for(name: str) -> str:
    return PRIORITY.get(name, "P2")


def why_priority(name: str, pct: float, absent: list[str], priority: str) -> str:
    bits = []
    if priority == "P0":
        bits.append("P0: core payment/revenue journey or high support dependency")
    elif priority == "P1":
        bits.append("P1: high DevEx / vertical impact after core journeys")
    else:
        bits.append("P2: improve later / lower urgency")
    if pct < 55:
        bits.append(f"large coverage gap ({pct}%)")
    elif pct < 75:
        bits.append(f"moderate coverage gap ({pct}%)")
    else:
        bits.append(f"coverage relatively stronger ({pct}%)")
    if absent:
        top = ", ".join(absent[:4])
        bits.append(f"missing: {top}")
    return "; ".join(bits)


def action_for(p: ProductBlock, absent: list[str]) -> str:
    if p.name in ACTIONS:
        return ACTIONS[p.name]
    if not absent:
        return "Maintain quality; keep changelog/FAQs current."
    # Use first progress-report recommendation if any
    for _label, val, rec in p.items:
        if as_num(val) == 0 and rec:
            return rec
    return f"Add missing content: {', '.join(absent[:5])}."


def tier_weight(pri: str) -> int:
    return {"P0": 1000, "P1": 700, "P2": 400, "P3": 100}.get(pri, 0)


def rank_score(name: str, pct: float, priority: str) -> float:
    core_boost = 40 if priority == "P0" else (20 if priority == "P1" else 0)
    return tier_weight(priority) + (100 - pct) * 2 + core_boost


def display(val) -> str:
    n = as_num(val)
    if n == 1:
        return "Present"
    if n == 0:
        return "Absent"
    return "N/A"


def apply_present_fill(cell, text: str):
    if text == "Present":
        cell.fill = FILL_GREEN
    elif text == "Absent":
        cell.fill = FILL_RED
    else:
        cell.fill = FILL_GREY
    cell.alignment = CENTER
    cell.border = THIN
    cell.font = FONT_BODY


def apply_maturity(cell, v: str):
    if v == "Complete":
        cell.fill = FILL_GREEN
    elif v == "Partial":
        cell.fill = FILL_YELLOW
    else:
        cell.fill = FILL_RED
    cell.alignment = CENTER
    cell.border = THIN
    cell.font = FONT_BODY


def apply_priority(cell, v: str):
    if v == "P0":
        cell.fill = FILL_P0
        cell.font = FONT_WHITE_BOLD
    elif v == "P1":
        cell.fill = FILL_P1
        cell.font = FONT_BOLD
    else:
        cell.fill = FILL_P2
        cell.font = FONT_WHITE_BOLD
    cell.alignment = CENTER
    cell.border = THIN


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = CENTER
        cell.border = THIN


def build_report(products: list[ProductBlock]) -> Workbook:
    rows = []
    for p in products:
        applicable, score, pct, maturity, present, absent = coverage_stats(p)
        pri = priority_for(p.name)
        rows.append(
            {
                "product": p,
                "applicable": applicable,
                "docs_score": score,
                "pct": pct,
                "maturity": maturity,
                "present": present,
                "absent": absent,
                "priority": pri,
                "why": why_priority(p.name, pct, absent, pri),
                "action": action_for(p, absent),
                "rank": rank_score(p.name, pct, pri),
            }
        )

    ranked = sorted(rows, key=lambda r: (-r["rank"], r["pct"], r["product"].name))
    total = len(rows)
    port_applicable = sum(r["applicable"] for r in rows)
    port_score = sum(r["docs_score"] for r in rows)
    overall = round(100.0 * port_score / port_applicable, 1) if port_applicable else 0.0
    complete = sum(1 for r in rows if r["maturity"] == "Complete")
    partial = sum(1 for r in rows if r["maturity"] == "Partial")
    critical = sum(1 for r in rows if r["maturity"] == "Critical gaps")
    p0 = [r for r in ranked if r["priority"] == "P0"]
    p1 = [r for r in ranked if r["priority"] == "P1"]

    # Portfolio checklist rollup (standard items only)
    checklist_keys = [
        "Intro",
        "Use Cases",
        "How it Works",
        "Video",
        "Dashboard",
        "Features",
        "API",
        "Webhooks",
        "FAQs",
        "Glossary",
        "Changelog",
        "S2S/Hosted",
    ]
    checklist_present = {k: 0 for k in checklist_keys}
    checklist_applicable = {k: 0 for k in checklist_keys}
    for r in rows:
        for label, val, _ in r["product"].items:
            short = SHORT.get(label, None)
            if short not in checklist_applicable:
                continue
            n = as_num(val)
            if n is None:
                continue
            checklist_applicable[short] += 1
            if n == 1:
                checklist_present[short] += 1

    wb = Workbook()

    # =====================================================================
    # Sheet 1 — Executive Dashboard (lean, numbers from Progress Report)
    # =====================================================================
    ws = wb.active
    ws.title = "Executive Dashboard"

    ws.merge_cells("A1:H1")
    ws["A1"] = "PayU Product Documentation — Leadership Coverage Report"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Coverage source: Progress Report checklist (Intro, Use Cases, How it Works, Video, Dashboard, "
        f"Features, API, Webhooks, FAQs, Glossary, Changelog, S2S/Hosted). "
        f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}. "
        f"Goal: developers integrate without support."
    )
    ws["A2"].font = Font(name="Calibri", italic=True, size=10)
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 36

    ws.merge_cells("A3:H3")
    ws["A3"] = (
        "LEGEND: Present/Absent = whether that documentation type exists for the product. "
        "Complete ≥85% · Partial 40–84.9% · Critical gaps <40% of applicable checklist items. "
        "N/A items are excluded from coverage %."
    )
    ws["A3"].fill = FILL_BLUE
    ws["A3"].font = Font(name="Calibri", size=10, color="0B3D5C")
    ws["A3"].alignment = WRAP
    ws.row_dimensions[3].height = 36

    kpis = [
        ("Total Products", total),
        ("Overall Coverage %", f"{overall}%"),
        ("Complete", complete),
        ("Partial", partial),
        ("Critical gaps", critical),
        ("P0 priorities", len(p0)),
        ("P1 priorities", len(p1)),
        ("Checklist items scored", f"{port_score}/{port_applicable}"),
    ]
    for i, (label, val) in enumerate(kpis):
        c = i + 1
        ws.cell(row=5, column=c, value=label).font = FONT_HEADER
        ws.cell(row=5, column=c).fill = FILL_HEADER
        ws.cell(row=5, column=c).alignment = CENTER
        ws.cell(row=5, column=c).border = THIN
        cell = ws.cell(row=6, column=c, value=val)
        cell.font = FONT_KPI
        cell.fill = FILL_KPI
        cell.alignment = CENTER
        cell.border = THIN
    ws.row_dimensions[6].height = 28

    # Where we stand
    ws["A8"] = "1. Where we stand"
    ws["A8"].font = FONT_SECTION
    stand = [
        f"Portfolio documentation coverage is {overall}% across {total} products ({port_score} of {port_applicable} applicable checklist items Present).",
        f"Maturity mix: {complete} Complete · {partial} Partial · {critical} Critical gaps.",
        "Strongest patterns: Introduction / How it Works / API pages are often Present for core products.",
        "Weakest patterns: Product Video, Glossary, Changelog, and often Use Cases / Go-live-style readiness content are widely Absent.",
        "This is a documentation maturity gap — products exist; required developer content is incomplete.",
    ]
    for i, line in enumerate(stand):
        ws.cell(row=9 + i, column=1, value=f"• {line}").alignment = WRAP
        ws.merge_cells(start_row=9 + i, start_column=1, end_row=9 + i, end_column=8)
        ws.row_dimensions[9 + i].height = 28

    # What we are chasing
    ws["A15"] = "2. What we are chasing"
    ws["A15"].font = FONT_SECTION
    chase = [
        "Close Critical gaps and lift Partial products on P0 journeys to Complete (≥85%).",
        "Dedicated Integration Guides only for high-adoption / high-complexity products (not every product).",
        "Fill the portfolio-wide Absents that block self-serve: Video (where useful), Use Cases, Glossary, Changelog, FAQs/Troubleshooting depth.",
        "Remove naming/duplicate noise on Checkout Plus/Bolt, CommercePro/Express, Partner API trees so merchants find one path.",
    ]
    for i, line in enumerate(chase):
        ws.cell(row=16 + i, column=1, value=f"• {line}").alignment = WRAP
        ws.merge_cells(start_row=16 + i, start_column=1, end_row=16 + i, end_column=8)
        ws.row_dimensions[16 + i].height = 28

    # Top 10 priorities table
    ws["A21"] = "3. Top priorities to fix first"
    ws["A21"].font = FONT_SECTION
    top_h = ["Rank", "Product", "Priority", "Coverage %", "Maturity", "What's absent", "Why prioritize", "Action"]
    for c, h in enumerate(top_h, 1):
        ws.cell(row=22, column=c, value=h)
    style_header(ws, 22, 8)

    for i, r in enumerate(ranked[:12], 1):
        rr = 22 + i
        vals = [
            i,
            r["product"].name,
            r["priority"],
            r["pct"],
            r["maturity"],
            ", ".join(r["absent"][:5]) if r["absent"] else "—",
            r["why"],
            r["action"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=rr, column=c, value=v)
            cell.border = THIN
            cell.font = FONT_BODY
            cell.alignment = WRAP if c in {2, 6, 7, 8} else CENTER
        apply_priority(ws.cell(row=rr, column=3), r["priority"])
        apply_maturity(ws.cell(row=rr, column=5), r["maturity"])
        ws.row_dimensions[rr].height = 42

    # Portfolio checklist health
    ws["A36"] = "4. Portfolio checklist health (% Present where applicable)"
    ws["A36"].font = FONT_SECTION
    ws["A37"] = "Checklist item"
    ws["B37"] = "% Present"
    ws["C37"] = "Present / Applicable"
    style_header(ws, 37, 3)
    for i, key in enumerate(checklist_keys):
        r = 38 + i
        app = checklist_applicable[key]
        pre = checklist_present[key]
        pct = round(100.0 * pre / app, 1) if app else None
        ws.cell(row=r, column=1, value=key).border = THIN
        cell = ws.cell(row=r, column=2, value=f"{pct}%" if pct is not None else "—")
        cell.border = THIN
        cell.alignment = CENTER
        if pct is None:
            cell.fill = FILL_GREY
        elif pct >= 85:
            cell.fill = FILL_GREEN
        elif pct >= 40:
            cell.fill = FILL_YELLOW
        else:
            cell.fill = FILL_RED
        ws.cell(row=r, column=3, value=f"{pre}/{app}" if app else "—").border = THIN
        ws.cell(row=r, column=3).alignment = CENTER

    # Maturity chart data
    ws["E37"] = "Maturity"
    ws["F37"] = "Count"
    style_header(ws, 37, 6)
    # overwrite E37-F37 properly
    for c, h in enumerate(["Maturity", "Count"], 5):
        cell = ws.cell(row=37, column=c, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = CENTER
        cell.border = THIN
    ws["E38"] = "Complete"
    ws["F38"] = complete
    ws["E39"] = "Partial"
    ws["F39"] = partial
    ws["E40"] = "Critical gaps"
    ws["F40"] = critical
    for r in range(38, 41):
        for c in range(5, 7):
            ws.cell(row=r, column=c).border = THIN
            ws.cell(row=r, column=c).alignment = CENTER
        apply_maturity(ws.cell(row=r, column=5), ws.cell(row=r, column=5).value)

    chart = DoughnutChart()
    chart.title = "Maturity mix"
    labels = Reference(ws, min_col=5, min_row=38, max_row=40)
    data = Reference(ws, min_col=6, min_row=37, max_row=40)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.width = 12
    chart.height = 8
    ws.add_chart(chart, "E42")

    for col, width in {
        "A": 10,
        "B": 38,
        "C": 10,
        "D": 12,
        "E": 14,
        "F": 28,
        "G": 42,
        "H": 48,
    }.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A5"

    # =====================================================================
    # Sheet 2 — Coverage Snapshot (Present / Absent from Excel 1)
    # =====================================================================
    cov = wb.create_sheet("Coverage Snapshot")
    cov.merge_cells("A1:R1")
    cov["A1"] = "Coverage Snapshot — what is Present vs Absent (Progress Report checklist)"
    cov["A1"].font = FONT_TITLE
    cov["A1"].fill = FILL_TITLE
    cov.row_dimensions[1].height = 28

    cov.merge_cells("A2:R2")
    cov["A2"] = (
        "Coverage % recalculated from Progress Report scores only. "
        "Present=1, Absent=0, N/A excluded. Sorted by leadership priority then lowest coverage."
    )
    cov["A2"].font = Font(name="Calibri", italic=True, size=10)
    cov["A2"].alignment = WRAP

    # Dynamic short columns used across products
    used_shorts = []
    for key in checklist_keys:
        used_shorts.append(key)
    # also include signup-specific if needed
    for extra in ["Sign-up flow", "All steps", "Documents", "How-tos", "Flows", "Integrations"]:
        if any(
            SHORT.get(label, label) == extra
            for r in rows
            for label, _v, _rec in r["product"].items
        ):
            if extra not in used_shorts:
                used_shorts.append(extra)

    headers = (
        ["Product", "Priority", "Coverage %", "Maturity", "Docs Score", "Applicable"]
        + used_shorts
        + ["What's absent", "Action"]
    )
    for c, h in enumerate(headers, 1):
        cov.cell(row=4, column=c, value=h)
    style_header(cov, 4, len(headers))
    cov.row_dimensions[4].height = 32
    cov.freeze_panes = "A5"
    cov.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{4 + len(ranked)}"

    for i, r in enumerate(ranked, 1):
        rr = 4 + i
        # map short -> value
        short_map = {}
        for label, val, _rec in r["product"].items:
            short_map[SHORT.get(label, label)] = display(val)

        vals = [
            r["product"].name,
            r["priority"],
            r["pct"],
            r["maturity"],
            r["docs_score"],
            r["applicable"],
        ]
        for key in used_shorts:
            vals.append(short_map.get(key, "N/A"))
        vals.append(", ".join(r["absent"]) if r["absent"] else "—")
        vals.append(r["action"])

        for c, v in enumerate(vals, 1):
            cell = cov.cell(row=rr, column=c, value=v)
            cell.border = THIN
            cell.font = FONT_BODY
            cell.alignment = WRAP if c in {1, len(headers) - 1, len(headers)} else CENTER
        apply_priority(cov.cell(row=rr, column=2), r["priority"])
        apply_maturity(cov.cell(row=rr, column=4), r["maturity"])
        pct_cell = cov.cell(row=rr, column=3)
        if r["pct"] >= 85:
            pct_cell.fill = FILL_GREEN
        elif r["pct"] >= 40:
            pct_cell.fill = FILL_YELLOW
        else:
            pct_cell.fill = FILL_RED
        # present/absent cells
        for c in range(7, 7 + len(used_shorts)):
            apply_present_fill(cov.cell(row=rr, column=c), cov.cell(row=rr, column=c).value)

    cov.column_dimensions["A"].width = 40
    cov.column_dimensions["B"].width = 10
    for c in range(3, 7):
        cov.column_dimensions[get_column_letter(c)].width = 11
    for c in range(7, 7 + len(used_shorts)):
        cov.column_dimensions[get_column_letter(c)].width = 11
    cov.column_dimensions[get_column_letter(len(headers) - 1)].width = 36
    cov.column_dimensions[get_column_letter(len(headers))].width = 46

    # =====================================================================
    # Sheet 3 — Priorities & Action Items
    # =====================================================================
    act = wb.create_sheet("Priorities & Action Items")
    act.merge_cells("A1:G1")
    act["A1"] = "Priorities & Action Items — what leadership should chase"
    act["A1"].font = FONT_TITLE
    act["A1"].fill = FILL_TITLE

    act.merge_cells("A2:G2")
    act["A2"] = (
        "Priority basis: P0 = core payment/revenue / high support dependency — do now. "
        "P1 = next wave DevEx/vertical impact. P2 = later. "
        "Within a tier, lower Progress Report coverage % is chased first. "
        "Dedicated Integration Guides only where they materially reduce support."
    )
    act["A2"].alignment = WRAP
    act["A2"].font = Font(name="Calibri", italic=True, size=10)
    act.row_dimensions[2].height = 40

    # P0 section
    act["A4"] = "P0 — Must chase now"
    act["A4"].font = FONT_WHITE_BOLD
    act["A4"].fill = FILL_P0
    act.merge_cells("A4:G4")

    act_headers = [
        "Rank",
        "Product",
        "Coverage %",
        "Maturity",
        "What's absent (from Progress Report)",
        "Why this priority",
        "Action item",
    ]
    for c, h in enumerate(act_headers, 1):
        act.cell(row=5, column=c, value=h)
    style_header(act, 5, 7)

    r = 6
    for i, row in enumerate([x for x in ranked if x["priority"] == "P0"], 1):
        vals = [
            i,
            row["product"].name,
            row["pct"],
            row["maturity"],
            ", ".join(row["absent"]) if row["absent"] else "—",
            row["why"],
            row["action"],
        ]
        for c, v in enumerate(vals, 1):
            cell = act.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.font = FONT_BODY
            cell.alignment = WRAP if c >= 2 else CENTER
        apply_maturity(act.cell(row=r, column=4), row["maturity"])
        act.row_dimensions[r].height = 44
        r += 1

    r += 1
    act.cell(row=r, column=1, value="P1 — Next").font = FONT_BOLD
    act.cell(row=r, column=1).fill = FILL_P1
    act.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1
    for c, h in enumerate(act_headers, 1):
        cell = act.cell(row=r, column=c, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = CENTER
        cell.border = THIN
    r += 1
    for i, row in enumerate([x for x in ranked if x["priority"] == "P1"], 1):
        vals = [
            i,
            row["product"].name,
            row["pct"],
            row["maturity"],
            ", ".join(row["absent"]) if row["absent"] else "—",
            row["why"],
            row["action"],
        ]
        for c, v in enumerate(vals, 1):
            cell = act.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.font = FONT_BODY
            cell.alignment = WRAP if c >= 2 else CENTER
        apply_maturity(act.cell(row=r, column=4), row["maturity"])
        act.row_dimensions[r].height = 40
        r += 1

    r += 2
    act.cell(row=r, column=1, value="Leadership ask (summary)").font = FONT_SECTION
    r += 1
    asks = [
        f"1. Approve focus on {len(p0)} P0 products first — these drive core collect/disburse/partner journeys and support load.",
        "2. Measure success as Progress Report Coverage % moving Partial/Critical → Complete (target: P0 products ≥85%).",
        "3. Fund Integration Guide writing + IA cleanup (naming/duplicates) in parallel — guides fail if merchants cannot find the canonical path.",
        "4. Do not commission dedicated IGs for every plugin/utility — use install guides and cross-links there.",
        "5. Re-score monthly using the Progress Report checklist (Baseline already set from the docs repository).",
    ]
    for line in asks:
        act.cell(row=r, column=1, value=line).alignment = WRAP
        act.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        act.row_dimensions[r].height = 28
        r += 1

    act.column_dimensions["A"].width = 8
    act.column_dimensions["B"].width = 40
    act.column_dimensions["C"].width = 12
    act.column_dimensions["D"].width = 14
    act.column_dimensions["E"].width = 40
    act.column_dimensions["F"].width = 44
    act.column_dimensions["G"].width = 48
    act.freeze_panes = "A6"

    return wb, {
        "total": total,
        "overall": overall,
        "complete": complete,
        "partial": partial,
        "critical": critical,
        "p0": len(p0),
        "p1": len(p1),
        "port_score": port_score,
        "port_applicable": port_applicable,
    }


def main():
    products = build_products()
    wb, stats = build_report(products)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(
        f"Products={stats['total']} Coverage={stats['overall']}% "
        f"Complete/Partial/Critical={stats['complete']}/{stats['partial']}/{stats['critical']} "
        f"P0/P1={stats['p0']}/{stats['p1']} "
        f"Score={stats['port_score']}/{stats['port_applicable']}"
    )


if __name__ == "__main__":
    main()
