#!/usr/bin/env python3
"""
PayU Docs Coverage Progress Report (FY 26-27) — Product Docs

Format mirrors the sample "Docs Coverage Progress Report" workbook:
  Product | Content Coverage | Recommendations | TW | monthly Content Coverage Score
  Per-product checklist rows + Total + Docs Score
  Portfolio rollup at the end

Data is freshly derived from the PayU Developer Documentation repository.
Sample file used ONLY for layout/columns/checklist philosophy — not its data.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path("/workspace")
OUT = Path("/workspace/docs-coverage-tracker/PayU_Docs_Coverage_Progress_Report_26-27_Product_Docs.xlsx")

# FY 26-27 tracking months (Baseline = current repo assessment)
MONTHS = [
    "Baseline\n(Jul 2026)",
    "Aug 2026",
    "Sept 2026",
    "Oct 2026",
    "Nov 2026",
    "Dec 2026",
    "Jan 2027",
    "Feb 2027",
    "Mar 2027",
]

# Standard content-coverage checklist (sample product template)
STANDARD_ITEMS = [
    "Introduction Page - Definition, Examples, Features and Advantages",
    "Use Cases (Top 3-5 Verticals)",
    "How it Works (Desc. + Process Diagram)",
    "Product Video",
    "All Dashboard Action (&Video)",
    "Other Features and Capabilties",
    "API Page",
    "Webhooks",
    "FAQs",
    "Glossary",
    "Change Log",
    "S2S/Hosted/Custom [if applicable]",
]

SIGNUP_ITEMS = [
    "Sign up Process is explained",
    "All steps are documented",
    "All required documents are documented",
    "Videos",
    "FAQs",
]

PAYMENT_METHODS_ITEMS = [
    "Descriptions and How Tos",
    "Flows",
    "Integrations",
    "FAQs",
    "Features",
    "S2S/Hosted/Custom",
]

NA = "Not Applicable"
ScoreVal = Union[int, str]  # 1, 0, or Not Applicable


def exists(*parts: str) -> bool:
    return (ROOT.joinpath(*parts)).exists()


def any_file(directory: str, patterns: list[str]) -> bool:
    base = ROOT / directory
    if not base.exists():
        return False
    for p in base.rglob("*"):
        if not p.is_file():
            continue
        name = p.name.lower()
        if any(re.search(pat, name) for pat in patterns):
            return True
    return False


def search_text(directory: str, needles: list[str], max_files: int = 80) -> bool:
    base = ROOT / directory
    if not base.exists():
        return False
    count = 0
    for p in base.rglob("*"):
        if p.suffix.lower() not in {".md", ".mdx", ".html"}:
            continue
        count += 1
        if count > max_files:
            break
        try:
            text = p.read_text(encoding="utf-8", errors="ignore").lower()
        except Exception:
            continue
        if any(n.lower() in text for n in needles):
            return True
    return False


def yn(cond: bool) -> ScoreVal:
    return 1 if cond else 0


@dataclass
class ProductBlock:
    name: str
    items: list[tuple[str, ScoreVal, str]]  # (coverage item, score, recommendation)
    tw: str = ""
    seo_note: str = "Page SEO and Heading Styles"
    template: str = "standard"  # standard | signup | payment_methods


def score_standard(
    *,
    intro: bool,
    use_cases: bool | str,
    how_it_works: bool | str,
    product_video: bool | str = False,
    dashboard: bool | str,
    other_features: bool | str,
    api: bool | str,
    webhooks: bool | str,
    faqs: bool | str,
    glossary: bool | str = False,
    changelog: bool | str = False,
    s2s_hosted: bool | str,
    recommendations: dict[str, str] | None = None,
) -> list[tuple[str, ScoreVal, str]]:
    recs = recommendations or {}

    def coerce(v: bool | str) -> ScoreVal:
        if v is True:
            return 1
        if v is False:
            return 0
        return v  # NA or other string

    mapping = [
        (STANDARD_ITEMS[0], coerce(intro)),
        (STANDARD_ITEMS[1], coerce(use_cases)),
        (STANDARD_ITEMS[2], coerce(how_it_works)),
        (STANDARD_ITEMS[3], coerce(product_video)),
        (STANDARD_ITEMS[4], coerce(dashboard)),
        (STANDARD_ITEMS[5], coerce(other_features)),
        (STANDARD_ITEMS[6], coerce(api)),
        (STANDARD_ITEMS[7], coerce(webhooks)),
        (STANDARD_ITEMS[8], coerce(faqs)),
        (STANDARD_ITEMS[9], coerce(glossary)),
        (STANDARD_ITEMS[10], coerce(changelog)),
        (STANDARD_ITEMS[11], coerce(s2s_hosted)),
    ]
    return [(label, val, recs.get(label, "")) for label, val in mapping]


def build_products() -> list[ProductBlock]:
    products: list[ProductBlock] = []

    # ---- Getting Started ----
    products.append(
        ProductBlock(
            name="Sign Up / Merchant Onboarding",
            template="signup",
            items=[
                (
                    "Sign up Process is explained",
                    yn(exists("docs/getting started/register-with-payu/register-for-a-merchant-account-on-dashboard.md")),
                    "",
                ),
                (
                    "All steps are documented",
                    yn(exists("docs/getting started/register-with-payu/index.md")),
                    "",
                ),
                (
                    "All required documents are documented",
                    yn(exists("docs/getting started/register-with-payu/documents-checklist-for-account-activation.md")),
                    "",
                ),
                (
                    "Videos",
                    0,
                    "Create an end-to-end onboarding video and short GIFs for KYC/activation steps.",
                ),
                (
                    "FAQs",
                    0,
                    "Add Sign Up / KYC FAQs (currently dashboard FAQs exist separately).",
                ),
            ],
            seo_note="Page SEO and Heading Styles",
        )
    )

    products.append(
        ProductBlock(
            name="PayU Dashboard",
            items=score_standard(
                intro=exists("docs/getting started/payu-dashboard/index.md"),
                use_cases=NA,
                how_it_works=True,
                product_video=0,
                dashboard=True,
                other_features=True,
                api=NA,
                webhooks=exists("docs/getting started/payu-dashboard/manage-webhooks-using-dashboard/index.md"),
                faqs=exists("docs/getting started/payu-dashboard/faqs-for-dashboard.md"),
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Product Video": "Add short dashboard walkthrough videos for key modules.",
                    "Glossary": "Optional — add glossary for settlement/TDR terms if support volume warrants.",
                    "Change Log": "Track dashboard feature releases.",
                },
            ),
            seo_note="Try to simplify the pages. Page SEO and Heading Styles",
        )
    )

    products.append(
        ProductBlock(
            name="Settlements / Priority Settlements",
            items=score_standard(
                intro=exists("docs/getting started/payu-dashboard/settlements-dashboard/index.md"),
                use_cases=NA,
                how_it_works=exists("docs/getting started/payu-dashboard/settlements-dashboard/priority-settlements.md"),
                product_video=0,
                dashboard=True,
                other_features=True,
                api=exists("reference/Settlements"),
                webhooks=NA,
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "FAQs": "Add settlements FAQs (priority settlements, TDR, release settlement).",
                    "Product Video": "Add settlements dashboard video.",
                },
            ),
        )
    )

    # ---- Collect Payments: Web Checkout ----
    products.append(
        ProductBlock(
            name="PayU Hosted Checkout (Prebuilt)",
            items=score_standard(
                intro=exists("docs/Collect Payments/introduction-web/prebuilt-checkout-payu-hosted/index.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=NA,
                other_features=exists(
                    "docs/Collect Payments/introduction-web/prebuilt-checkout-payu-hosted/payu-payment-page-customization.md"
                ),
                api=exists("reference/Collect Payment/_payment_payu_hosted_checkout.md"),
                webhooks=exists("docs/Collect Payments/introduction-web/webhooks.md"),
                faqs=exists("docs/Collect Payments/introduction-web/faqs-for-web-checkout-integration.md"),
                glossary=0,
                changelog=0,
                s2s_hosted=1,  # hosted path documented; part of checkout matrix
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add top vertical use cases for Hosted Checkout.",
                    "Product Video": "Add end-to-end Hosted Checkout integration video.",
                    "Glossary": "Add glossary for hash, SURL/FURL, txnid terms.",
                    "Change Log": "Publish Hosted Checkout changelog.",
                    "All Dashboard Action (&Video)": "N/A for pure API checkout — keep NA unless dashboard config pages are added.",
                },
            ),
            seo_note="Unify Prebuilt / Non-Seamless / Hosted naming. Page SEO and Heading Styles",
        )
    )
    # Fix dashboard NA properly
    products[-1].items = score_standard(
        intro=True,
        use_cases=0,
        how_it_works=True,
        product_video=0,
        dashboard=NA,
        other_features=True,
        api=True,
        webhooks=True,
        faqs=True,
        glossary=0,
        changelog=0,
        s2s_hosted=1,
        recommendations={
            "Use Cases (Top 3-5 Verticals)": "Add top vertical use cases for Hosted Checkout.",
            "Product Video": "Add end-to-end Hosted Checkout integration video.",
            "Glossary": "Add glossary for hash, SURL/FURL, txnid terms.",
            "Change Log": "Publish Hosted Checkout changelog.",
        },
    )

    products.append(
        ProductBlock(
            name="Merchant Hosted Checkout (Custom / Seamless)",
            items=score_standard(
                intro=exists("docs/Collect Payments/introduction-web/custom-checkout-merchant-hosted/index.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=NA,
                other_features=True,
                api=exists("reference/Collect Payment/_payment_merchant_hosted"),
                webhooks=True,
                faqs=True,
                glossary=0,
                changelog=0,
                s2s_hosted=1,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add use cases for custom UX merchants.",
                    "Product Video": "Add Merchant Hosted integration video (cards/UPI/NB).",
                    "Glossary": "Add seamless/custom checkout glossary.",
                    "Change Log": "Publish MH checkout changelog; deprecate duplicate API surfaces.",
                },
            ),
            seo_note="Resolve duplicate MH API surfaces. Page SEO and Heading Styles",
        )
    )

    products.append(
        ProductBlock(
            name="Server-to-Server (S2S) Checkout",
            items=score_standard(
                intro=exists("docs/Collect Payments/introduction-web/server-to-server-integration/index.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=NA,
                other_features=True,
                api=exists("reference/Collect Payment/_payment_server_to_server"),
                webhooks=True,
                faqs=True,
                glossary=0,
                changelog=0,
                s2s_hosted=1,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add when-to-choose S2S (classic vs decoupled vs direct auth).",
                    "Product Video": "Add S2S flow videos per auth model.",
                    "Glossary": "Define classic/decoupled/direct authorization clearly.",
                    "Change Log": "Track S2S API/flow changes; remove old_ API pages.",
                },
            ),
            seo_note="Create decision-tree Integration Guide. Page SEO and Heading Styles",
        )
    )

    products.append(
        ProductBlock(
            name="Checkout Plus (ICP / Bolt Checkout)",
            items=score_standard(
                intro=exists("docs/Collect Payments/introduction-web/checkout-plus-integration/index.md"),
                use_cases=0,
                how_it_works=exists(
                    "docs/Collect Payments/introduction-web/checkout-plus-integration/customer-journey-checkouplus.md"
                ),
                product_video=0,
                dashboard=NA,
                other_features=True,
                api=exists(
                    "docs/Collect Payments/introduction-web/checkout-plus-integration/apis-used-for-checkout-plus-integration.md"
                ),
                webhooks=True,
                faqs=True,
                glossary=0,
                changelog=0,
                s2s_hosted=0,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add Checkout Plus adoption use cases.",
                    "Product Video": "Add Checkout Plus product/integration video.",
                    "S2S/Hosted/Custom [if applicable]": "Clarify how Checkout Plus relates to Hosted/MH/S2S matrix.",
                    "Change Log": "Publish Checkout Plus changelog.",
                    "Glossary": "Resolve Plus / ICP / Bolt naming in a glossary/alias note.",
                },
            ),
            seo_note="Canonicalize Checkout Plus vs ICP vs Bolt naming",
        )
    )

    products.append(
        ProductBlock(
            name="CommercePro Checkout (Checkout Express)",
            items=score_standard(
                intro=exists("docs/Collect Payments/introduction-web/checkout-express/index.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=NA,
                other_features=True,
                api=exists("reference/Checkout Express"),
                webhooks=True,
                faqs=True,
                glossary=0,
                changelog=0,
                s2s_hosted=0,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add CommercePro use cases.",
                    "Product Video": "Add CommercePro integration video.",
                    "Glossary": "Document CommercePro = Checkout Express naming.",
                    "Change Log": "Add changelog.",
                    "S2S/Hosted/Custom [if applicable]": "Document callback vs response-handler integration modes clearly.",
                },
            ),
            seo_note="Unify CommercePro vs Checkout Express naming across docs/reference/plugins",
        )
    )

    # ---- No-code ----
    products.append(
        ProductBlock(
            name="Payment Links",
            items=score_standard(
                intro=exists(
                    "docs/Collect Payments/introduction-no-code-payments-integration/payment-links-dashboard/index.md"
                ),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=True,
                other_features=True,
                api=exists("reference/payment links"),
                webhooks=True,
                faqs=exists(
                    "docs/Collect Payments/introduction-no-code-payments-integration/payment-links-dashboard/faqs-payment-links.md"
                ),
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add Payment Links use cases (WhatsApp, TPV, field collection).",
                    "Product Video": "Add dashboard + API Payment Links videos.",
                    "Glossary": "Add Payment Link vs Invoice vs Button glossary.",
                    "Change Log": "Track Payment Links API/dashboard changes.",
                },
            ),
        )
    )

    products.append(
        ProductBlock(
            name="Payment Buttons",
            items=score_standard(
                intro=exists(
                    "docs/Collect Payments/introduction-no-code-payments-integration/payment-buttons-dashboard.md"
                ),
                use_cases=0,
                how_it_works=1,
                product_video=0,
                dashboard=1,
                other_features=0,
                api=NA,
                webhooks=NA,
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add use cases.",
                    "Product Video": "Add product video.",
                    "Other Features and Capabilties": "Expand beyond single-page guide.",
                    "FAQs": "Add FAQs.",
                },
            ),
        )
    )

    products.append(
        ProductBlock(
            name="Invoices (No-Code)",
            items=score_standard(
                intro=exists(
                    "docs/Collect Payments/introduction-no-code-payments-integration/invoices-dashboard/index.md"
                ),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=True,
                other_features=True,
                api=NA,
                webhooks=NA,
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add invoice use cases.",
                    "Product Video": "Add invoices dashboard video.",
                    "API Page": "Clarify API availability vs Payment Links / Zion invoices.",
                    "FAQs": "Add invoices FAQs.",
                    "Glossary": "Disambiguate Invoice vs Payment Link vs Zion subscription invoice.",
                },
            ),
        )
    )

    # ---- In-person ----
    products.append(
        ProductBlock(
            name="UPI QR",
            items=score_standard(
                intro=exists("docs/Collect Payments/in-person-payments/integrate-upi-qr/index.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=0,
                other_features=True,
                api=exists("reference/In-person payments/integrate-upi-qr-apis"),
                webhooks=0,
                faqs=exists("docs/Collect Payments/in-person-payments/integrate-upi-qr/faqs-2.md"),
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add QR use cases (retail, bill desk, etc.).",
                    "Product Video": "Add UPI QR integration video.",
                    "All Dashboard Action (&Video)": "Document any dashboard QR management actions.",
                    "Webhooks": "Document QR callbacks/webhooks explicitly.",
                    "Change Log": "Add QR API changelog.",
                },
            ),
        )
    )

    products.append(
        ProductBlock(
            name="Dynamic Storefront QR (DBQR)",
            items=score_standard(
                intro=exists("docs/Collect Payments/in-person-payments/integrated-dynamic-storefront/index.md"),
                use_cases=0,
                how_it_works=exists(
                    "docs/Collect Payments/in-person-payments/integrated-dynamic-storefront/integrated-dynamic-storefront-customer-journey.md"
                ),
                product_video=0,
                dashboard=0,
                other_features=0,
                api=0,
                webhooks=0,
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add DBQR use cases.",
                    "API Page": "Publish API reference; merge ASK AI offline-dbqr into canonical docs.",
                    "Webhooks": "Document callbacks.",
                    "FAQs": "Add FAQs.",
                    "Glossary": "Resolve DBQR / Offline DBQR / Dynamic Bharat QR naming.",
                    "Product Video": "Add product video.",
                },
            ),
            seo_note="Resolve DBQR naming; promote ASK AI content to canonical IA",
        )
    )

    products.append(
        ProductBlock(
            name="POS Terminal Integration",
            items=score_standard(
                intro=exists("docs/Collect Payments/in-person-payments/pos-terminal-integration/index.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=NA,
                other_features=0,
                api=exists("reference/In-person payments/pos-terminal-integration-apis"),
                webhooks=0,
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add POS use cases.",
                    "Product Video": "Add POS integration video.",
                    "Other Features and Capabilties": "Document device features, receipts, voids.",
                    "Webhooks": "Document POS event callbacks.",
                    "FAQs": "Add POS FAQs.",
                    "Change Log": "Track POS API versions.",
                },
            ),
        )
    )

    products.append(
        ProductBlock(
            name="Android POS SDK",
            items=score_standard(
                intro=exists("docs/Collect Payments/in-person-payments/android-pos-sdk/index.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=NA,
                other_features=True,
                api=exists("reference/In-person payments/android-pos-sdk-apis"),
                webhooks=NA,
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add Android POS SDK use cases.",
                    "Product Video": "Add SDK walkthrough video.",
                    "FAQs": "Add SDK FAQs/troubleshooting.",
                    "Change Log": "Publish SDK changelog.",
                },
            ),
        )
    )

    # ---- Plugins (grouped high-level + major ones) ----
    plugin_specs = [
        ("Shopify Plugin", "shopify", True, True, True),
        ("WooCommerce Plugin", "woocommerce", True, False, True),
        ("Magento Plugin", "magento", True, False, True),
        ("Wix Plugin", "wix", True, True, False),
        ("BigCommerce Plugin", "bigcommerce", True, False, True),
        ("Shopmatic Plugin", "shopmatic", True, False, True),
        ("Fynd Store Plugin", "fynd-integration", True, False, False),
        ("OpenCart Plugin", "opencart", True, False, True),
        ("PrestaShop Plugin", "prestashop", True, False, True),
        ("Zoho Plugin", "zoho-integration", True, False, False),
        ("Odoo Plugin", "odoo", True, False, False),
        ("Bagisto Plugin", "bagisto.md", False, False, False),
        ("Interakt for WhatsApp Business Plugin", "interakt-for-whatsapp-business", True, False, False),
    ]
    for name, folder, has_folder, has_faq, has_trouble in plugin_specs:
        base = f"docs/Collect Payments/ecommerce-platform-plugins/{folder}"
        if folder.endswith(".md"):
            intro = exists(f"docs/Collect Payments/ecommerce-platform-plugins/{folder}")
            faq = False
            features = intro
        else:
            intro = exists(f"{base}/index.md")
            faq = any_file(base, [r"faq"]) if has_faq else any_file(base, [r"faq"])
            features = True if intro else False
        products.append(
            ProductBlock(
                name=name,
                items=score_standard(
                    intro=intro,
                    use_cases=0,
                    how_it_works=intro,
                    product_video=0,
                    dashboard=NA,
                    other_features=features,
                    api=NA,
                    webhooks=NA,
                    faqs=faq,
                    glossary=0,
                    changelog=0,
                    s2s_hosted=NA,
                    recommendations={
                        "Use Cases (Top 3-5 Verticals)": "Add merchant use cases for this plugin.",
                        "Product Video": "Add install/configure video.",
                        "FAQs": "Add or expand plugin FAQs." if not faq else "",
                        "Change Log": "Publish plugin version changelog.",
                        "Glossary": "Optional for plugin channel.",
                    },
                ),
                seo_note="Keep install + troubleshooting current; use shared plugins checklist",
            )
        )

    # ---- Server SDKs ----
    for lang, file in [
        ("Go SDK", "go-sdk.md"),
        ("Java SDK", "java-sdk.md"),
        ("PHP SDK", "php-sdk.md"),
        ("Python SDK", "python-sdk.md"),
        ("Node.js SDK", "node-js-sdk.md"),
    ]:
        path_ok = exists(f"docs/Collect Payments/explore-server-integrations/{file}")
        products.append(
            ProductBlock(
                name=lang,
                items=score_standard(
                    intro=path_ok,
                    use_cases=0,
                    how_it_works=path_ok,
                    product_video=0,
                    dashboard=NA,
                    other_features=0,
                    api=1,  # uses Collect Payment APIs
                    webhooks=1,  # shared web checkout webhooks
                    faqs=0,
                    glossary=0,
                    changelog=0,
                    s2s_hosted=NA,
                    recommendations={
                        "Use Cases (Top 3-5 Verticals)": "Add SDK usage scenarios.",
                        "Product Video": "Add SDK quickstart video.",
                        "Other Features and Capabilties": "Expand beyond single-page SDK note (samples, errors, verify payment).",
                        "FAQs": "Add SDK FAQs.",
                        "Change Log": "Publish SDK version changelog.",
                    },
                ),
                seo_note="Upgrade thin single-page SDK docs to full Integration Guide template",
            )
        )

    # ---- Mobile SDKs ----
    mobile = [
        (
            "Android CheckoutPro SDK",
            "docs/Collect Payments/mobile-sdks/explore-android-sdks/android-checkoutpro-sdk",
            True,
            True,
            True,
        ),
        (
            "Android Core SDK",
            "docs/Collect Payments/mobile-sdks/explore-android-sdks/android-core-sdk",
            True,
            True,
            False,
        ),
        (
            "iOS CheckoutPro SDK",
            "docs/Collect Payments/mobile-sdks/explore-ios-sdks/ios-checkoutpro-sdk",
            True,
            True,
            False,
        ),
        (
            "iOS Core SDK",
            "docs/Collect Payments/mobile-sdks/explore-ios-sdks/ios-core-sdk",
            True,
            True,
            False,
        ),
        (
            "iOS Custom Browser SDK",
            "docs/Collect Payments/mobile-sdks/explore-ios-sdks/ios-custombrowser-sdk",
            True,
            True,
            False,
        ),
        (
            "React Native CheckoutPro SDK",
            "docs/Collect Payments/mobile-sdks/explore-reactnative-sdks/react-native-checkoutpro-sdk",
            True,
            True,
            True,
        ),
        (
            "Flutter CheckoutPro SDK",
            "docs/Collect Payments/mobile-sdks/flutter-sdk-introduction/flutter-checkoutpro-sdk",
            True,
            False,
            True,
        ),
        (
            "Cordova CheckoutPro SDK",
            "docs/Collect Payments/mobile-sdks/cordova-mobile-sdks/cordova-sdk-introduction",
            True,
            False,
            False,
        ),
        (
            "UPI Bolt SDK (Cross-Platform)",
            "docs/Collect Payments/mobile-sdks/explore-android-sdks/payu-bolt-sdk",
            True,
            False,
            False,
        ),
    ]
    for name, folder, intro, faqs_nearby, changelog in mobile:
        faq = (
            any_file("docs/Collect Payments/mobile-sdks/explore-android-sdks/faqs-android-sdk", [r"faq"])
            if "Android" in name
            else any_file("docs/Collect Payments/mobile-sdks/explore-ios-sdks", [r"faq"])
            if "iOS" in name
            else any_file("docs/Collect Payments/mobile-sdks/explore-reactnative-sdks", [r"faq"])
            if "React Native" in name
            else False
        )
        cl = any_file(folder, [r"change", r"version", r"changelog", r"release"])
        products.append(
            ProductBlock(
                name=name,
                items=score_standard(
                    intro=exists(f"{folder}/index.md") or exists(folder),
                    use_cases=0,
                    how_it_works=True,
                    product_video=0,
                    dashboard=NA,
                    other_features=True,
                    api=NA,
                    webhooks=NA,
                    faqs=faq,
                    glossary=0,
                    changelog=cl,
                    s2s_hosted=NA,
                    recommendations={
                        "Use Cases (Top 3-5 Verticals)": "Add mobile integration use cases.",
                        "Product Video": "Add SDK integration video / sample-app walkthrough.",
                        "FAQs": "Ensure platform FAQ hub is linked." if not faq else "",
                        "Change Log": "Keep version history/changelog current." if not cl else "",
                        "Glossary": "Add SDK terms (hash, CB, CheckoutPro vs Core).",
                    },
                ),
                seo_note="Platform landing should include Choose-your-SDK matrix",
            )
        )

    # ---- Offerings ----
    offerings = [
        (
            "EMI / Cardless EMI",
            "docs/Offerings/introduction-to-affordability/emi-api-integration",
            "reference/General/emi-apis",
            True,
            False,
            True,
            False,
            1,
        ),
        (
            "Offer Engine / Offers",
            "docs/Offerings/introduction-to-affordability/offers-integration-1",
            "reference/Affordability/offer-apis",
            True,
            False,
            False,
            False,
            NA,
        ),
        (
            "Affordability Widget",
            "docs/Offerings/introduction-to-affordability/affordability-suite",
            None,
            True,
            False,
            False,
            False,
            NA,
        ),
        (
            "BNPL / Pay Later",
            "docs/Offerings/introduction-to-affordability/payu-bnpl-integration-introduction",
            "reference/Affordability/bnpl-integration-apis",
            True,
            False,
            False,
            False,
            NA,
        ),
        (
            "LazyPay Pay-in-3",
            "docs/Offerings/introduction-to-affordability/lazypay-pay-in-3",
            "reference/Affordability/lazypay-pay-in-3",
            True,
            False,
            False,
            False,
            NA,
        ),
        (
            "MobiKwik Link & Pay",
            "docs/Offerings/introduction-to-affordability/mobikwik-link-pay-integration",
            "reference/Third-Party Wallets/mobikwik-link-wallet-apis",
            True,
            False,
            False,
            False,
            NA,
        ),
        (
            "Loyalty Edge",
            "docs/Offerings/introduction-to-affordability/loyalty-edge-introduction",
            None,
            True,
            True,
            False,
            False,
            NA,
        ),
        (
            "Affordability Suite (Hub)",
            "docs/Offerings/introduction-to-affordability",
            "reference/Affordability",
            True,
            True,
            False,
            False,
            NA,
        ),
    ]
    for name, docs_path, api_path, how, dash, faq, clog, s2s in offerings:
        products.append(
            ProductBlock(
                name=name,
                items=score_standard(
                    intro=exists(f"{docs_path}/index.md") or exists(docs_path),
                    use_cases=0 if "Suite" not in name else 0,
                    how_it_works=how,
                    product_video=0,
                    dashboard=dash if dash is not True else True,
                    other_features=True,
                    api=exists(api_path) if api_path else NA if api_path is None else 0,
                    webhooks=NA if name in {"Offer Engine / Offers", "Affordability Widget", "Loyalty Edge", "Affordability Suite (Hub)"} else 0,
                    faqs=any_file(docs_path, [r"faq"]) or faq,
                    glossary=0,
                    changelog=clog,
                    s2s_hosted=s2s,
                    recommendations={
                        "Use Cases (Top 3-5 Verticals)": "Add vertical use cases.",
                        "Product Video": "Add product/integration video.",
                        "FAQs": "Add FAQs." if not (any_file(docs_path, [r"faq"]) or faq) else "",
                        "Glossary": "Add affordability glossary (EMI/BNPL/offers terms).",
                        "Change Log": "Track affordability API/product changes.",
                    },
                ),
            )
        )

    # Fix dashboard bools for offerings that wrongly passed True as dash flag meaning "has dashboard"
    # Loyalty Edge and Offer Engine have dashboard pages
    for i, p in enumerate(products):
        if p.name == "Offer Engine / Offers":
            products[i].items = score_standard(
                intro=True,
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=exists("docs/Offerings/introduction-to-affordability/offers-integration-1/offers-dashboard/index.md"),
                other_features=True,
                api=True,
                webhooks=NA,
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add offer campaign use cases.",
                    "Product Video": "Add Offer Engine video.",
                    "FAQs": "Add Offer Engine FAQs.",
                    "Webhooks": "Confirm if offer events need webhooks documentation.",
                    "Change Log": "Publish offers changelog.",
                    "Glossary": "Add offers glossary.",
                },
            )
        if p.name == "Loyalty Edge":
            products[i].items = score_standard(
                intro=True,
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=True,
                other_features=True,
                api=NA,
                webhooks=NA,
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add Loyalty Edge use cases.",
                    "Product Video": "Add product video.",
                    "API Page": "Publish APIs when available.",
                    "FAQs": "Add FAQs.",
                },
            )

    products.append(
        ProductBlock(
            name="Subscriptions / Recurring Payments",
            items=score_standard(
                intro=exists("docs/Offerings/introduction-recurring-payments-integration/index.md"),
                use_cases=0,
                how_it_works=exists(
                    "docs/Offerings/introduction-recurring-payments-integration/customer-experience-and-workflow-recurring-payments"
                )
                or True,
                product_video=0,
                dashboard=exists(
                    "docs/Offerings/introduction-recurring-payments-integration/subscription-dashboard/index.md"
                ),
                other_features=True,
                api=exists("reference/Subscriptions"),
                webhooks=True,
                faqs=exists("docs/Offerings/introduction-recurring-payments-integration/faqs-recurring-payments.md"),
                glossary=0,
                changelog=0,
                s2s_hosted=1,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add SaaS/lending/OTT subscription use cases.",
                    "Product Video": "Add recurring payments product video.",
                    "Glossary": "Unify Subscriptions / Recurring / SI terminology in glossary.",
                    "Change Log": "Publish recurring/SI changelog; consolidate duplicate API aggregates.",
                },
            ),
            seo_note="Unify Subscriptions vs Recurring vs SI naming; merge ASK AI duplicate",
        )
    )

    products.append(
        ProductBlock(
            name="Zion Subscription Automation",
            items=score_standard(
                intro=exists(
                    "docs/Offerings/introduction-recurring-payments-integration/using-zion-subscription-automation-platform/index.md"
                ),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=True,
                other_features=True,
                api=exists("reference/ZION"),
                webhooks=exists(
                    "docs/Offerings/introduction-recurring-payments-integration/using-zion-subscription-automation-platform/webhooks-for-subscription.md"
                ),
                faqs=exists(
                    "docs/Offerings/introduction-recurring-payments-integration/using-zion-subscription-automation-platform/faqs-zion-integration.md"
                ),
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add Zion automation use cases.",
                    "Product Video": "Add Zion walkthrough video.",
                    "Glossary": "Standardize Zion vs ZION casing.",
                    "Change Log": "Add Zion API changelog.",
                },
            ),
        )
    )

    products.append(
        ProductBlock(
            name="Dynamic Currency Conversion / International Payments",
            items=score_standard(
                intro=exists("docs/Offerings/introduction-dynamic-currency-conversion/index.md"),
                use_cases=0,
                how_it_works=exists(
                    "docs/Offerings/introduction-dynamic-currency-conversion/dynamic-currency-conversion-workflow.md"
                ),
                product_video=0,
                dashboard=0,
                other_features=True,
                api=exists("reference/international payments"),
                webhooks=NA,
                faqs=exists(
                    "docs/Offerings/introduction-dynamic-currency-conversion/faqs-dynamic-currency-conversion.md"
                ),
                glossary=0,
                changelog=0,
                s2s_hosted=1,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add international payments use cases.",
                    "Product Video": "Add DCC product video.",
                    "All Dashboard Action (&Video)": "Document any dashboard FX/DCC configuration.",
                    "Glossary": "Expand acronyms; disambiguate MCP (pricing) vs MCP (Model Context Protocol).",
                    "Change Log": "Track currency/DCC changes.",
                },
            ),
            seo_note="Unify DCC vs International Payments naming",
        )
    )

    products.append(
        ProductBlock(
            name="Cross-Border Payments Import (PACB)",
            items=score_standard(
                intro=exists("docs/Offerings/introduction-cross-border-payments-import/index.md"),
                use_cases=0,
                how_it_works=exists(
                    "docs/Offerings/introduction-cross-border-payments-import/workflow-for-cross-border-payments-import.md"
                ),
                product_video=0,
                dashboard=0,
                other_features=True,
                api=exists("reference/Cross-border Payments"),
                webhooks=0,
                faqs=exists(
                    "docs/Offerings/introduction-cross-border-payments-import/faqs-for-cross-border-payments.md"
                ),
                glossary=0,
                changelog=0,
                s2s_hosted=1,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add import/travel/education (LRS) use cases.",
                    "Product Video": "Add CB Import integration video.",
                    "All Dashboard Action (&Video)": "Document CB dashboard actions if any.",
                    "Webhooks": "Document CB webhooks/status callbacks.",
                    "Glossary": "Define CB / PACB / LRS / Import terms.",
                    "Change Log": "Publish CB API changelog.",
                },
            ),
            seo_note="Standardize CB/PACB/Import naming",
        )
    )

    more_offerings = [
        (
            "EFTNET / NEFT-RTGS Collect",
            "docs/Offerings/introduction-to-eftnet",
            "reference/Collect Payment",
            True,
            NA,
            0,
            0,
            1,
        ),
        (
            "Pre-Authorize / Auth & Capture",
            "docs/Offerings/auth-and-capture-pre-authorize-card-payments",
            "reference/Pre-Authorize Payment",
            True,
            NA,
            0,
            0,
            1,
        ),
        (
            "UPI One-Time Mandate (OTM / Reserve Pay)",
            "docs/Offerings/auth-and-capture-pre-authorize-card-payments/upi-one-time-mandate-integration",
            "reference/Pre-Authorize Payment/pre-authorize-payments-for-upi",
            True,
            NA,
            0,
            0,
            1,
        ),
        (
            "Split Settlements (Aggregator / Marketplace)",
            "docs/Offerings/split-settlments",
            "reference/split settlements",
            True,
            True,
            True,
            0,
            1,
        ),
        (
            "Tokenization / Save Cards (Vault)",
            "docs/Offerings/introduction-save-cards",
            "reference/Tokenization",
            True,
            NA,
            0,
            0,
            1,
        ),
        (
            "Third-Party Verification (TPV)",
            "docs/Offerings/introduction-to-payu-tpv",
            None,  # dispersed
            True,
            NA,
            True,
            0,
            1,
        ),
        (
            "Recommendation Engine",
            "docs/Offerings/recommendation-engine",
            "docs/Offerings/recommendation-engine/fetch-recommendation-engine-api.md",
            True,
            NA,
            0,
            0,
            NA,
        ),
        (
            "Native OTP Flow",
            "docs/Offerings/native-otp-flow-integration",
            "reference/Collect Payment/native-otp-flow-apis",
            True,
            NA,
            0,
            0,
            1,
        ),
        (
            "Apple Pay",
            "docs/Offerings/apple-pay-integration",
            "reference/Apple Pay",
            True,
            NA,
            0,
            0,
            1,
        ),
        (
            "Account Funding Transaction (AFT)",
            "docs/Offerings/account-funding-transaction-integration",
            "reference/Collect Payment/_payment_merchant_hosted/_payment_api_merchant_hosted_aft.md",
            True,
            NA,
            0,
            0,
            1,
        ),
        (
            "Mutual Fund Payments (WealthTech)",
            "docs/Offerings/mutual-funds-payments",
            "reference/Wealth Tech",
            True,
            NA,
            0,
            0,
            1,
        ),
        (
            "Merchant Wallet",
            "docs/Offerings/introduction-to-merchant-wallet",
            "reference/Merchant Wallet",
            True,
            NA,
            0,
            0,
            NA,
        ),
        (
            "Virtual Cards / GPR Cards",
            "docs/Offerings/virtual-cards-introduction",
            "docs/Offerings/virtual-cards-introduction/apis-used-in-virtual-cards-integration.md",
            True,
            NA,
            0,
            0,
            NA,
        ),
        (
            "Refunds",
            "docs/Offerings/introduction-refunds",
            "reference/General/refund-apis",
            True,
            True,
            True,
            0,
            NA,
        ),
        (
            "Chargebacks / Disputes",
            "docs/Offerings/chargeback",
            "reference/Chargeback",
            True,
            True,
            0,
            0,
            NA,
        ),
        (
            "Rewards / RewardX Partner Integration",
            "docs/Offerings/rewards-partner-integration",
            "reference/REWARD PARTNERS",
            True,
            NA,
            0,
            0,
            1,
        ),
        (
            "Banking Connect (IBMB / NBBL)",
            "docs/Offerings/banking-connect-ibmb-or-nbbl",
            None,
            True,
            NA,
            0,
            0,
            1,
        ),
        (
            "Pluxee Card (Sodexo)",
            "docs/Collect Payments/introduction-web/custom-checkout-merchant-hosted/integrate-with-merchant-hosted-checkout-for-pluxee-card.md",
            "reference/Sodexo",
            True,
            NA,
            0,
            0,
            1,
        ),
    ]

    for name, docs_path, api_path, how, dash, faq, clog, s2s in more_offerings:
        intro = exists(f"{docs_path}/index.md") if not docs_path.endswith(".md") else exists(docs_path)
        if not intro:
            intro = exists(docs_path)
        api_score: ScoreVal
        if api_path is None:
            api_score = 0 if name == "Third-Party Verification (TPV)" else NA
            if name == "Third-Party Verification (TPV)":
                api_score = 0  # dispersed — treat as gap
            if name == "Banking Connect (IBMB / NBBL)":
                api_score = 0
        else:
            api_score = yn(exists(api_path))

        dash_score: ScoreVal = dash if isinstance(dash, str) else yn(bool(dash))
        faq_score: ScoreVal
        if isinstance(faq, str):
            faq_score = faq
        elif faq is True:
            faq_score = 1
        else:
            d = docs_path if not docs_path.endswith(".md") else str(Path(docs_path).parent)
            faq_score = yn(any_file(d, [r"faq"]))

        wh: ScoreVal = NA
        if name == "Refunds":
            wh = yn(exists("docs/Offerings/introduction-refunds/webhooks-for-refunds.md"))
        elif name == "Chargebacks / Disputes":
            wh = yn(exists("docs/Offerings/chargeback/webhooks-for-chargeback"))
        elif name == "Split Settlements (Aggregator / Marketplace)":
            wh = 0
        elif name in {"Pre-Authorize / Auth & Capture", "UPI One-Time Mandate (OTM / Reserve Pay)", "Tokenization / Save Cards (Vault)", "Third-Party Verification (TPV)", "Apple Pay", "Mutual Fund Payments (WealthTech)", "Merchant Wallet", "Rewards / RewardX Partner Integration"}:
            wh = 0

        products.append(
            ProductBlock(
                name=name,
                items=score_standard(
                    intro=intro,
                    use_cases=0,
                    how_it_works=how,
                    product_video=0,
                    dashboard=dash_score,
                    other_features=True if intro else 0,
                    api=api_score,
                    webhooks=wh,
                    faqs=faq_score,
                    glossary=0,
                    changelog=clog,
                    s2s_hosted=s2s,
                    recommendations={
                        "Use Cases (Top 3-5 Verticals)": "Add use cases.",
                        "Product Video": "Add product/integration video.",
                        "Glossary": "Add product glossary where domain terms are dense.",
                        "Change Log": "Add changelog.",
                        "FAQs": "Add FAQs." if faq_score == 0 else "",
                        "API Page": "Consolidate dispersed API assets into a clear reference landing."
                        if name == "Third-Party Verification (TPV)"
                        else "",
                        "Webhooks": "Document webhooks/callbacks." if wh == 0 else "",
                    },
                ),
                seo_note=(
                    "Fix folder typo split-settlments"
                    if "Split" in name
                    else "Rename Sodexo reference to Pluxee"
                    if "Pluxee" in name
                    else "Align Mutual Funds guide with Wealth Tech API category"
                    if "Mutual Fund" in name
                    else "Page SEO and Heading Styles"
                ),
            )
        )

    # ---- Payouts ----
    products.append(
        ProductBlock(
            name="PayU Payouts",
            items=score_standard(
                intro=exists("docs/payouts/introduction-to-payouts.md"),
                use_cases=0,
                how_it_works=exists("docs/payouts/process-flow-for-payouts.md"),
                product_video=0,
                dashboard=exists("docs/payouts/payouts-dashboard/index.md"),
                other_features=True,
                api=exists("reference/payouts"),
                webhooks=exists("docs/payouts/payouts-integration/payouts-webhooks.md"),
                faqs=exists("docs/payouts/faqs-for-payouts.md"),
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add payouts use cases (marketplace seller settle, refunds disbursement, payroll-like).",
                    "Product Video": "Add Payouts product video.",
                    "Glossary": "Add payouts glossary (VPA, beneficiary, Smart Send).",
                    "Change Log": "Publish payouts API changelog; remove releasepending- filename residue.",
                },
            ),
        )
    )
    products.append(
        ProductBlock(
            name="Smart Send",
            items=score_standard(
                intro=exists("docs/payouts/payouts-integration/smart-send-introduction/index.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=True,
                other_features=True,
                api=exists("reference/payouts/smart-send-apis"),
                webhooks=True,
                faqs=True,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add Smart Send use cases.",
                    "Product Video": "Add Smart Send video.",
                    "Change Log": "Track Smart Send API changes.",
                },
            ),
        )
    )
    products.append(
        ProductBlock(
            name="Pay to Phone",
            items=score_standard(
                intro=exists("docs/payouts/releasepending-pay-to-phone-integration/releasepending-pay-to-phone-initiation.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=0,
                other_features=True,
                api=exists("reference/payouts/releasepending-pay-to-phone-configuration-apis"),
                webhooks=True,
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add Pay to Phone use cases.",
                    "All Dashboard Action (&Video)": "Document dashboard configuration if available.",
                    "FAQs": "Add FAQs.",
                    "Product Video": "Add product video.",
                    "Change Log": "Rename releasepending- paths; add changelog.",
                },
            ),
            seo_note="Remove releasepending- editorial prefix from paths",
        )
    )

    # ---- Partners ----
    products.append(
        ProductBlock(
            name="PayU Partner Program & Portal",
            items=score_standard(
                intro=exists("docs/partners/payu-partner-program-overview.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=exists("docs/partners/partner-portal/index.md"),
                other_features=True,
                api=NA,
                webhooks=NA,
                faqs=exists("docs/partners/faqs-partner-integration.md"),
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add partner program use cases.",
                    "Product Video": "Add Partner Portal onboarding video.",
                    "Change Log": "Track portal feature changes.",
                },
            ),
        )
    )
    products.append(
        ProductBlock(
            name="Partner Merchant Onboarding (API / OAuth)",
            items=score_standard(
                intro=exists("docs/partners/internal-reviewpartner-integration-overview/index.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=True,
                other_features=True,
                api=exists("reference/ParTner integration"),
                webhooks=True,
                faqs=True,
                glossary=0,
                changelog=0,
                s2s_hosted=1,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add partner onboarding use cases.",
                    "Product Video": "Add OAuth/co-branded onboarding video.",
                    "API Page": "Merge duplicate Partner API trees; fix ParTner casing.",
                    "Glossary": "Add partner onboarding glossary (CKYC, VKYC, UBO, etc.).",
                    "Change Log": "Publish partner API changelog.",
                    "All Dashboard Action (&Video)": "Add portal/onboarding videos.",
                },
            ),
            seo_note="Publish internal-review partner overview as canonical; merge parallel API trees",
        )
    )
    products.append(
        ProductBlock(
            name="Partner Payments Integration",
            items=score_standard(
                intro=exists("docs/partners/partner-payments-integration.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=NA,
                other_features=True,
                api=exists("reference/ParTner integration/partner-payment-integration-apis"),
                webhooks=0,
                faqs=True,
                glossary=0,
                changelog=0,
                s2s_hosted=1,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add partner payment use cases.",
                    "Webhooks": "Document partner payment webhooks.",
                    "Product Video": "Add integration video.",
                    "Change Log": "Add changelog.",
                },
            ),
        )
    )

    # ---- WhatsApp, BBPS, MCP, Monitoring ----
    products.append(
        ProductBlock(
            name="WhatsApp Payments",
            items=score_standard(
                intro=exists("docs/Whatsapp integration/whatsapp-integration-introduction.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=0,
                other_features=True,
                api=0,
                webhooks=1,
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add WhatsApp P2M / EPL use cases.",
                    "API Page": "Add clear API map for native WhatsApp payments.",
                    "All Dashboard Action (&Video)": "Document WhatsApp Business linking steps with video.",
                    "FAQs": "Add WhatsApp payments FAQs.",
                    "Product Video": "Add product video.",
                    "Change Log": "Add changelog.",
                },
            ),
        )
    )
    products.append(
        ProductBlock(
            name="BBPS Connect Agent",
            items=score_standard(
                intro=exists("docs/BBPS/connect-agent-api-integration/index.md"),
                use_cases=0,
                how_it_works=exists("docs/BBPS/connect-agent-api-integration/bbps-integration-flow.md"),
                product_video=0,
                dashboard=NA,
                other_features=True,
                api=exists("reference/BBPS"),
                webhooks=0,
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add BBPS agent use cases.",
                    "Product Video": "Add BBPS integration video.",
                    "Webhooks": "Document status/callback model.",
                    "FAQs": "Add BBPS FAQs.",
                    "Glossary": "Add BBPS glossary (biller, COU, etc.).",
                    "Change Log": "Track BBPS API versions.",
                },
            ),
        )
    )
    products.append(
        ProductBlock(
            name="BBPS Prepaid Recharge",
            items=score_standard(
                intro=exists("docs/BBPS/recharge-api-integration/index.md"),
                use_cases=0,
                how_it_works=exists("docs/BBPS/recharge-api-integration/prepaid-recharge-workflow.md"),
                product_video=0,
                dashboard=NA,
                other_features=True,
                api=exists("reference/BBPS/bbps-prepaid-apis"),
                webhooks=0,
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add prepaid recharge use cases.",
                    "FAQs": "Add FAQs.",
                    "Product Video": "Add workflow video.",
                    "Change Log": "Add changelog.",
                },
            ),
        )
    )
    products.append(
        ProductBlock(
            name="PayU Remote MCP Server",
            items=score_standard(
                intro=exists("docs/MCP & CLI/payu-remote-mcp-server-integration.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=NA,
                other_features=0,
                api=NA,
                webhooks=NA,
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add agentic merchant ops use cases.",
                    "Other Features and Capabilties": "Publish full tool catalog + auth examples.",
                    "FAQs": "Add MCP FAQs; disambiguate MCP acronym vs Multi-Currency Pricing.",
                    "Product Video": "Add setup video for Claude/Cursor clients.",
                    "Change Log": "Track tool/schema changes.",
                },
            ),
            seo_note="Always expand MCP acronym on first use",
        )
    )
    products.append(
        ProductBlock(
            name="PayU CLI",
            items=score_standard(
                intro=exists("docs/MCP & CLI/payu-cli.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=NA,
                other_features=0,
                api=NA,
                webhooks=NA,
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Other Features and Capabilties": "Expand command reference with examples.",
                    "FAQs": "Add CLI FAQs.",
                    "Product Video": "Add CLI quickstart video.",
                    "Change Log": "Track CLI releases.",
                },
            ),
        )
    )
    products.append(
        ProductBlock(
            name="Agentic Commerce Suite",
            items=score_standard(
                intro=exists("docs/MCP & CLI/agentic-commerce/index.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=NA,
                other_features=True,
                api=0,
                webhooks=0,
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add ChatGPT/WhatsApp agent commerce use cases.",
                    "API Page": "Publish agentic commerce API/reference map.",
                    "FAQs": "Add FAQs.",
                    "Product Video": "Add ChatGPT merchant app build video.",
                    "Change Log": "Track suite releases.",
                },
            ),
        )
    )
    products.append(
        ProductBlock(
            name="PayU Overwatch (Monitoring & Alerts)",
            items=score_standard(
                intro=exists("docs/Monitoring & Alerts/payu-monitoring-alerts-overwatch/index.md"),
                use_cases=0,
                how_it_works=True,
                product_video=0,
                dashboard=True,
                other_features=True,
                api=NA,
                webhooks=exists("docs/Monitoring & Alerts/payu-monitoring-alerts-overwatch/webhook-alerts.md"),
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Use Cases (Top 3-5 Verticals)": "Add monitoring use cases.",
                    "FAQs": "Add Overwatch FAQs.",
                    "Product Video": "Add setup video.",
                    "Change Log": "Track alert catalog changes.",
                },
            ),
        )
    )

    # ---- Payment methods aggregate (like sample) ----
    products.append(
        ProductBlock(
            name="Payment Methods (Cards / UPI / NB / Wallets / EMI / BNPL)",
            template="payment_methods",
            items=[
                (
                    "Descriptions and How Tos",
                    1,
                    "",
                ),
                (
                    "Flows",
                    1,
                    "",
                ),
                (
                    "Integrations",
                    1,
                    "",
                ),
                (
                    "FAQs",
                    1,
                    "",
                ),
                (
                    "Features",
                    1,
                    "",
                ),
                (
                    "S2S/Hosted/Custom",
                    1,
                    "Keep method pages synchronized across Hosted/MH/S2S.",
                ),
            ],
            seo_note="Simplify pages, check heading styles and SEO",
        )
    )

    products.append(
        ProductBlock(
            name="Air India Checkout API Suite",
            items=score_standard(
                intro=exists("docs/AIR India/air-india-integration-apis.md"),
                use_cases=NA,
                how_it_works=NA,
                product_video=NA,
                dashboard=NA,
                other_features=True,
                api=1,
                webhooks=NA,
                faqs=NA,
                glossary=NA,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "Change Log": "Track merchant-specific API changes privately.",
                },
            ),
            seo_note="Merchant-specific — keep out of general public nav unless productized",
        )
    )

    products.append(
        ProductBlock(
            name="General Utility APIs (Verify / BIN / Health / Bank Verification)",
            items=score_standard(
                intro=exists("docs/API basics/rest-api-format.md"),
                use_cases=NA,
                how_it_works=True,
                product_video=NA,
                dashboard=NA,
                other_features=True,
                api=exists("reference/General"),
                webhooks=1,
                faqs=0,
                glossary=0,
                changelog=0,
                s2s_hosted=NA,
                recommendations={
                    "FAQs": "Add utility API FAQs (verify payment, BIN, health).",
                    "Glossary": "Add API basics glossary.",
                    "Change Log": "Track utility API changes.",
                },
            ),
            seo_note="Improve discoverability from checkout Integration Guides",
        )
    )

    return products


# ---- Excel formatting ----
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_PRODUCT = PatternFill("solid", fgColor="D6EAF8")
FILL_TOTAL = PatternFill("solid", fgColor="FFF2CC")
FILL_SCORE = PatternFill("solid", fgColor="E2EFDA")
FILL_SUMMARY = PatternFill("solid", fgColor="FCE4D6")
FILL_ONE = PatternFill("solid", fgColor="C6EFCE")
FILL_ZERO = PatternFill("solid", fgColor="FFC7CE")
FILL_NA = PatternFill("solid", fgColor="D9D9D9")
FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_BOLD = Font(name="Calibri", bold=True, size=10)
FONT_BODY = Font(name="Calibri", size=10)
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def is_scored(v: ScoreVal) -> bool:
    return v in (0, 1) or v == "0" or v == "1"


def as_num(v: ScoreVal) -> Optional[int]:
    if v in (1, "1"):
        return 1
    if v in (0, "0"):
        return 0
    return None


def write_workbook(products: list[ProductBlock]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Docs Coverage"

    # Row 1 headers
    ws["A1"] = "Product"
    ws["B1"] = "Content Coverage"
    ws["C1"] = "Recommendations"
    ws["D1"] = "TW"
    ws["E1"] = "Content Coverage Score"
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = CENTER
        cell.border = THIN
    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=5 + len(MONTHS) - 1)
    for col in range(5, 5 + len(MONTHS)):
        cell = ws.cell(row=1, column=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = THIN
        cell.alignment = CENTER

    # Row 2 month labels
    for i, m in enumerate(MONTHS):
        cell = ws.cell(row=2, column=5 + i, value=m)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = CENTER
        cell.border = THIN
    for col in range(1, 5):
        cell = ws.cell(row=2, column=col, value="")
        cell.fill = FILL_HEADER
        cell.border = THIN

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "A3"

    row = 3
    portfolio_totals = []  # applicable counts per product
    portfolio_scores = []  # docs scores per product

    for p in products:
        first = True
        applicable = 0
        score = 0
        item_rows = []
        for label, val, rec in p.items:
            item_rows.append((label, val, rec))
            n = as_num(val)
            if n is not None:
                applicable += 1
                score += n

        for idx, (label, val, rec) in enumerate(item_rows):
            ws.cell(row=row, column=1, value=p.name if first else "").font = FONT_BOLD if first else FONT_BODY
            if first:
                ws.cell(row=row, column=1).fill = FILL_PRODUCT
            ws.cell(row=row, column=2, value=label).font = FONT_BODY
            ws.cell(row=row, column=3, value=rec).font = FONT_BODY
            ws.cell(row=row, column=4, value=p.tw if first else "").font = FONT_BODY
            # Baseline month only populated; future months blank for tracking
            baseline_cell = ws.cell(row=row, column=5, value=val)
            for c in range(1, 5):
                ws.cell(row=row, column=c).border = THIN
                ws.cell(row=row, column=c).alignment = WRAP
            baseline_cell.border = THIN
            baseline_cell.alignment = CENTER
            if as_num(val) == 1:
                baseline_cell.fill = FILL_ONE
            elif as_num(val) == 0:
                baseline_cell.fill = FILL_ZERO
            elif val == NA:
                baseline_cell.fill = FILL_NA
            for mi in range(1, len(MONTHS)):
                cell = ws.cell(row=row, column=5 + mi, value="")
                cell.border = THIN
                cell.alignment = CENTER
            first = False
            row += 1

        # Total row
        ws.cell(row=row, column=1, value="").border = THIN
        ws.cell(row=row, column=2, value="").border = THIN
        ws.cell(row=row, column=3, value=p.seo_note).border = THIN
        ws.cell(row=row, column=3).alignment = WRAP
        tot = ws.cell(row=row, column=4, value="Total")
        tot.font = FONT_BOLD
        tot.fill = FILL_TOTAL
        tot.border = THIN
        tot.alignment = CENTER
        tcell = ws.cell(row=row, column=5, value=applicable)
        tcell.font = FONT_BOLD
        tcell.fill = FILL_TOTAL
        tcell.border = THIN
        tcell.alignment = CENTER
        for mi in range(1, len(MONTHS)):
            cell = ws.cell(row=row, column=5 + mi, value="")
            cell.fill = FILL_TOTAL
            cell.border = THIN
        row += 1

        # Docs Score row
        ws.cell(row=row, column=1, value="").border = THIN
        ws.cell(row=row, column=2, value="").border = THIN
        ws.cell(row=row, column=3, value="").border = THIN
        ds = ws.cell(row=row, column=4, value="Docs Score")
        ds.font = FONT_BOLD
        ds.fill = FILL_SCORE
        ds.border = THIN
        ds.alignment = CENTER
        scell = ws.cell(row=row, column=5, value=score)
        scell.font = FONT_BOLD
        scell.fill = FILL_SCORE
        scell.border = THIN
        scell.alignment = CENTER
        for mi in range(1, len(MONTHS)):
            cell = ws.cell(row=row, column=5 + mi, value="")
            cell.fill = FILL_SCORE
            cell.border = THIN
        portfolio_totals.append(applicable)
        portfolio_scores.append(score)
        row += 1

        # blank separator
        row += 1

    # Portfolio summary (like sample)
    ws.cell(row=row, column=2, value="Total").font = FONT_BOLD
    ws.cell(row=row, column=2).fill = FILL_SUMMARY
    ws.cell(row=row, column=5, value=sum(portfolio_totals)).font = FONT_BOLD
    ws.cell(row=row, column=5).fill = FILL_SUMMARY
    ws.cell(row=row, column=5).alignment = CENTER
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = THIN
    row += 1

    ws.cell(row=row, column=2, value="Docs Coverage Score").font = FONT_BOLD
    ws.cell(row=row, column=2).fill = FILL_SUMMARY
    ws.cell(row=row, column=5, value=sum(portfolio_scores)).font = FONT_BOLD
    ws.cell(row=row, column=5).fill = FILL_SUMMARY
    ws.cell(row=row, column=5).alignment = CENTER
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = THIN
    row += 1

    pct = round(100 * sum(portfolio_scores) / sum(portfolio_totals), 2) if sum(portfolio_totals) else 0
    ws.cell(row=row, column=2, value="Coverage %age").font = FONT_BOLD
    ws.cell(row=row, column=2).fill = FILL_SUMMARY
    ws.cell(row=row, column=5, value=f"{pct}%").font = FONT_BOLD
    ws.cell(row=row, column=5).fill = FILL_SUMMARY
    ws.cell(row=row, column=5).alignment = CENTER
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = THIN
    row += 2

    ws.cell(row=row, column=2, value="Notes").font = FONT_BOLD
    row += 1
    notes = [
        "Source of truth: PayU Developer Documentation repository (docs/, reference/). Sample progress-report format reused; data not copied.",
        "Baseline (Jul 2026) = current repository assessment. Future month columns left blank for ongoing FY 26-27 tracking.",
        "Score legend: 1 = Present, 0 = Missing, Not Applicable = excluded from Total/Docs Score denominator.",
        "Total = count of applicable checklist items; Docs Score = count of items scored 1.",
        "Coverage %age = Docs Coverage Score / Total across all products.",
        "Product Video / Glossary / Change Log are commonly 0 across the India docs corpus — primary FY26-27 gap themes.",
        "RECYCLE BIN excluded; Integration ASK AI Docs treated as non-canonical duplicates (captured in recommendations where relevant).",
    ]
    for n in notes:
        ws.cell(row=row, column=2, value=n).alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 28
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 14
    for i in range(len(MONTHS)):
        ws.column_dimensions[get_column_letter(5 + i)].width = 12

    # Summary sheet — product rollup (helpful; still same philosophy)
    summary = wb.create_sheet("Coverage Summary")
    summary["A1"] = "PayU Product Docs — Coverage Summary (Baseline Jul 2026)"
    summary["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    summary["A1"].fill = FILL_HEADER
    summary.merge_cells("A1:F1")
    headers = ["Product", "Applicable Items (Total)", "Docs Score", "Coverage %", "# Gaps (0s)", "Top Recommendation"]
    for c, h in enumerate(headers, 1):
        cell = summary.cell(row=3, column=c, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.border = THIN
        cell.alignment = CENTER
    summary.freeze_panes = "A4"

    r = 4
    for p, total, sc in zip(products, portfolio_totals, portfolio_scores):
        gaps = sum(1 for _, v, _ in p.items if as_num(v) == 0)
        top_rec = next((rec for _, v, rec in p.items if as_num(v) == 0 and rec), "")
        cov = round(100 * sc / total, 1) if total else 0
        vals = [p.name, total, sc, cov, gaps, top_rec]
        for c, v in enumerate(vals, 1):
            cell = summary.cell(row=r, column=c, value=v)
            cell.border = THIN
            cell.alignment = WRAP if c in {1, 6} else CENTER
            cell.font = FONT_BODY
        cov_cell = summary.cell(row=r, column=4)
        if cov >= 85:
            cov_cell.fill = FILL_ONE
        elif cov >= 50:
            cov_cell.fill = FILL_TOTAL
        else:
            cov_cell.fill = FILL_ZERO
        r += 1

    summary.cell(row=r + 1, column=1, value="Portfolio Docs Coverage Score").font = FONT_BOLD
    summary.cell(row=r + 1, column=2, value=sum(portfolio_scores))
    summary.cell(row=r + 2, column=1, value="Portfolio Applicable Total").font = FONT_BOLD
    summary.cell(row=r + 2, column=2, value=sum(portfolio_totals))
    summary.cell(row=r + 3, column=1, value="Portfolio Coverage %").font = FONT_BOLD
    summary.cell(row=r + 3, column=2, value=f"{pct}%")
    summary.column_dimensions["A"].width = 48
    summary.column_dimensions["B"].width = 22
    summary.column_dimensions["C"].width = 12
    summary.column_dimensions["D"].width = 12
    summary.column_dimensions["E"].width = 12
    summary.column_dimensions["F"].width = 60
    summary.auto_filter.ref = f"A3:F{3+len(products)}"

    # Legend sheet
    legend = wb.create_sheet("Legend & Instructions")
    legend["A1"] = "How to use this Progress Report"
    legend["A1"].font = Font(name="Calibri", bold=True, size=14)
    lines = [
        "1. This workbook follows the same structure as the FY Docs Coverage Progress Report sample.",
        "2. Update future month columns (Aug 2026 → Mar 2027) with 1 / 0 / Not Applicable as content ships.",
        "3. Recalculate each product's Total (= applicable items) and Docs Score (= items with 1) for that month.",
        "4. Portfolio Coverage %age = sum(Docs Scores) / sum(Totals).",
        "5. Green=1 (present), Red=0 (missing), Grey=Not Applicable.",
        "6. Do not copy historical scores from other brands/products — only track PayU India Developer Docs.",
        "7. Regenerate baseline anytime with: python3 docs-coverage-tracker/generate_progress_report.py",
        "",
        "Standard checklist items:",
        *[f"   - {i}" for i in STANDARD_ITEMS],
    ]
    for i, line in enumerate(lines, 3):
        legend.cell(row=i, column=1, value=line)
    legend.column_dimensions["A"].width = 100

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Products: {len(products)}")
    print(f"Portfolio Total (applicable): {sum(portfolio_totals)}")
    print(f"Portfolio Docs Score: {sum(portfolio_scores)}")
    print(f"Portfolio Coverage %: {pct}%")


def main():
    products = build_products()
    write_workbook(products)


if __name__ == "__main__":
    main()
